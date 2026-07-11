from django.shortcuts import render, redirect, get_object_or_404
from django.forms import modelformset_factory
from django.http import HttpResponse, HttpResponseRedirect
from django.template.loader import render_to_string
from django.db.models import Q
try:
    from xhtml2pdf import pisa
except ImportError:
    pisa = None
from django.contrib import messages
from django.urls import reverse
import io
import re
from urllib.parse import quote
from .models import Exam, ExamGrade, ExamType
from classroom.models import Classroom
from courses.models import Subject
from students.models import Student
from accounts.models import Studentenrollment
from .forms import ExamForm, ExamGradesForm, ExamGradesFormSet
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from django.utils import timezone


def _ensure_exam_grades_for_classroom(exam):
    """
    Ensure every active student in the exam's classroom has an ExamGrade row.
    This helps keep pages like /exams/exam/<id>/grades/ populated even
    if students were added after the exam was created.
    """
    classroom_students = exam.classroom.students.all()
    for student in classroom_students:
        ExamGrade.objects.get_or_create(
            exam=exam,
            student=student,
            defaults={'grade': None, 'notes': '', 'status': 'present'}
        )

def exams_dashboard(request):
    """لوحة التحكم الرئيسية للاختبارات"""
    classrooms = Classroom.objects.filter(is_active=True)
    if not request.user.is_superuser:
        classrooms = classrooms.filter(is_visible=True)
        
    current_year = getattr(request, 'current_academic_year', None)
    if current_year:
        classrooms = classrooms.filter(
            Q(course__academic_year=current_year) |
            Q(enrollments__student__academic_year=current_year)
        ).distinct()
        
    classrooms = classrooms.order_by('name')
    return render(request, 'exams/dashboard.html', {'classrooms': classrooms})

def exam_list(request, classroom_id):
    """قائمة الاختبارات للشعبة"""
    classroom = get_object_or_404(Classroom, pk=classroom_id)
    exams_list = Exam.objects.filter(classroom=classroom).order_by('-exam_date')
    
    return render(request, 'exams/exam_list.html', {
        'classroom': classroom,
        'exams': exams_list
    })

def create_exam(request, classroom_id):
    """إنشاء اختبار جديد"""
    classroom = get_object_or_404(Classroom, pk=classroom_id)
    
    if request.method == 'POST':
        form = ExamForm(request.POST)
        if form.is_valid():
            exam = form.save(commit=False)
            exam.classroom = classroom
            exam.save()
            
            # إنشاء سجلات العلامات للطلاب النشطين والمسجلين بدورات فعّالة
            students = Student.objects.filter(
                classroom_enrollments__classroom=classroom,
                is_active=True
            ).distinct()
            for student in students:
                ExamGrade.objects.create(
                    exam=exam, 
                    student=student,
                    status='present',
                    grade=None,
                    notes=''
                )
            
            messages.success(request, 'تم إنشاء الاختبار بنجاح')
            return redirect('exams:exam_list', classroom_id=classroom_id)
        else:
            messages.error(request, 'حدث خطأ في إنشاء الاختبار')
    else:
        form = ExamForm()
    
    return render(request, 'exams/create_exam.html', {
        'classroom': classroom,
        'form': form,
        'exam': None,
        'page_title': 'إنشاء اختبار جديد',
        'submit_label': 'حفظ الاختبار'
    })


def edit_exam(request, exam_id):
    """تعديل بيانات اختبار"""
    exam = get_object_or_404(Exam, pk=exam_id)
    classroom = exam.classroom

    if request.method == 'POST':
        form = ExamForm(request.POST, instance=exam)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تعديل بيانات الاختبار بنجاح')
            return redirect('exams:exam_detail', exam_id=exam.id)
        else:
            messages.error(request, 'حدث خطأ أثناء تعديل بيانات الاختبار')
    else:
        form = ExamForm(instance=exam)

    return render(request, 'exams/create_exam.html', {
        'classroom': classroom,
        'form': form,
        'exam': exam,
        'page_title': 'تعديل الاختبار',
        'submit_label': 'تحديث الاختبار'
    })

def exam_detail(request, exam_id):
    """تفاصيل الاختبار"""
    exam = get_object_or_404(Exam, pk=exam_id)
    _ensure_exam_grades_for_classroom(exam)
    exam_grades = ExamGrade.objects.filter(
        exam=exam
    ).distinct()
    
    # حساب الإحصائيات
    total_students = exam_grades.count()
    entered_grades = exam_grades.filter(status='present', grade__isnull=False).count()
    absent_count = exam_grades.filter(status='absent').count()
    not_submitted_count = exam_grades.filter(status='not_submitted').count()
    missing_grades = exam_grades.filter(status='present', grade__isnull=True).count()
    percentage_entered = (entered_grades / total_students * 100) if total_students > 0 else 0
    
    return render(request, 'exams/exam_detail.html', {
        'exam': exam,
        'total_students': total_students,
        'entered_grades': entered_grades,
        'absent_count': absent_count,
        'not_submitted_count': not_submitted_count,
        'missing_grades': missing_grades,
        'percentage_entered': round(percentage_entered, 1)
    })

def view_exam_grades(request, exam_id):
    """عرض علامات اختبار"""
    exam = get_object_or_404(Exam, pk=exam_id)
    _ensure_exam_grades_for_classroom(exam)
    exam_grades = ExamGrade.objects.filter(
        exam=exam
    ).select_related('student').order_by('student__full_name').distinct()
    
    # حساب الإحصائيات
    total_students = exam_grades.count()
    entered_grades = exam_grades.filter(status='present', grade__isnull=False).count()
    absent_count = exam_grades.filter(status='absent').count()
    not_submitted_count = exam_grades.filter(status='not_submitted').count()
    missing_grades = exam_grades.filter(status='present', grade__isnull=True).count()
    percentage_entered = (entered_grades / total_students * 100) if total_students > 0 else 0
    
    return render(request, 'exams/view_exam_grades.html', {
        'exam': exam,
        'exam_grades': exam_grades,
        'total_students': total_students,
        'entered_grades': entered_grades,
        'absent_count': absent_count,
        'not_submitted_count': not_submitted_count,
        'missing_grades': missing_grades,
        'percentage_entered': round(percentage_entered, 1)
    })
def edit_exam_grades(request, exam_id):
    """تعديل علامات اختبار"""
    exam = get_object_or_404(Exam, pk=exam_id)
    _ensure_exam_grades_for_classroom(exam)
    
    if request.method == 'POST':
        formset = ExamGradesFormSet(
            request.POST,
            queryset=ExamGrade.objects.filter(
                exam=exam
            ).distinct()
        )
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                # التحقق من أن العلامة لا تتجاوز الحد الأقصى
                if instance.grade and instance.grade > exam.max_grade:
                    messages.warning(request, f'العلامة للطالب {instance.student.full_name} تتجاوز الحد الأقصى ({exam.max_grade})')
                instance.save()
            
            formset.save_m2m()
            messages.success(request, 'تم حفظ العلامات بنجاح')
            return redirect('exams:view_exam_grades', exam_id=exam.id)
        else:
            messages.error(request, 'حدث خطأ في حفظ العلامات')
    else:
        formset = ExamGradesFormSet(
            queryset=ExamGrade.objects.filter(
                exam=exam
            ).order_by('student__full_name').distinct()
        )
    
    return render(request, 'exams/edit_exam_grades.html', {
        'exam': exam,
        'formset': formset
    })

def export_exam_grades(request, exam_id):
    """تصدير علامات الاختبار إلى إكسل"""
    exam = get_object_or_404(Exam, pk=exam_id)
    _ensure_exam_grades_for_classroom(exam)
    exam_grades = ExamGrade.objects.filter(
        exam=exam
    ).select_related('student').order_by('student__full_name').distinct()
    
    # إنشاء ملف إكسل
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "علامات الاختبار"

    # Styles
    title_font = Font(size=16, bold=True)
    title_fill = PatternFill(fill_type="solid", fgColor="F5F7FA")
    info_font = Font(size=10, bold=True)
    info_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    text_alignment = Alignment(horizontal="right", vertical="center")
    wrap_alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # كتابة العنوان
    ws['A1'] = f"علامات اختبار {exam.name}"
    ws.merge_cells('A1:D1')
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28
    
    # معلومات الاختبار
    ws['A2'] = f"الشعبة: {exam.classroom.name}"
    ws['B2'] = f"المادة: {exam.subject.name}"
    ws['C2'] = f"تاريخ الاختبار: {exam.exam_date}"
    ws['D2'] = f"العلامة القصوى: {exam.max_grade}"
    for col in range(1, 5):
        cell = ws.cell(row=2, column=col)
        cell.font = info_font
        cell.fill = info_fill
        cell.alignment = text_alignment
        cell.border = border
    
    # كتابة العناوين
    headers = ['م', 'اسم الطالب', 'العلامة', 'النسبة المئوية', 'ملاحظات']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    ws.row_dimensions[4].height = 22

    # كتابة البيانات
    row = 5
    for i, exam_grade in enumerate(exam_grades, 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=exam_grade.student.full_name)
        
        # كتابة العلامة (بما فيها 0)
        if exam_grade.grade is not None:
            # التحقق من القيمة 0
            if float(exam_grade.grade) == 0:
                ws.cell(row=row, column=3, value=0)
            else:
                ws.cell(row=row, column=3, value=float(exam_grade.grade))
        else:
            ws.cell(row=row, column=3, value='-')
        
        # حساب النسبة المئوية
        if exam_grade.grade is not None:
            grade_value = float(exam_grade.grade)
            percentage = (grade_value / float(exam.max_grade)) * 100
            percent_cell = ws.cell(row=row, column=4, value=(percentage / 100))
            percent_cell.number_format = "0.0%"
        else:
            ws.cell(row=row, column=4, value='')
        
        ws.cell(row=row, column=5, value=exam_grade.notes if exam_grade.notes else '')

        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col in (1, 3, 4):
                cell.alignment = cell_alignment
            elif col == 5:
                cell.alignment = wrap_alignment
            else:
                cell.alignment = text_alignment

        if row % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
        row += 1
    
    # ضبط عرض الأعمدة
    # Auto-fit column widths based on content
    max_widths = [0, 0, 0, 0, 0]
    for row_cells in ws.iter_rows(min_row=1, max_row=row - 1, max_col=5):
        for idx, cell in enumerate(row_cells):
            value = "" if cell.value is None else str(cell.value)
            text_width = int(len(value) * 1.2) + 2
            max_widths[idx] = max(max_widths[idx], text_width)
    for i, width in enumerate(max_widths, 1):
        width = max(8, min(width, 60))
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    

    # Sheet layout
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{row - 1}"

    # Conditional formatting for grade column
    if row > 5:
        grade_range = f"C5:C{row - 1}"
        ws.conditional_formatting.add(
            grade_range,
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

    # إعداد الاستجابة
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    raw_name = f"{exam.name}_{exam.classroom.name}"
    safe_base = re.sub(r'[\\/:*?"<>|]', "_", raw_name).strip()
    safe_name = f"{safe_base}.xlsx"
    ascii_base = re.sub(r'[^A-Za-z0-9._-]+', "_", safe_base).strip("_") or "exam"
    fallback_name = f"{ascii_base}.xlsx"
    response['Content-Disposition'] = (
        "attachment; "
        f"filename=\"{fallback_name}\"; "
        f"filename*=UTF-8''{quote(safe_name)}"
    )
    
    wb.save(response)
    return response

def print_exam_grades(request, exam_id):
    """طباعة علامات الاختبار"""
    exam = get_object_or_404(Exam, pk=exam_id)
    _ensure_exam_grades_for_classroom(exam)
    exam_grades = ExamGrade.objects.filter(
        exam=exam
    ).select_related('student').order_by('student__full_name').distinct()
    
    html_string = render_to_string('exams/print_exam_grades.html', {
        'exam': exam,
        'exam_grades': exam_grades,
        'max_grade': exam.max_grade,
        'print_date': timezone.now().date()
    })
    
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html_string.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename=exam_grades_{exam.name}.pdf'
        return response
    
    return HttpResponse('Error generating PDF: %s' % pdf.err)

def exam_stats(request, exam_id):
    """إحصائيات الاختبار"""
    exam = get_object_or_404(Exam, pk=exam_id)
    _ensure_exam_grades_for_classroom(exam)
    exam_grades = ExamGrade.objects.filter(
        exam=exam
    ).exclude(grade__isnull=True).distinct()
    
    # حساب الإحصائيات
    total_students = ExamGrade.objects.filter(
        exam=exam
    ).distinct().count()
    entered_grades = exam_grades.count()
    
    if entered_grades > 0:
        grades_list = [float(eg.grade) for eg in exam_grades]
        average = sum(grades_list) / entered_grades
        highest = max(grades_list)
        lowest = min(grades_list)
        
        # توزيع العلامات بناءً على النسبة المئوية
        max_grade = float(exam.max_grade)
        distribution = {
            'ممتاز (90-100%)': len([g for g in grades_list if (g/max_grade)*100 >= 90]),
            'جيد جداً (80-89%)': len([g for g in grades_list if 80 <= (g/max_grade)*100 < 90]),
            'جيد (70-79%)': len([g for g in grades_list if 70 <= (g/max_grade)*100 < 80]),
            'مقبول (60-69%)': len([g for g in grades_list if 60 <= (g/max_grade)*100 < 70]),
            'راسب (أقل من 60%)': len([g for g in grades_list if (g/max_grade)*100 < 60]),
        }
    else:
        average = highest = lowest = 0
        distribution = {}
    
    return render(request, 'exams/exam_stats.html', {
        'exam': exam,
        'total_students': total_students,
        'entered_grades': entered_grades,
        'average': round(average, 2),
        'highest': highest,
        'lowest': lowest,
        'distribution': distribution,
        'max_grade': exam.max_grade
    })

def delete_exam(request, exam_id):
    """حذف اختبار"""
    exam = get_object_or_404(Exam, pk=exam_id)
    classroom_id = exam.classroom.id
    
    if request.method == 'POST':
        exam_name = exam.name
        exam.delete()
        messages.success(request, f'تم حذف الاختبار "{exam_name}" بنجاح')
        return redirect('exams:exam_list', classroom_id=classroom_id)
    
    return render(request, 'exams/delete_exam_confirm.html', {
        'exam': exam
    })


def missed_exams_report(request):
    """تقرير المتخلفين عن الاختبارات"""
    classrooms = Classroom.objects.filter(is_active=True)
    if not request.user.is_superuser:
        classrooms = classrooms.filter(is_visible=True)
        
    current_year = getattr(request, 'current_academic_year', None)
    if current_year:
        classrooms = classrooms.filter(
            Q(course__academic_year=current_year) |
            Q(enrollments__student__academic_year=current_year)
        ).distinct()
        
    classrooms = classrooms.order_by('name')
    subjects = Subject.objects.all().order_by('name')
    exam_types = ExamType.objects.all().order_by('name')
    
    # Filters from GET
    classroom_id = request.GET.get('classroom')
    subject_id = request.GET.get('subject')
    exam_type_id = request.GET.get('exam_type')
    status_filter = request.GET.get('status')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Query missed grades
    grades_qs = ExamGrade.objects.select_related('student', 'exam', 'exam__classroom', 'exam__subject', 'exam__exam_type')
    
    # Filter by classrooms active/visible
    grades_qs = grades_qs.filter(exam__classroom__in=classrooms)
    
    if classroom_id:
        grades_qs = grades_qs.filter(exam__classroom_id=classroom_id)
    if subject_id:
        grades_qs = grades_qs.filter(exam__subject_id=subject_id)
    if exam_type_id:
        grades_qs = grades_qs.filter(exam__exam_type_id=exam_type_id)
        
    if status_filter == 'absent':
        grades_qs = grades_qs.filter(status='absent')
    elif status_filter == 'not_submitted':
        grades_qs = grades_qs.filter(status='not_submitted')
    elif status_filter == 'incomplete':
        grades_qs = grades_qs.filter(status='present', grade__isnull=True)
    else:  # 'all_missed' or default
        grades_qs = grades_qs.filter(Q(status__in=['absent', 'not_submitted']) | Q(grade__isnull=True))
        
    if start_date:
        grades_qs = grades_qs.filter(exam__exam_date__gte=start_date)
    if end_date:
        grades_qs = grades_qs.filter(exam__exam_date__lte=end_date)
        
    grades_qs = grades_qs.order_by('-exam__exam_date', 'student__full_name')
    
    report_data = []
    for grade in grades_qs:
        student = grade.student
        exam = grade.exam
        
        status_label = "ناقصة (لم ترصد بعد)"
        if grade.status == 'absent':
            status_label = "غائب"
        elif grade.status == 'not_submitted':
            status_label = "غير مقدم"
            
        phones = {
            'student': student.phone or '',
            'father': student.father_phone or '',
            'mother': student.mother_phone or '',
        }
        
        # Build text templates
        warning_msg = f"السلام عليكم، نود إعلامكم بأن الطالب {student.full_name} قد تخلف ({status_label}) عن تقديم اختبار {exam.name} لمادة {exam.subject.name} بتاريخ {exam.exam_date}. يرجى الالتزام والمتابعة."
        parent_msg = f"السلام عليكم، يرجى من ولي أمر الطالب {student.full_name} مراجعة إدارة المعهد بخصوص تخلف الطالب ({status_label}) عن تقديم اختبار {exam.name} لمادة {exam.subject.name} بتاريخ {exam.exam_date}."
        
        wa_urls = {}
        for role, phone in phones.items():
            if phone:
                clean_phone = phone.strip().replace('+', '').replace(' ', '')
                if not clean_phone.startswith('963') and not clean_phone.startswith('00963'):
                    if clean_phone.startswith('0'):
                        clean_phone = '963' + clean_phone[1:]
                    else:
                        clean_phone = '963' + clean_phone
                
                wa_urls[f"{role}_warning"] = f"https://wa.me/{clean_phone}?text={quote(warning_msg)}"
                wa_urls[f"{role}_parent"] = f"https://wa.me/{clean_phone}?text={quote(parent_msg)}"
            else:
                wa_urls[f"{role}_warning"] = ""
                wa_urls[f"{role}_parent"] = ""
                
        report_data.append({
            'grade': grade,
            'student': student,
            'exam': exam,
            'status_label': status_label,
            'phones': phones,
            'wa_urls': wa_urls
        })
        
    context = {
        'classrooms': classrooms,
        'subjects': subjects,
        'exam_types': exam_types,
        'report_data': report_data,
        'filters': {
            'classroom': classroom_id or '',
            'subject': subject_id or '',
            'exam_type': exam_type_id or '',
            'status': status_filter or 'all_missed',
            'start_date': start_date or '',
            'end_date': end_date or '',
        }
    }
    return render(request, 'exams/missed_exams.html', context)
