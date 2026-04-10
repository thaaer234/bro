from django import forms 
from django.views.generic import ListView, CreateView, DeleteView, UpdateView
from django.views.generic.edit import FormView
from django.urls import reverse, reverse_lazy
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Value, DecimalField, Count, Max
from django.db.models.functions import Coalesce
from django.db import transaction
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import UserPassesTestMixin, LoginRequiredMixin
from django.contrib.auth.decorators import login_required  # ← أضف هذا السطر
from attendance.models import Attendance
from classroom.models import Classroomenrollment, Classroom
from django.http import JsonResponse, Http404, HttpResponse
from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import View, TemplateView, ListView, DetailView
# from .models import QuickStudent, QuickEnrollment, QuickCourse, AcademicYear
from django.contrib import messages
from django.utils.dateparse import parse_date
from .forms import (
    AcademicYearForm,
    QuickCourseForm,
    QuickClassroomForm,
    QuickCourseTimeOptionForm,
    QuickCourseSessionForm,
    QuickSessionAssignStudentsForm,
    QuickSessionAttendanceBulkForm,
    QuickSessionTransferForm,
    QuickStudentForm,
    QuickEnrollmentForm,
    _normalize_quick_name,
    _normalize_quick_phone,
)
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from itertools import combinations
from math import ceil
import logging
import time
from django.views.decorators.http import require_POST
from django.views.decorators.http import require_GET
from django.views.decorators.cache import never_cache
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from urllib.parse import urlencode
from django.db.models import Prefetch
from django.conf import settings
from django.db import transaction, connection, OperationalError, close_old_connections
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from accounts.models import Transaction, JournalEntry, Account, get_user_cash_account
from .models import (
    QuickCourse,
    QuickCourseTimeOption,
    QuickCourseSession,
    QuickCourseSessionAttendance,
    QuickCourseSessionEnrollment,
    QuickEnrollment,
    QuickManualSortingSelection,
    QuickReceiptPrintJob,
    QuickStudent,
    QuickStudentReceipt,
    AcademicYear,
)
from accounts.models import Course, CostCenter
from .services.receipt_printer import QuickReceiptPrinterError, print_many_receipts
from employ.decorators import require_superuser
User = get_user_model()
logger = logging.getLogger('django')


def _parse_quick_posted_int(raw_value):
    if raw_value in (None, ''):
        raise ValueError('empty value')
    normalized = str(raw_value).strip()
    for token in (' ', ',', '.', '٬', '،'):
        normalized = normalized.replace(token, '')
    return int(normalized)


def _get_employee_cash_account(user):
    """Return the current user's cash account or raise."""
    if not user or not getattr(user, 'is_authenticated', False):
        raise ValueError('user must be authenticated to fetch cash account')
    cash_account = get_user_cash_account(user, fallback_code='121')
    if not cash_account:
        raise ValueError('Cash account missing for the current user')
    return cash_account


def _process_quick_refund(student, enrollment, refund_amount, refund_reason, user):
    """Apply a refund for a quick student enrollment and create the journal entry."""
    if refund_amount <= 0:
        raise ValueError('المبلغ المسترد يجب أن يكون أكبر من الصفر')

    receipts_data = _adjust_quick_receipts_for_refund(student, enrollment, refund_amount)
    actual_refund = receipts_data['refunded_amount']

    if actual_refund <= 0:
        raise ValueError('لا يوجد مبالغ مدفوعة كافية ليتم استردادها')

    cash_account = _get_employee_cash_account(user)
    description = f"استرداد مبلغ - {student.full_name} - {enrollment.course.name}"
    if refund_reason:
        description += f" - {refund_reason}"

    refund_entry = JournalEntry.objects.create(
        date=timezone.now().date(),
        description=description,
        entry_type='ADJUSTMENT',
        total_amount=actual_refund,
        created_by=user
    )

    Transaction.objects.create(
        journal_entry=refund_entry,
        account=student.ar_account,
        amount=actual_refund,
        is_debit=True,
        description=f"استرداد مبلغ - {enrollment.course.name}"
    )

    Transaction.objects.create(
        journal_entry=refund_entry,
        account=cash_account,
        amount=actual_refund,
        is_debit=False,
        description=f"استرداد نقدي - {student.full_name}"
    )

    refund_entry.post_entry(user)

    net_amount = enrollment.net_amount or Decimal('0.00')
    previous_balance = max(Decimal('0.00'), net_amount - receipts_data['previous_paid'])
    new_balance = max(Decimal('0.00'), net_amount - receipts_data['new_total_paid'])

    return {
        'refund_entry': refund_entry,
        'refund_amount': actual_refund,
        'previous_paid': receipts_data['previous_paid'],
        'new_total_paid': receipts_data['new_total_paid'],
        'previous_balance': previous_balance,
        'new_balance': new_balance
    }


def _adjust_quick_receipts_for_refund(student, enrollment, refund_amount):
    """Decrease QuickStudentReceipt paid amounts to match the refund target."""
    refund_amount = max(Decimal('0'), refund_amount or Decimal('0'))
    receipts_qs = QuickStudentReceipt.objects.filter(
        quick_student=student,
        quick_enrollment=enrollment,
        course=enrollment.course
    ).order_by('-date', '-id')

    total_paid = receipts_qs.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    refundable = min(refund_amount, total_paid)

    remaining = refundable
    for receipt in receipts_qs:
        if remaining <= 0:
            break

        available = min(remaining, receipt.paid_amount)
        if available <= 0:
            continue

        receipt.paid_amount -= available
        receipt.save(update_fields=['paid_amount'])
        remaining -= available

    return {
        'previous_paid': total_paid,
        'new_total_paid': total_paid - refundable,
        'refunded_amount': refundable
    }


def _is_quick_legacy_withdrawal_account(account):
    if not account:
        return False
    code = str(getattr(account, 'code', '') or '')
    if code == '4201' or code.startswith('4201-'):
        return True
    account_text = ' '.join([
        code,
        str(getattr(account, 'name', '') or ''),
        str(getattr(account, 'name_ar', '') or ''),
    ]).casefold()
    return (
        account.account_type == 'REVENUE' and (
            'withdrawal' in account_text or
            'انسحاب' in account_text
        )
    )


def _get_legacy_withdrawal_amount(entry):
    amount = Decimal('0')
    for tx in entry.transactions.select_related('account').all():
        if tx.is_debit and _is_quick_legacy_withdrawal_account(tx.account):
            amount += tx.amount or Decimal('0')
    return amount


def _deactivate_quick_legacy_withdrawal_accounts():
    deactivated = []
    legacy_accounts = Account.objects.filter(
        Q(code='4201') |
        Q(code__startswith='4201-') |
        (
            Q(account_type='REVENUE') &
            (
                Q(name__icontains='Withdrawal') |
                Q(name_ar__icontains='انسحاب')
            )
        )
    ).order_by('-code')

    for account in legacy_accounts:
        try:
            live_balance = account.get_net_balance() or Decimal('0')
        except Exception:
            live_balance = account.balance or Decimal('0')

        if abs(live_balance) >= Decimal('0.01'):
            continue
        if not account.is_active:
            continue

        account.is_active = False
        account.save(update_fields=['is_active'])
        deactivated.append(account.code)

    return deactivated


def _find_quick_withdrawal_entries(enrollment):
    student = getattr(enrollment, 'student', None)
    course = getattr(enrollment, 'course', None)
    if not student or not course:
        return JournalEntry.objects.none()

    deferred_account = Account.get_or_create_quick_course_deferred_account(course)
    return JournalEntry.objects.filter(
        Q(description__icontains=student.full_name) &
        Q(description__icontains=course.name) &
        (
            (
                Q(transactions__account=deferred_account) &
                (
                    Q(transactions__description__icontains='عكس') |
                    Q(transactions__description__icontains='المدينة')
                )
            ) |
            (
                Q(transactions__account__code='4201') |
                Q(transactions__account__code__startswith='4201-') |
                (
                    Q(transactions__account__account_type='REVENUE') &
                    (
                        Q(transactions__account__name__icontains='Withdrawal') |
                        Q(transactions__account__name_ar__icontains='انسحاب')
                    )
                )
            )
        )
    ).distinct()


def _find_quick_generated_withdraw_fix_entries(enrollment):
    return JournalEntry.objects.filter(
        Q(description__icontains=f'[QUICK_WITHDRAW #{enrollment.id}]') |
        Q(description__icontains=f'[QUICK_WITHDRAW_FIX #{enrollment.id}]') |
        Q(description__icontains=f'[QUICK_WITHDRAW_CLEANUP #{enrollment.id}]')
    ).distinct()


def _cleanup_quick_withdrawal_entries(enrollment, user):
    cleaned = {
        'reversed_ids': [],
        'deleted_ids': [],
    }
    candidate_entries = {}
    for entry in _find_quick_withdrawal_entries(enrollment):
        candidate_entries[entry.id] = entry
    for entry in _find_quick_generated_withdraw_fix_entries(enrollment):
        candidate_entries[entry.id] = entry

    for entry in candidate_entries.values():
        entry_id = entry.id
        entry.delete()
        cleaned['deleted_ids'].append(entry_id)

    return cleaned


def _ensure_quick_enrollment_entry(enrollment, user=None):
    entry = JournalEntry.objects.filter(reference=f'QE-{enrollment.id}').first()
    if entry:
        return entry

    if user is not None:
        entry = enrollment.create_accrual_enrollment_entry(user)
        return _normalize_quick_enrollment_entry_arabic(entry, enrollment)

    return None


def _find_quick_enrollment_entry(enrollment):
    return JournalEntry.objects.filter(reference=f'QE-{enrollment.id}').first()


def _normalize_quick_enrollment_entry_arabic(entry, enrollment):
    if not entry or not enrollment:
        return entry

    entry.description = f"تسجيل سريع - {enrollment.student.full_name} في {enrollment.course.name}"
    entry.save(update_fields=['description'])

    for tx in entry.transactions.all():
        if tx.is_debit:
            tx.description = f"تسجيل سريع - {enrollment.student.full_name}"
        else:
            tx.description = f"إيرادات مؤجلة - {enrollment.course.name}"
        tx.save(update_fields=['description'])

    return entry


def _normalize_quick_receipt_entry_arabic(entry, receipt):
    if not entry or not receipt:
        return entry

    entry.description = f"إيصال سريع - {receipt.student_name} - {receipt.course_name}"
    entry.save(update_fields=['description'])

    for tx in entry.transactions.all():
        if tx.is_debit:
            tx.description = f"إيصال سريع - {receipt.student_name}"
        else:
            tx.description = f"تسديد ذمم - {receipt.course_name}"
        tx.save(update_fields=['description'])

    return entry


def _normalize_quick_withdrawal_entry_arabic(entry, enrollment, receipt=None):
    if not entry or not enrollment:
        return entry

    student_name = enrollment.student.full_name
    course_name = enrollment.course.name

    if receipt is None:
        entry.description = (
            f"[QUICK_WITHDRAW #{enrollment.id}] "
            f"عكس قيد تسجيل عند السحب - {student_name} - {course_name}"
        )
        entry.save(update_fields=['description'])

        for tx in entry.transactions.all():
            if tx.is_debit:
                tx.description = f"عكس تسجيل سريع - {course_name}"
            else:
                tx.description = f"عكس تسجيل سريع - {student_name}"
            tx.save(update_fields=['description'])
        return entry

    receipt_label = receipt.receipt_number or str(receipt.id)
    entry.description = (
        f"[QUICK_WITHDRAW #{enrollment.id}] "
        f"عكس قيد قبض عند السحب - {student_name} - {course_name} - إيصال {receipt_label}"
    )
    entry.save(update_fields=['description'])

    for tx in entry.transactions.all():
        if tx.is_debit:
            tx.description = f"عكس قبض سريع - {course_name}"
        else:
            tx.description = f"رد قبض سريع - {student_name}"
        tx.save(update_fields=['description'])

    return entry


def _make_journal_reference_available(entry, suffix='OLD'):
    if not entry or not entry.reference:
        return

    base_reference = entry.reference
    max_length = JournalEntry._meta.get_field('reference').max_length
    counter = 1
    while True:
        extra = f'-{suffix}{counter}'
        candidate = f'{base_reference[:max_length - len(extra)]}{extra}'
        if not JournalEntry.objects.exclude(id=entry.id).filter(reference=candidate).exists():
            entry.reference = candidate
            entry.save(update_fields=['reference'])
            return
        counter += 1


def _free_journal_reference(reference, exclude_id=None):
    reference = str(reference or '').strip()
    if not reference:
        return
    queryset = JournalEntry.objects.filter(reference=reference)
    if exclude_id:
        queryset = queryset.exclude(id=exclude_id)
    conflict_entry = queryset.first()
    if conflict_entry:
        _make_journal_reference_available(conflict_entry, suffix='REPLACED')


def _rebuild_quick_enrollment_entry(enrollment, user):
    entry = _find_quick_enrollment_entry(enrollment)
    if entry:
        if entry.is_posted:
            entry.reverse_entry(
                user,
                description=(
                    f"[QUICK_REG_CLEANUP #{enrollment.id}] "
                    f"إلغاء قيد تسجيل خاطئ - {enrollment.student.full_name} - {enrollment.course.name}"
                ),
            )
            _make_journal_reference_available(entry, suffix='REPLACED')
        else:
            entry.delete()
    entry = enrollment.create_accrual_enrollment_entry(user)
    return _normalize_quick_enrollment_entry_arabic(entry, enrollment)


def _fix_quick_receipt_entry(receipt, student_ar_account):
    if not receipt or not receipt.journal_entry_id:
        return False

    entry = receipt.journal_entry
    if not entry:
        return False

    credit_transactions = list(entry.transactions.filter(is_debit=False).select_related('account'))
    debit_transactions = list(entry.transactions.filter(is_debit=True).select_related('account'))

    if len(credit_transactions) != 1 or len(debit_transactions) != 1:
        return False

    credit_tx = credit_transactions[0]
    debit_tx = debit_transactions[0]
    expected_amount = receipt.paid_amount or Decimal('0')

    if debit_tx.amount != expected_amount or credit_tx.amount != expected_amount:
        return False

    if credit_tx.account_id == student_ar_account.id:
        return True

    credit_tx.account = student_ar_account
    credit_tx.save(update_fields=['account'])
    try:
        debit_tx.account.recalculate_tree_balances()
    except Exception:
        pass
    try:
        student_ar_account.recalculate_tree_balances()
    except Exception:
        pass
    return True


def _rebuild_quick_receipt_entry(receipt, user):
    if not receipt:
        return False

    entry = receipt.journal_entry
    if entry:
        if entry.is_posted:
            entry.reverse_entry(
                user,
                description=(
                    f"[QUICK_RECEIPT_CLEANUP #{receipt.id}] "
                    f"إلغاء قيد قبض خاطئ - {receipt.student_name} - {receipt.course_name}"
                ),
            )
            _make_journal_reference_available(entry, suffix='REPLACED')
        else:
            entry.delete()
        QuickStudentReceipt.objects.filter(id=receipt.id).update(journal_entry=None)
        receipt.journal_entry = None

    if (receipt.paid_amount or Decimal('0')) <= 0:
        return True

    _free_journal_reference(receipt.receipt_number, exclude_id=receipt.journal_entry_id)
    entry = receipt.create_accrual_journal_entry(user)
    _normalize_quick_receipt_entry_arabic(entry, receipt)
    return True


def _cap_quick_enrollment_receipts_to_net(enrollment, user):
    target_total = max(Decimal('0'), enrollment.net_amount or Decimal('0'))
    receipts = list(
        QuickStudentReceipt.objects.filter(
            quick_student=enrollment.student,
            quick_enrollment=enrollment,
        ).select_related('journal_entry').order_by('-date', '-id')
    )

    current_total = sum((receipt.paid_amount or Decimal('0')) for receipt in receipts)
    overflow = current_total - target_total
    if overflow <= 0:
        return 0

    fixed = 0
    for receipt in receipts:
        if overflow <= 0:
            break

        current_paid = receipt.paid_amount or Decimal('0')
        if current_paid <= 0:
            continue

        reduction = min(current_paid, overflow)
        new_paid = current_paid - reduction
        QuickStudentReceipt.objects.filter(id=receipt.id).update(paid_amount=new_paid)
        receipt.paid_amount = new_paid
        _rebuild_quick_receipt_entry(receipt, user)
        overflow -= reduction
        fixed += 1

    return fixed


def _build_quick_withdrawal_entry(enrollment, user, refunded_amount, description):
    created_entries = []

    enrollment_entry = _find_quick_enrollment_entry(enrollment)
    if enrollment_entry and enrollment_entry.is_posted:
        reversed_entry = enrollment_entry.reverse_entry(
            user,
            description=(
                f"[QUICK_WITHDRAW #{enrollment.id}] "
                f"عكس قيد تسجيل عند السحب - {enrollment.student.full_name} - {enrollment.course.name}"
            ),
        )
        created_entries.append(
            _normalize_quick_withdrawal_entry_arabic(
                reversed_entry,
                enrollment,
            )
        )

    receipts = list(
        QuickStudentReceipt.objects.filter(
            quick_student=enrollment.student,
            quick_enrollment=enrollment,
        ).select_related('journal_entry').order_by('date', 'id')
    )
    for receipt in receipts:
        entry = receipt.journal_entry
        if not entry or not entry.is_posted:
            continue
        reversed_entry = entry.reverse_entry(
            user,
            description=(
                f"[QUICK_WITHDRAW #{enrollment.id}] "
                f"عكس قيد قبض عند السحب - {enrollment.student.full_name} - {enrollment.course.name} - إيصال {receipt.receipt_number or receipt.id}"
            ),
        )
        created_entries.append(
            _normalize_quick_withdrawal_entry_arabic(
                reversed_entry,
                enrollment,
                receipt=receipt,
            )
        )

    return created_entries


def _get_quick_enrollment_paid_total(enrollment, student=None):
    """Return the paid total for one specific quick enrollment only."""
    if not enrollment:
        return Decimal('0')

    filters = {
        'quick_enrollment': enrollment,
    }
    if student is not None:
        filters['quick_student'] = student

    return QuickStudentReceipt.objects.filter(
        **filters
    ).aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')


def _normalize_phone(phone):
    if not phone:
        return ''
    digits = ''.join(ch for ch in str(phone) if ch.isdigit())
    return digits


def _build_regular_phone_set():
    from students.models import Student
    phones = set()
    regular_students = Student.objects.filter(quick_student_profile__isnull=True)
    for student in regular_students:
        for value in (student.phone, student.father_phone, student.mother_phone, student.home_phone):
            normalized = _normalize_phone(value)
            if normalized:
                phones.add(normalized)
    return phones


def _safe_sheet_title(title, existing_titles):
    base = (title or '').strip() or 'Sheet'
    if len(base) > 31:
        base = base[:31]
    name = base
    counter = 1
    while name in existing_titles:
        suffix = f"_{counter}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    return name


def _format_money(value):
    try:
        return Decimal(value)
    except Exception:
        return Decimal('0')


def _normalize_quick_student_name(name):
    value = ' '.join((name or '').split()).casefold()
    translation = str.maketrans({
        'أ': 'ا',
        'إ': 'ا',
        'آ': 'ا',
        'ٱ': 'ا',
        'ى': 'ي',
        'ئ': 'ي',
        'ؤ': 'و',
        'ة': 'ه',
        'ـ': '',
    })
    value = value.translate(translation)
    return ' '.join(value.split())


def _get_quick_enrollment_entry(enrollment):
    if not enrollment:
        return None
    return JournalEntry.objects.filter(reference=f'QE-{enrollment.id}').first()


def _retarget_journal_account(entry, old_account, new_account):
    if not entry or not old_account or not new_account or old_account == new_account:
        return False

    updated = False
    for tx in entry.transactions.select_related('account').all():
        if tx.account_id != old_account.id:
            continue
        tx.account = new_account
        tx.save(update_fields=['account'])
        updated = True

    if updated:
        old_account.recalculate_tree_balances()
        new_account.recalculate_tree_balances()
    return updated


def _retarget_all_account_transactions(old_account, new_account):
    if not old_account or not new_account or old_account == new_account:
        return 0

    updated = 0
    for tx in Transaction.objects.filter(account=old_account).select_related('journal_entry'):
        tx.account = new_account
        tx.save(update_fields=['account'])
        updated += 1

    if updated:
        old_account.recalculate_tree_balances()
        new_account.recalculate_tree_balances()
    return updated


def _reverse_quick_enrollment_entry(enrollment, student_account, user, note=''):
    entry = _get_quick_enrollment_entry(enrollment)
    if not entry:
        return None

    deferred_account = Account.get_or_create_quick_course_deferred_account(enrollment.course)
    description = f'إلغاء تسجيل مكرر لطالب سريع - {enrollment.student.full_name} - {enrollment.course.name}'
    if note:
        description = f'{description} - {note}'

    reversing_entry = JournalEntry.objects.create(
        date=timezone.now().date(),
        description=description,
        entry_type='ADJUSTMENT',
        total_amount=enrollment.net_amount or Decimal('0'),
        created_by=user,
    )
    Transaction.objects.create(
        journal_entry=reversing_entry,
        account=deferred_account,
        amount=enrollment.net_amount or Decimal('0'),
        is_debit=True,
        description=f'عكس إيراد مؤجل - {enrollment.course.name}',
    )
    Transaction.objects.create(
        journal_entry=reversing_entry,
        account=student_account,
        amount=enrollment.net_amount or Decimal('0'),
        is_debit=False,
        description=f'عكس ذمة تسجيل مكرر - {enrollment.student.full_name}',
    )
    reversing_entry.post_entry(user)
    return reversing_entry


def _reverse_quick_receipt_entry(receipt, student_account, user, note=''):
    if not receipt or not receipt.journal_entry_id or (receipt.paid_amount or Decimal('0')) <= 0:
        return None

    entry = receipt.journal_entry
    if not entry or not entry.is_posted:
        return None

    cash_account = None
    for tx in entry.transactions.select_related('account').all():
        if tx.is_debit:
            cash_account = tx.account
            break

    if not cash_account:
        return None

    description = f'إلغاء إيصال مكرر لطالب سريع - {receipt.student_name} - {receipt.course_name or "-"}'
    if note:
        description = f'{description} - {note}'

    reversing_entry = JournalEntry.objects.create(
        date=timezone.now().date(),
        description=description,
        entry_type='ADJUSTMENT',
        total_amount=receipt.paid_amount or Decimal('0'),
        created_by=user,
    )
    Transaction.objects.create(
        journal_entry=reversing_entry,
        account=student_account,
        amount=receipt.paid_amount or Decimal('0'),
        is_debit=True,
        description=f'عكس تسديد ذمة لإيصال مكرر - {receipt.course_name or "-"}',
    )
    Transaction.objects.create(
        journal_entry=reversing_entry,
        account=cash_account,
        amount=receipt.paid_amount or Decimal('0'),
        is_debit=False,
        description=f'عكس قبض إيصال مكرر - {receipt.student_name}',
    )
    reversing_entry.post_entry(user)
    return reversing_entry


def _pick_merge_target(students):
    return sorted(
        students,
        key=lambda student: (
            not student.is_active,
            student.created_at or timezone.now(),
            student.id,
        )
    )[0]


def _merge_quick_students_by_name(normalized_name, user):
    _configure_sqlite_busy_timeout()

    students = list(
        QuickStudent.objects.select_related('student', 'academic_year', 'created_by')
        .filter(full_name__isnull=False)
        .order_by('created_at', 'id')
    )
    matched_students = [
        student for student in students
        if _normalize_quick_student_name(student.full_name) == normalized_name
    ]
    if len(matched_students) < 2:
        raise ValueError('لم يعد يوجد سجلات مكررة لهذا الاسم.')

    target = _pick_merge_target(matched_students)
    sources = [student for student in matched_students if student.id != target.id]
    target_ar = Account.get_or_create_quick_student_ar_account(target)

    touched_accounts = {target_ar.id: target_ar}
    target_enrollments = {
        enrollment.course_id: enrollment
        for enrollment in QuickEnrollment.objects.select_related('course').filter(student=target)
    }

    merged_enrollments = 0
    merged_receipts = 0
    reversed_duplicates = 0
    deactivated_sources = []

    with transaction.atomic():
        for source in sources:
            source_ar = Account.get_or_create_quick_student_ar_account(source)
            touched_accounts[source_ar.id] = source_ar
            deleted_receipt_ids = set()

            source_receipts = list(
                QuickStudentReceipt.objects.select_related('course', 'quick_enrollment', 'journal_entry')
                .filter(quick_student=source)
                .order_by('date', 'id')
            )
            source_receipts_by_enrollment = defaultdict(list)
            source_orphan_receipts = []
            for receipt in source_receipts:
                if receipt.quick_enrollment_id:
                    source_receipts_by_enrollment[receipt.quick_enrollment_id].append(receipt)
                else:
                    source_orphan_receipts.append(receipt)

            enrollment_map = {}
            source_enrollments = list(
                QuickEnrollment.objects.select_related('course')
                .filter(student=source)
                .order_by('enrollment_date', 'id')
            )
            for enrollment in source_enrollments:
                existing = target_enrollments.get(enrollment.course_id)
                source_receipt_total = sum(
                    (receipt.paid_amount or Decimal('0')) for receipt in source_receipts_by_enrollment.get(enrollment.id, [])
                )

                if existing:
                    _reverse_quick_enrollment_entry(
                        enrollment=enrollment,
                        student_account=source_ar,
                        user=user,
                        note='دمج حسابات الطلاب السريعين',
                    )
                    duplicate_receipts = list(source_receipts_by_enrollment.get(enrollment.id, []))
                    duplicate_orphan_receipts = [
                        receipt for receipt in source_orphan_receipts
                        if receipt.course_id == enrollment.course_id
                    ]
                    for duplicate_receipt in duplicate_receipts + duplicate_orphan_receipts:
                        _reverse_quick_receipt_entry(
                            receipt=duplicate_receipt,
                            student_account=source_ar,
                            user=user,
                            note='دمج حسابات الطلاب السريعين',
                        )
                        if duplicate_receipt in source_orphan_receipts:
                            source_orphan_receipts.remove(duplicate_receipt)
                        deleted_receipt_ids.add(duplicate_receipt.id)
                        duplicate_receipt.delete()

                    reversed_duplicates += 1
                    enrollment_map[enrollment.id] = existing
                    continue

                QuickEnrollment.objects.filter(pk=enrollment.pk).update(student=target)
                enrollment.student = target
                entry = _get_quick_enrollment_entry(enrollment)
                _retarget_journal_account(entry, source_ar, target_ar)
                target_enrollments[enrollment.course_id] = enrollment
                enrollment_map[enrollment.id] = enrollment
                merged_enrollments += 1

            for receipt in source_receipts:
                if receipt.id in deleted_receipt_ids:
                    continue
                target_enrollment = None
                if receipt.quick_enrollment_id:
                    target_enrollment = enrollment_map.get(receipt.quick_enrollment_id)
                elif receipt.course_id:
                    target_enrollment = target_enrollments.get(receipt.course_id)

                QuickStudentReceipt.objects.filter(pk=receipt.pk).update(
                    quick_student=target,
                    student_name=target.full_name,
                    quick_enrollment=target_enrollment,
                )
                receipt.quick_student = target
                receipt.student_name = target.full_name
                receipt.quick_enrollment = target_enrollment
                _retarget_journal_account(receipt.journal_entry, source_ar, target_ar)
                merged_receipts += 1

            _retarget_all_account_transactions(source_ar, target_ar)
            QuickReceiptPrintJob.objects.filter(quick_student=source).update(quick_student=target)

            source_notes = ((source.notes or '').strip() + f'\n[MERGED_INTO #{target.id}]').strip()
            QuickStudent.objects.filter(pk=source.pk).update(
                is_active=False,
                notes=source_notes,
            )
            source.is_active = False
            source.notes = source_notes

            if getattr(source, 'student', None):
                source_student_notes = ((source.student.notes or '').strip() + f'\nQuick merged into #{target.id}').strip()
                type(source.student).objects.filter(pk=source.student.pk).update(
                    is_active=False,
                    notes=source_student_notes,
                )
                source.student.is_active = False
                source.student.notes = source_student_notes

            source_id = source.pk
            source.delete()
            deactivated_sources.append(source_id)

        if getattr(target, 'student', None):
            updated_fields = []
            if not target.student.phone and target.phone:
                target.student.phone = target.phone
                updated_fields.append('phone')
            if not target.student.full_name and target.full_name:
                target.student.full_name = target.full_name
                updated_fields.append('full_name')
            if updated_fields:
                type(target.student).objects.filter(pk=target.student.pk).update(
                    **{field: getattr(target.student, field) for field in updated_fields}
                )

    for account in touched_accounts.values():
        try:
            account.recalculate_tree_balances()
        except Exception:
            continue

    return {
        'target': target,
        'sources': deactivated_sources,
        'merged_enrollments': merged_enrollments,
        'merged_receipts': merged_receipts,
        'reversed_duplicates': reversed_duplicates,
    }


def _merge_quick_students_by_name_with_retry(normalized_name, user, attempts=3):
    last_error = None
    for attempt in range(attempts):
        try:
            return _merge_quick_students_by_name(normalized_name, user)
        except OperationalError as exc:
            if 'database is locked' not in str(exc).lower():
                raise
            last_error = exc
            connection.close()
            time.sleep(0.75 * (attempt + 1))
    if last_error:
        raise last_error


def _configure_sqlite_busy_timeout(timeout_ms=30000):
    if connection.vendor == 'sqlite':
        with connection.cursor() as cursor:
            cursor.execute(f'PRAGMA busy_timeout = {int(timeout_ms)}')


def _relink_quick_name_group_to_target(normalized_name, target_id, user):
    _configure_sqlite_busy_timeout()

    students = list(
        QuickStudent.objects.select_related('student', 'academic_year', 'created_by')
        .filter(full_name__isnull=False)
        .order_by('created_at', 'id')
    )
    matched_students = [
        student for student in students
        if _normalize_quick_student_name(student.full_name) == normalized_name
    ]
    target = next((student for student in matched_students if student.id == target_id), None)
    if not target:
        raise ValueError('الحساب الهدف غير موجود ضمن نفس الاسم.')

    sources = [student for student in matched_students if student.id != target.id]
    target_ar = Account.get_or_create_quick_student_ar_account(target)
    touched_accounts = {target_ar.id: target_ar}
    target_enrollments = {
        enrollment.course_id: enrollment
        for enrollment in QuickEnrollment.objects.select_related('course').filter(student=target)
    }

    moved_enrollments = 0
    moved_receipts = 0
    skipped_conflicting_enrollments = 0
    relinked_print_jobs = 0
    touched_source_ids = []
    repaired_enrollments = 0
    repaired_receipts = 0
    assigned_sessions = 0
    reactivated_target = False
    active_enrollments_after = 0

    def repair_target_links():
        nonlocal repaired_enrollments, repaired_receipts, assigned_sessions, reactivated_target, active_enrollments_after

        target_enrollment_list = list(
            QuickEnrollment.objects.select_related('course').filter(student=target).order_by('enrollment_date', 'id')
        )
        for enrollment in target_enrollment_list:
            entry = _find_quick_enrollment_entry(enrollment)
            deferred_account = Account.get_or_create_quick_course_deferred_account(enrollment.course)
            expected_amount = enrollment.net_amount or Decimal('0')
            entry_ok = False
            if entry:
                debit_ok = entry.transactions.filter(
                    account=target_ar,
                    is_debit=True,
                    amount=expected_amount,
                ).exists()
                credit_ok = entry.transactions.filter(
                    account=deferred_account,
                    is_debit=False,
                    amount=expected_amount,
                ).exists()
                entry_ok = debit_ok and credit_ok

            if not entry:
                if _ensure_quick_enrollment_entry(enrollment, user=user):
                    repaired_enrollments += 1
            elif not entry_ok:
                if _rebuild_quick_enrollment_entry(enrollment, user):
                    repaired_enrollments += 1

            receipts = list(
                QuickStudentReceipt.objects.filter(
                    quick_student=target,
                    course_id=enrollment.course_id,
                ).select_related('journal_entry').order_by('date', 'id')
            )
            for receipt in receipts:
                updated_fields = []
                if receipt.quick_enrollment_id != enrollment.id:
                    receipt.quick_enrollment = enrollment
                    updated_fields.append('quick_enrollment')
                if receipt.student_name != target.full_name:
                    receipt.student_name = target.full_name
                    updated_fields.append('student_name')
                expected_course_name = enrollment.course.name if enrollment.course else receipt.course_name
                if expected_course_name and receipt.course_name != expected_course_name:
                    receipt.course_name = expected_course_name
                    updated_fields.append('course_name')
                if updated_fields:
                    receipt.save(update_fields=updated_fields)
                    repaired_receipts += 1

                if not receipt.journal_entry_id:
                    if _rebuild_quick_receipt_entry(receipt, user):
                        repaired_receipts += 1
                    continue

                if not _fix_quick_receipt_entry(receipt, target_ar):
                    if _rebuild_quick_receipt_entry(receipt, user):
                        repaired_receipts += 1

            if not enrollment.is_completed and not getattr(enrollment, 'session_assignment', None):
                assignment = _assign_enrollment_to_available_session(enrollment, user)
                if assignment is not None:
                    assigned_sessions += 1

        active_enrollments_after = sum(1 for enrollment in target_enrollment_list if not enrollment.is_completed)
        if active_enrollments_after > 0:
            student_updates = {}
            if not target.is_active:
                target.is_active = True
                student_updates['is_active'] = True
            if student_updates:
                QuickStudent.objects.filter(pk=target.pk).update(**student_updates)
                reactivated_target = True

            if getattr(target, 'student', None) and not target.student.is_active:
                type(target.student).objects.filter(pk=target.student.pk).update(is_active=True)
                target.student.is_active = True

    if not matched_students:
        raise ValueError('لا يوجد أي سجل مطابق لهذا الاسم.')

    with transaction.atomic():
        for source in sources:
            source_ar = Account.get_or_create_quick_student_ar_account(source)
            touched_accounts[source_ar.id] = source_ar
            touched_source_ids.append(source.id)

            source_receipts = list(
                QuickStudentReceipt.objects.select_related('course', 'quick_enrollment', 'journal_entry')
                .filter(quick_student=source)
                .order_by('date', 'id')
            )
            source_receipts_by_enrollment = defaultdict(list)
            source_orphan_receipts = []
            for receipt in source_receipts:
                if receipt.quick_enrollment_id:
                    source_receipts_by_enrollment[receipt.quick_enrollment_id].append(receipt)
                else:
                    source_orphan_receipts.append(receipt)

            source_enrollments = list(
                QuickEnrollment.objects.select_related('course')
                .filter(student=source)
                .order_by('enrollment_date', 'id')
            )
            enrollment_map = {}

            for enrollment in source_enrollments:
                existing = target_enrollments.get(enrollment.course_id)
                if existing:
                    enrollment_map[enrollment.id] = existing
                    skipped_conflicting_enrollments += 1
                    continue

                QuickEnrollment.objects.filter(pk=enrollment.pk).update(student=target)
                enrollment.student = target
                _retarget_journal_account(_get_quick_enrollment_entry(enrollment), source_ar, target_ar)
                target_enrollments[enrollment.course_id] = enrollment
                enrollment_map[enrollment.id] = enrollment
                moved_enrollments += 1

            for receipt in source_receipts:
                target_enrollment = None
                if receipt.quick_enrollment_id:
                    target_enrollment = enrollment_map.get(receipt.quick_enrollment_id)
                elif receipt.course_id:
                    target_enrollment = target_enrollments.get(receipt.course_id)

                QuickStudentReceipt.objects.filter(pk=receipt.pk).update(
                    quick_student=target,
                    student_name=target.full_name,
                    quick_enrollment=target_enrollment,
                )
                receipt.quick_student = target
                receipt.student_name = target.full_name
                receipt.quick_enrollment = target_enrollment
                _retarget_journal_account(receipt.journal_entry, source_ar, target_ar)
                moved_receipts += 1

            relinked_print_jobs += QuickReceiptPrintJob.objects.filter(quick_student=source).update(
                quick_student=target
            )

        if getattr(target, 'student', None):
            updated_fields = []
            if not target.student.phone and target.phone:
                target.student.phone = target.phone
                updated_fields.append('phone')
            if not target.student.full_name and target.full_name:
                target.student.full_name = target.full_name
                updated_fields.append('full_name')
            if updated_fields:
                type(target.student).objects.filter(pk=target.student.pk).update(
                    **{field: getattr(target.student, field) for field in updated_fields}
                )

        repair_target_links()

    for account in touched_accounts.values():
        try:
            account.recalculate_tree_balances()
        except Exception:
            continue

    return {
        'target': target,
        'sources': touched_source_ids,
        'moved_enrollments': moved_enrollments,
        'moved_receipts': moved_receipts,
        'skipped_conflicting_enrollments': skipped_conflicting_enrollments,
        'relinked_print_jobs': relinked_print_jobs,
        'repaired_enrollments': repaired_enrollments,
        'repaired_receipts': repaired_receipts,
        'assigned_sessions': assigned_sessions,
        'reactivated_target': reactivated_target,
        'active_enrollments_after': active_enrollments_after,
    }


def _run_quick_student_checking(target, user):
    _configure_sqlite_busy_timeout()

    target_ar = Account.get_or_create_quick_student_ar_account(target)
    enrollments = list(
        QuickEnrollment.objects.select_related('course').filter(student=target).order_by('enrollment_date', 'id')
    )
    receipts = list(
        QuickStudentReceipt.objects.filter(
            Q(quick_student=target) | Q(student_name=target.full_name)
        ).select_related('course', 'quick_enrollment', 'journal_entry').order_by('date', 'id')
    )

    touched_accounts = {target_ar.id: target_ar}
    reactivated_enrollments = 0
    synthesized_receipts = 0
    deleted_journal_entries = 0
    rebuilt_enrollment_entries = 0
    relinked_receipts = 0
    rebuilt_receipt_entries = 0
    assigned_sessions = 0

    with transaction.atomic():
        enrollment_creator_by_id = {}
        for enrollment in enrollments:
            existing_entry = _find_quick_enrollment_entry(enrollment)
            creator = getattr(existing_entry, 'created_by', None) or user
            enrollment_creator_by_id[enrollment.id] = creator

        receipt_creator_by_id = {}
        for receipt in receipts:
            creator = getattr(receipt, 'created_by', None)
            if not creator and receipt.journal_entry_id:
                creator = getattr(receipt.journal_entry, 'created_by', None)
            receipt_creator_by_id[receipt.id] = creator or user

        if not target.is_active:
            QuickStudent.objects.filter(pk=target.pk).update(is_active=True)
            target.is_active = True
        if getattr(target, 'student', None) and not target.student.is_active:
            type(target.student).objects.filter(pk=target.student.pk).update(is_active=True)
            target.student.is_active = True

        enrollment_by_course = {}
        for enrollment in enrollments:
            if enrollment.is_completed:
                QuickEnrollment.objects.filter(pk=enrollment.pk).update(is_completed=False, completion_date=None)
                enrollment.is_completed = False
                enrollment.completion_date = None
                reactivated_enrollments += 1

            cleanup_result = _cleanup_quick_withdrawal_entries(enrollment, user)
            deleted_journal_entries += (
                len(cleanup_result['deleted_ids']) +
                len(cleanup_result['reversed_ids'])
            )

            entry_user = enrollment_creator_by_id.get(enrollment.id) or user
            if _find_quick_enrollment_entry(enrollment):
                deleted_journal_entries += 1
            if _rebuild_quick_enrollment_entry(enrollment, entry_user):
                rebuilt_enrollment_entries += 1

            enrollment_by_course[enrollment.course_id] = enrollment

        if not receipts:
            for payload in _extract_legacy_quick_receipt_payloads(target, enrollments):
                creator = payload.pop('created_by', None) or user
                QuickStudentReceipt.objects.create(created_by=creator, **payload)
                synthesized_receipts += 1

            if synthesized_receipts:
                receipts = list(
                    QuickStudentReceipt.objects.filter(
                        Q(quick_student=target) | Q(student_name=target.full_name)
                    ).select_related('course', 'quick_enrollment', 'journal_entry').order_by('date', 'id')
                )

        for receipt in receipts:
            target_enrollment = None
            if receipt.quick_enrollment_id and receipt.quick_enrollment.student_id == target.id:
                target_enrollment = receipt.quick_enrollment
            elif receipt.course_id:
                target_enrollment = enrollment_by_course.get(receipt.course_id)

            updates = {}
            if receipt.quick_student_id != target.id:
                updates['quick_student'] = target
                receipt.quick_student = target
            if receipt.student_name != target.full_name:
                updates['student_name'] = target.full_name
                receipt.student_name = target.full_name
            if target_enrollment and receipt.quick_enrollment_id != target_enrollment.id:
                updates['quick_enrollment'] = target_enrollment
                receipt.quick_enrollment = target_enrollment
            if receipt.course and receipt.course_name != receipt.course.name:
                updates['course_name'] = receipt.course.name
                receipt.course_name = receipt.course.name

            if updates:
                QuickStudentReceipt.objects.filter(pk=receipt.pk).update(
                    **{
                        key: (value.id if key in {'quick_student', 'quick_enrollment'} else value)
                        for key, value in updates.items()
                    }
                )
                relinked_receipts += 1

            entry_user = receipt_creator_by_id.get(receipt.id) or user
            if receipt.journal_entry_id:
                deleted_journal_entries += 1
            if _rebuild_quick_receipt_entry(receipt, entry_user):
                rebuilt_receipt_entries += 1

        for enrollment in enrollments:
            if not getattr(enrollment, 'session_assignment', None):
                assignment = _assign_enrollment_to_available_session(enrollment, user)
                if assignment is not None:
                    assigned_sessions += 1

        keep_references = {f'QE-{enrollment.id}' for enrollment in enrollments}
        keep_entry_ids = {
            row['journal_entry_id']
            for row in QuickStudentReceipt.objects.filter(quick_student=target)
            .values('journal_entry_id')
            if row['journal_entry_id']
        }
        purged_extra_ids = _purge_quick_extra_entries(
            target,
            keep_entry_ids=keep_entry_ids,
            keep_references=keep_references,
        )
        deleted_journal_entries += len(purged_extra_ids)

    for account in touched_accounts.values():
        try:
            account.recalculate_tree_balances()
        except Exception:
            continue

    return {
        'target': target,
        'enrollments_count': len(enrollments),
        'receipts_count': len(receipts),
        'reactivated_enrollments': reactivated_enrollments,
        'synthesized_receipts': synthesized_receipts,
        'deleted_journal_entries': deleted_journal_entries,
        'rebuilt_enrollment_entries': rebuilt_enrollment_entries,
        'relinked_receipts': relinked_receipts,
        'rebuilt_receipt_entries': rebuilt_receipt_entries,
        'assigned_sessions': assigned_sessions,
        'validation': _validate_quick_student_checking(target),
    }


def _validate_quick_student_checking(target):
    enrollments = list(
        QuickEnrollment.objects.select_related('course').filter(student=target).order_by('id')
    )
    receipts = list(
        QuickStudentReceipt.objects.filter(quick_student=target).select_related('journal_entry').order_by('id')
    )

    missing_enrollment_entry_ids = [
        enrollment.id
        for enrollment in enrollments
        if not _find_quick_enrollment_entry(enrollment)
    ]
    receipt_candidates = [receipt for receipt in receipts if (receipt.paid_amount or Decimal('0')) > 0]
    missing_receipt_entry_ids = [
        receipt.id
        for receipt in receipt_candidates
        if not receipt.journal_entry_id
    ]

    expected_receipt_entry_ids = {receipt.journal_entry_id for receipt in receipt_candidates if receipt.journal_entry_id}
    expected_enrollment_refs = {f'QE-{enrollment.id}' for enrollment in enrollments}
    related_entries = list(_get_quick_student_related_journal_entries(target))
    extra_entry_ids = [
        entry.id
        for entry in related_entries
        if entry.reference not in expected_enrollment_refs and entry.id not in expected_receipt_entry_ids
    ]

    return {
        'expected_enrollment_entries': len(enrollments),
        'actual_enrollment_entries': len(enrollments) - len(missing_enrollment_entry_ids),
        'expected_receipt_entries': len(receipt_candidates),
        'actual_receipt_entries': len(receipt_candidates) - len(missing_receipt_entry_ids),
        'missing_enrollment_entry_ids': missing_enrollment_entry_ids,
        'missing_receipt_entry_ids': missing_receipt_entry_ids,
        'extra_entry_ids': extra_entry_ids,
        'is_clean': not (
            missing_enrollment_entry_ids
            or missing_receipt_entry_ids
            or extra_entry_ids
        ),
    }


def _run_quick_student_checking_batch(students, user):
    results = []
    errors = []

    for student in students:
        try:
            result = _run_quick_student_checking_with_retry(student, user)
        except Exception as exc:
            errors.append({
                'student_id': student.id,
                'student_name': student.full_name,
                'error': str(exc),
            })
            continue
        results.append(result)

    return {
        'processed': len(results),
        'failed': len(errors),
        'results': results,
        'errors': errors,
        'reactivated_enrollments': sum(item['reactivated_enrollments'] for item in results),
        'synthesized_receipts': sum(item['synthesized_receipts'] for item in results),
        'deleted_journal_entries': sum(item['deleted_journal_entries'] for item in results),
        'rebuilt_enrollment_entries': sum(item['rebuilt_enrollment_entries'] for item in results),
        'relinked_receipts': sum(item['relinked_receipts'] for item in results),
        'rebuilt_receipt_entries': sum(item['rebuilt_receipt_entries'] for item in results),
        'assigned_sessions': sum(item['assigned_sessions'] for item in results),
        'clean_count': sum(1 for item in results if item['validation']['is_clean']),
    }


def _run_quick_student_checking_with_retry(student, user, attempts=4):
    last_error = None
    for attempt in range(attempts):
        try:
            close_old_connections()
            _configure_sqlite_busy_timeout(timeout_ms=45000)
            return _run_quick_student_checking(student, user)
        except OperationalError as exc:
            if 'database is locked' not in str(exc).lower():
                raise
            last_error = exc
            connection.close()
            time.sleep(0.6 * (attempt + 1))
    if last_error:
        raise last_error


def _extract_legacy_quick_receipt_payloads(target, enrollments):
    receipts = []
    seen_refs = set(
        QuickStudentReceipt.objects.filter(quick_student=target).values_list('receipt_number', flat=True)
    )
    enrollment_by_course = {enrollment.course_id: enrollment for enrollment in enrollments}
    related_entries = list(_get_quick_student_related_journal_entries(target).order_by('date', 'id'))

    for entry in related_entries:
        if entry.entry_type != 'receipt':
            continue
        reference = (entry.reference or '').strip()
        if reference and reference in seen_refs:
            continue

        matched_enrollment = None
        for enrollment in enrollments:
            course_name = str(getattr(enrollment.course, 'name', '') or '')
            if course_name and course_name in str(entry.description or ''):
                matched_enrollment = enrollment
                break

        if not matched_enrollment:
            continue

        amount = entry.total_amount or Decimal('0')
        if amount <= 0:
            continue

        receipts.append({
            'receipt_number': reference or None,
            'date': entry.date,
            'student_name': target.full_name,
            'quick_student': target,
            'course': matched_enrollment.course,
            'course_name': matched_enrollment.course.name,
            'quick_enrollment': matched_enrollment,
            'amount': amount,
            'paid_amount': amount,
            'payment_method': 'CASH',
            'notes': 'تم إنشاؤه تلقائيًا من قيد قبض قديم عبر أداة التشييك',
            'created_by': entry.created_by or target.created_by or None,
        })
        if reference:
            seen_refs.add(reference)

    return receipts


def _purge_quick_extra_entries(target, keep_entry_ids=None, keep_references=None):
    keep_entry_ids = {entry_id for entry_id in (keep_entry_ids or set()) if entry_id}
    keep_references = {str(ref).strip() for ref in (keep_references or set()) if str(ref).strip()}

    extra_entries = []
    for entry in _get_quick_student_related_journal_entries(target):
        reference = str(entry.reference or '').strip()
        if entry.id in keep_entry_ids:
            continue
        if reference and reference in keep_references:
            continue
        extra_entries.append(entry)

    if not extra_entries:
        return []

    extra_ids = [entry.id for entry in extra_entries]
    QuickStudentReceipt.objects.filter(journal_entry_id__in=extra_ids).update(journal_entry=None)
    JournalEntry.objects.filter(id__in=extra_ids).delete()
    return extra_ids


@require_superuser
def quick_checking_tool(request):
    eligible_students_qs = (
        QuickStudent.objects.filter(
            Q(enrollments__isnull=False) | Q(quickstudentreceipt__isnull=False)
        )
        .distinct()
        .order_by('id')
    )

    if request.method == 'POST':
        action = request.POST.get('action')
        target_id_raw = request.POST.get('target_student_id')
        search_query = (request.POST.get('q') or '').strip()
        if action == 'run_all':
            batch_size_raw = request.POST.get('batch_size') or '25'
            start_after_raw = request.POST.get('start_after_id') or '0'
            try:
                batch_size = max(1, min(100, int(batch_size_raw)))
            except (TypeError, ValueError):
                batch_size = 25
            try:
                start_after_id = max(0, int(start_after_raw))
            except (TypeError, ValueError):
                start_after_id = 0

            students = list(eligible_students_qs.filter(id__gt=start_after_id)[:batch_size])
            if not students:
                messages.success(request, 'اكتمل التشييك الجماعي على جميع الطلاب المطلوبين.')
                return redirect(reverse('quick:checking_tool'))

            batch = _run_quick_student_checking_batch(students, request.user)
            next_start_id = students[-1].id
            remaining_count = eligible_students_qs.filter(id__gt=next_start_id).count()
            messages.success(
                request,
                f'تم تشييك دفعة من {batch["processed"]} طالبًا: '
                f'إعادة تفعيل {batch["reactivated_enrollments"]} تسجيل، '
                f'وإنشاء {batch["synthesized_receipts"]} إيصال من قيود قبض قديمة، '
                f'وحذف {batch["deleted_journal_entries"]} قيد قديم، '
                f'وإعادة بناء {batch["rebuilt_enrollment_entries"]} قيد تسجيل، '
                f'وإعادة بناء {batch["rebuilt_receipt_entries"]} قيد قبض، '
                f'وتوزيع {batch["assigned_sessions"]} تسجيل على الكلاسات.'
            )
            if batch['failed']:
                sample_errors = ' | '.join(
                    f'#{item["student_id"]} {item["student_name"]}: {item["error"]}'
                    for item in batch['errors'][:3]
                )
                messages.warning(
                    request,
                    f'فشل التشييك على {batch["failed"]} طالب. أمثلة: {sample_errors}'
                )
            messages.info(
                request,
                f'التحقق النهائي النظيف تحقق لـ {batch["clean_count"]} من أصل {batch["processed"]} طالب تمت معالجتهم.'
            )
            if remaining_count > 0:
                messages.warning(
                    request,
                    f'تبقى {remaining_count} طالبًا للتشييك الجماعي. اضغط زر التشييك الجماعي مرة أخرى للمتابعة من بعد السجل #{next_start_id}.'
                )
                return redirect(
                    f"{reverse('quick:checking_tool')}?{urlencode({'batch_cursor': next_start_id, 'batch_size': batch_size})}"
                )
            messages.success(request, 'اكتمل التشييك الجماعي على جميع الطلاب المطلوبين.')
            return redirect(reverse('quick:checking_tool'))
        else:
            try:
                target_id = int(target_id_raw)
            except (TypeError, ValueError):
                messages.error(request, 'السجل الهدف غير صالح.')
                return redirect(f"{reverse('quick:checking_tool')}?{urlencode({'q': search_query})}")

            target = get_object_or_404(QuickStudent, pk=target_id)
            try:
                result = _run_quick_student_checking(target, request.user)
            except Exception as exc:
                messages.error(request, f'فشل تنفيذ أداة التشييك: {exc}')
            else:
                validation = result['validation']
                messages.success(
                    request,
                    f'تمت معالجة السجل #{result["target"].id}: '
                    f'إعادة تفعيل {result["reactivated_enrollments"]} تسجيل، '
                    f'وإنشاء {result["synthesized_receipts"]} إيصال من قيود قبض قديمة، '
                    f'وحذف {result["deleted_journal_entries"]} قيد قديم، '
                    f'وإعادة بناء {result["rebuilt_enrollment_entries"]} قيد تسجيل، '
                    f'وربط/فحص {result["relinked_receipts"]} إيصال، '
                    f'وإعادة بناء {result["rebuilt_receipt_entries"]} قيد قبض، '
                    f'وتوزيع {result["assigned_sessions"]} تسجيل على الكلاسات.'
                )
                if validation['is_clean']:
                    messages.success(
                        request,
                        f'التحقق النهائي ناجح: {validation["actual_enrollment_entries"]}/{validation["expected_enrollment_entries"]} قيد تسجيل '
                        f'و{validation["actual_receipt_entries"]}/{validation["expected_receipt_entries"]} قيد قبض، بدون قيود زائدة على هذا الحساب.'
                    )
                else:
                    messages.warning(
                        request,
                        f'نتيجة التحقق النهائي: '
                        f'نواقص قيود التسجيل {len(validation["missing_enrollment_entry_ids"])}, '
                        f'نواقص قيود القبض {len(validation["missing_receipt_entry_ids"])}, '
                        f'وقيود زائدة {len(validation["extra_entry_ids"])}.'
                    )
        return redirect(f"{reverse('quick:checking_tool')}?{urlencode({'q': search_query})}")

    search_query = (request.GET.get('q') or '').strip()
    normalized_search = _normalize_quick_student_name(search_query)
    batch_cursor_raw = request.GET.get('batch_cursor') or '0'
    batch_size_raw = request.GET.get('batch_size') or '25'
    try:
        batch_cursor = max(0, int(batch_cursor_raw))
    except (TypeError, ValueError):
        batch_cursor = 0
    try:
        batch_size = max(1, min(100, int(batch_size_raw)))
    except (TypeError, ValueError):
        batch_size = 25

    students = list(
        QuickStudent.objects.select_related('student', 'academic_year')
        .prefetch_related('enrollments', 'quickstudentreceipt_set')
        .order_by('full_name', 'id')
    )

    rows = []
    for student in students:
        normalized_name = _normalize_quick_student_name(student.full_name)
        if normalized_search and normalized_search not in normalized_name:
            continue
        if not normalized_search:
            continue
        rows.append({
            'student': student,
            'enrollments_count': student.enrollments.count(),
            'receipts_count': student.quickstudentreceipt_set.count(),
            'balance': student.balance,
        })

    return render(request, 'quick/quick_checking_tool.html', {
        'rows': rows,
        'search_query': search_query,
        'all_students_count': len(students),
        'eligible_students_count': eligible_students_qs.count(),
        'batch_cursor': batch_cursor,
        'batch_size': batch_size,
        'remaining_batch_count': eligible_students_qs.filter(id__gt=batch_cursor).count(),
    })


@require_superuser
def quick_name_link_tool(request):
    if request.method == 'POST':
        normalized_name = _normalize_quick_student_name(request.POST.get('group_name'))
        target_id_raw = request.POST.get('target_student_id')
        search_query = (request.POST.get('q') or '').strip()

        try:
            target_id = int(target_id_raw)
        except (TypeError, ValueError):
            messages.error(request, 'السجل الهدف غير صالح.')
            return redirect(f"{reverse('quick:name_link_tool')}?{urlencode({'q': search_query})}")

        if not normalized_name:
            messages.error(request, 'لم يتم تحديد الاسم المطلوب تصحيح ربطه.')
            return redirect(f"{reverse('quick:name_link_tool')}?{urlencode({'q': search_query})}")

        try:
            result = _relink_quick_name_group_to_target(normalized_name, target_id, request.user)
        except Exception as exc:
            messages.error(request, f'فشل تصحيح ربط التسجيلات والإيصالات: {exc}')
        else:
            messages.success(
                request,
                f'تمت معالجة السجل #{result["target"].id}: '
                f'نقل {result["moved_enrollments"]} تسجيل و{result["moved_receipts"]} إيصال، '
                f'وتصحيح {result["repaired_enrollments"]} قيد تسجيل و{result["repaired_receipts"]} عملية ربط/قيد للإيصالات، '
                f'وتوزيع {result["assigned_sessions"]} تسجيل على الكلاسات.'
            )
            if result['skipped_conflicting_enrollments']:
                messages.warning(
                    request,
                    f'تم تخطي {result["skipped_conflicting_enrollments"]} تسجيل بسبب وجود نفس الدورة مسبقًا على السجل الهدف.'
                )
            if result['reactivated_target']:
                messages.success(request, 'تم تفعيل السجل الهدف لأنه يحتوي على تسجيلات نشطة.')
            if result['active_enrollments_after'] == 0:
                messages.warning(
                    request,
                    'لا توجد تسجيلات نشطة على هذا السجل بعد المعالجة. إذا كان الطالب يجب أن يظهر كمسجل فعليًا، فلابد من إعادة فتح التسجيلات المكتملة أو إنشاء تسجيل جديد.'
                )

        return redirect(f"{reverse('quick:name_link_tool')}?{urlencode({'q': search_query})}")

    search_query = (request.GET.get('q') or '').strip()
    normalized_search = _normalize_quick_student_name(search_query)

    students = list(
        QuickStudent.objects.select_related('student', 'academic_year', 'created_by')
        .prefetch_related('enrollments', 'quickstudentreceipt_set')
        .order_by('full_name', 'created_at', 'id')
    )

    grouped_students = defaultdict(list)
    for student in students:
        normalized_name = _normalize_quick_student_name(student.full_name)
        if normalized_name:
            grouped_students[normalized_name].append(student)

    matched_single_records = []
    if normalized_search:
        matched_single_records = [
            student for student in students
            if normalized_search in _normalize_quick_student_name(student.full_name)
        ]

    groups = []
    for normalized_name, members in grouped_students.items():
        if not members:
            continue
        if normalized_search and normalized_search not in normalized_name:
            continue
        if not normalized_search and len(members) < 2:
            continue

        candidate_rows = []
        for member in members:
            candidate_rows.append({
                'student': member,
                'enrollments_count': member.enrollments.count(),
                'receipts_count': member.quickstudentreceipt_set.count(),
                'balance': member.balance,
            })

        groups.append({
            'normalized_name': normalized_name,
            'display_name': members[0].full_name,
            'members': candidate_rows,
            'members_count': len(candidate_rows),
        })

    groups.sort(key=lambda item: (-item['members_count'], item['display_name']))

    return render(request, 'quick/quick_name_link_tool.html', {
        'groups': groups,
        'search_query': search_query,
        'matched_single_records_count': len(matched_single_records),
    })


def _get_duplicate_groups(search_query='', phone_query='', scope='active'):
    include_inactive = scope == 'all'

    enrollment_queryset = (
        QuickEnrollment.objects.select_related('course')
        .prefetch_related(
            Prefetch(
                'quickstudentreceipt_set',
                queryset=QuickStudentReceipt.objects.select_related('created_by', 'journal_entry').order_by('date', 'id')
            )
        )
        .annotate(
            paid_total=Coalesce(
                Sum('quickstudentreceipt__paid_amount'),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=10, decimal_places=2),
            )
        )
        .order_by('course__name', 'id')
    )

    students_queryset = (
        QuickStudent.objects.select_related('student', 'academic_year', 'created_by')
        .prefetch_related(
            Prefetch('enrollments', queryset=enrollment_queryset),
            Prefetch(
                'quickstudentreceipt_set',
                queryset=QuickStudentReceipt.objects.select_related(
                    'created_by', 'journal_entry', 'course', 'quick_enrollment'
                ).order_by('date', 'id')
            )
        )
        .order_by('full_name', 'id')
    )
    if not include_inactive:
        students_queryset = students_queryset.filter(is_active=True)

    grouped_students = defaultdict(list)
    for quick_student in students_queryset:
        normalized_name = _normalize_quick_student_name(quick_student.full_name)
        if normalized_name:
            grouped_students[normalized_name].append(quick_student)

    duplicate_groups = []
    normalized_search = _normalize_quick_student_name(search_query)
    normalized_phone_search = _normalize_phone(phone_query)

    for normalized_name, students in grouped_students.items():
        if len(students) < 2:
            continue
        if normalized_search and normalized_search not in normalized_name:
            continue
        if normalized_phone_search and not any(
            normalized_phone_search in _normalize_phone(student.phone)
            for student in students
        ):
            continue

        members = []
        group_balance = Decimal('0')
        group_remaining = Decimal('0')
        group_enrollments = 0

        for quick_student in students:
            enrollments_data = []
            student_remaining = Decimal('0')
            account_created_by = '-'
            if quick_student.created_by:
                account_created_by = quick_student.created_by.get_full_name() or quick_student.created_by.username or '-'
            all_student_receipts = list(quick_student.quickstudentreceipt_set.all())
            used_receipt_ids = set()
            enrollments = list(quick_student.enrollments.all())
            enrollment_entry_refs = {f'QE-{enrollment.id}': enrollment for enrollment in enrollments}
            journal_entries = {
                entry.reference: entry
                for entry in JournalEntry.objects.filter(reference__in=enrollment_entry_refs.keys()).select_related('created_by')
            }

            for enrollment in enrollments:
                net_amount = enrollment.net_amount or Decimal('0')
                paid_amount = enrollment.paid_total or Decimal('0')
                remaining_amount = max(Decimal('0'), net_amount - paid_amount)
                student_remaining += remaining_amount
                group_enrollments += 1
                enrollment_created_by = '-'
                receipt_rows = []

                enrollment_entry = journal_entries.get(f'QE-{enrollment.id}')
                if enrollment_entry and enrollment_entry.created_by:
                    enrollment_created_by = (
                        enrollment_entry.created_by.get_full_name()
                        or enrollment_entry.created_by.username
                        or '-'
                    )

                receipts = [
                    receipt for receipt in all_student_receipts
                    if receipt.quick_enrollment_id == enrollment.id
                    or (
                        receipt.quick_enrollment_id is None
                        and receipt.course_id == enrollment.course_id
                    )
                ]
                if enrollment_created_by == '-' and receipts:
                    first_receipt_creator = receipts[0].created_by
                    if first_receipt_creator:
                        enrollment_created_by = (
                            first_receipt_creator.get_full_name()
                            or first_receipt_creator.username
                            or '-'
                        )

                for receipt in receipts:
                    receipt_created_by = '-'
                    if receipt.created_by:
                        receipt_created_by = receipt.created_by.get_full_name() or receipt.created_by.username or '-'
                    used_receipt_ids.add(receipt.id)
                    receipt_rows.append({
                        'receipt_number': receipt.receipt_number or '-',
                        'date': receipt.date,
                        'paid_amount': receipt.paid_amount or Decimal('0'),
                        'amount': receipt.amount or Decimal('0'),
                        'is_printed': receipt.is_printed,
                        'created_by': receipt_created_by,
                    })

                enrollments_data.append({
                    'course_name': enrollment.course.name if enrollment.course else '-',
                    'enrollment_date': enrollment.enrollment_date,
                    'net_amount': net_amount,
                    'paid_amount': paid_amount,
                    'remaining_amount': remaining_amount,
                    'is_completed': enrollment.is_completed,
                    'registered_by': enrollment_created_by,
                    'receipts': receipt_rows,
                    'has_receipts': bool(receipt_rows),
                    'receipts_count': len(receipt_rows),
                })

            orphan_receipts = []
            for receipt in all_student_receipts:
                if receipt.id in used_receipt_ids:
                    continue
                receipt_created_by = '-'
                if receipt.created_by:
                    receipt_created_by = receipt.created_by.get_full_name() or receipt.created_by.username or '-'
                orphan_receipts.append({
                    'receipt_number': receipt.receipt_number or '-',
                    'date': receipt.date,
                    'course_name': receipt.course_name or (receipt.course.name if receipt.course else '-'),
                    'paid_amount': receipt.paid_amount or Decimal('0'),
                    'amount': receipt.amount or Decimal('0'),
                    'is_printed': receipt.is_printed,
                    'created_by': receipt_created_by,
                })

            account_balance = quick_student.balance
            group_balance += account_balance
            group_remaining += student_remaining

            members.append({
                'student': quick_student,
                'account_created_by': account_created_by,
                'account_balance': account_balance,
                'remaining_total': student_remaining,
                'enrollments': enrollments_data,
                'enrollments_count': len(enrollments_data),
                'orphan_receipts': orphan_receipts,
                'has_orphan_receipts': bool(orphan_receipts),
            })

        members.sort(key=lambda item: item['student'].id)
        duplicate_groups.append({
            'display_name': students[0].full_name,
            'normalized_name': normalized_name,
            'students': members,
            'duplicate_count': len(members),
            'group_balance': group_balance,
            'group_remaining': group_remaining,
            'group_enrollments': group_enrollments,
        })

    duplicate_groups.sort(key=lambda item: (-item['duplicate_count'], item['display_name']))
    return duplicate_groups


def _build_quick_accounting_fix_rows():
    rows = []
    enrollments = (
        QuickEnrollment.objects.select_related('student', 'course')
        .order_by('-updated_at', '-id')
    )

    for enrollment in enrollments:
        issues = []
        suggested_actions = []
        linked_entry = _find_quick_enrollment_entry(enrollment)
        registration_ok = False
        receipts_ok = True
        withdrawal_ok = not enrollment.is_completed
        fixable_receipt_ids = []

        if not linked_entry:
            issues.append('قيد التسجيل مفقود')
            suggested_actions.append('إنشاء قيد التسجيل')
        else:
            student_ar = enrollment.student.ar_account
            deferred_account = Account.get_or_create_quick_course_deferred_account(enrollment.course)
            debit_ok = linked_entry.transactions.filter(
                account=student_ar,
                is_debit=True,
                amount=enrollment.net_amount or Decimal('0'),
            ).exists()
            credit_ok = linked_entry.transactions.filter(
                account=deferred_account,
                is_debit=False,
                amount=enrollment.net_amount or Decimal('0'),
            ).exists()
            registration_ok = debit_ok and credit_ok
            if not registration_ok:
                issues.append('قيد التسجيل لا يطابق الخطة المطلوبة')
                suggested_actions.append('إعادة بناء قيد التسجيل')

        receipts = list(
            QuickStudentReceipt.objects.filter(
                quick_student=enrollment.student,
                quick_enrollment=enrollment,
            ).select_related('journal_entry')
        )
        retained_amount = sum((receipt.paid_amount or Decimal('0')) for receipt in receipts)
        receipt_count = len(receipts)
        audited_receipts = 0
        receipt_issue_count = 0
        for receipt in receipts:
            audited_receipts += 1
            if not receipt.journal_entry_id:
                receipts_ok = False
                receipt_issue_count += 1
                issues.append(f'إيصال #{receipt.id} بدون قيد قبض')
                fixable_receipt_ids.append(receipt.id)
                continue

            receipt_entry = receipt.journal_entry
            credit_ok = receipt_entry.transactions.filter(
                account=enrollment.student.ar_account,
                is_debit=False,
                amount=receipt.paid_amount or Decimal('0'),
            ).exists()
            debit_ok = receipt_entry.transactions.filter(
                is_debit=True,
                amount=receipt.paid_amount or Decimal('0'),
            ).exclude(account=enrollment.student.ar_account).exists()
            if not (credit_ok and debit_ok):
                receipts_ok = False
                receipt_issue_count += 1
                issues.append(f'إيصال #{receipt.id} لا يطابق قيد القبض المطلوب')
                if receipt.journal_entry_id:
                    fixable_receipt_ids.append(receipt.id)

        legacy_entries = list(_find_quick_withdrawal_entries(enrollment))
        refunded_amount = sum((_get_legacy_withdrawal_amount(entry) for entry in legacy_entries), Decimal('0'))

        correction_amount = retained_amount + refunded_amount
        existing_fix_entries = list(_find_quick_generated_withdraw_fix_entries(enrollment))
        already_fixed = bool(existing_fix_entries)

        if enrollment.is_completed and correction_amount > 0:
            if refunded_amount > 0 and retained_amount > 0:
                issues.append('سحب قديم جزئي لم يُستكمل بالكامل وفق الطريقة الجديدة')
            elif refunded_amount > 0:
                issues.append('سحب قديم أبقى أثراً على الدورة أو على حساب الانسحاب القديم')
            elif retained_amount > 0:
                issues.append('سحب قديم لم يصفّر الدفعات المسجلة بالكامل')
            if already_fixed:
                suggested_actions.append('إلغاء قيود السحب القديمة وإعادة إنشائها')
            else:
                suggested_actions.append('إنشاء قيد تصحيح سحب')
            withdrawal_ok = False
        elif enrollment.is_completed:
            if not legacy_entries and not existing_fix_entries and (enrollment.net_amount or Decimal('0')) > 0:
                issues.append('تسجيل مسحوب بدون قيد سحب واضح')
                suggested_actions.append('إنشاء قيد سحب مفقود')
                withdrawal_ok = False
            else:
                withdrawal_ok = True

        if not enrollment.is_completed and retained_amount > (enrollment.net_amount or Decimal('0')):
            issues.append('إجمالي المقبوض أكبر من صافي التسجيل')
            suggested_actions.append('تصحيح مبالغ القبض الزائدة')

        if not suggested_actions and issues:
            suggested_actions.append('مراجعة يدوية')

        rows.append({
            'enrollment': enrollment,
            'missing_entry': not bool(linked_entry),
            'legacy_entries': legacy_entries,
            'legacy_entries_count': len(legacy_entries),
            'existing_fix_entries_count': len(existing_fix_entries),
            'has_legacy_withdrawal': refunded_amount > 0,
            'retained_amount': retained_amount,
            'refunded_amount': refunded_amount,
            'correction_amount': correction_amount,
            'already_fixed': already_fixed,
            'issues': issues,
            'suggested_actions': suggested_actions,
            'is_compliant': not issues,
            'registration_ok': registration_ok,
            'receipts_ok': receipts_ok,
            'withdrawal_ok': withdrawal_ok,
            'receipt_count': receipt_count,
            'audited_receipts': audited_receipts,
            'receipt_issue_count': receipt_issue_count,
            'fixable_receipt_ids': fixable_receipt_ids,
        })

    return rows


def _apply_quick_accounting_fixes(user):
    rows = _build_quick_accounting_fix_rows()
    fixed_links = 0
    fixed_withdrawals = 0
    fixed_receipts = 0
    created_entries = []
    errors = []
    cleaned_withdraw_entries = 0
    deactivated_legacy_accounts = []

    for row in rows:
        enrollment = row['enrollment']
        success = False
        last_exc = None
        for attempt in range(3):
            try:
                close_old_connections()
                _configure_sqlite_busy_timeout()
                with transaction.atomic():
                    if row['missing_entry'] or not row['registration_ok']:
                        if row['missing_entry']:
                            entry = _ensure_quick_enrollment_entry(enrollment, user=user)
                        else:
                            entry = _rebuild_quick_enrollment_entry(enrollment, user=user)
                        if entry:
                            fixed_links += 1

                    if row['fixable_receipt_ids']:
                        for receipt in QuickStudentReceipt.objects.filter(id__in=row['fixable_receipt_ids']).select_related('journal_entry'):
                            if not receipt.journal_entry_id:
                                if _rebuild_quick_receipt_entry(receipt, user):
                                    fixed_receipts += 1
                                continue

                            student_ar = enrollment.student.ar_account
                            if _fix_quick_receipt_entry(receipt, student_ar):
                                fixed_receipts += 1
                            elif _rebuild_quick_receipt_entry(receipt, user):
                                fixed_receipts += 1

                    if enrollment.is_completed and row['correction_amount'] > 0:
                        if row['retained_amount'] > 0:
                            _adjust_quick_receipts_for_refund(
                                enrollment.student,
                                enrollment,
                                row['retained_amount'],
                            )

                        cleanup_result = _cleanup_quick_withdrawal_entries(enrollment, user)
                        cleaned_withdraw_entries += (
                            len(cleanup_result['reversed_ids']) +
                            len(cleanup_result['deleted_ids'])
                        )

                        description = (
                            f"[QUICK_WITHDRAW #{enrollment.id}] "
                            f"تصحيح سحب طالب سريع - {enrollment.student.full_name} - {enrollment.course.name}"
                        )

                        entries = _build_quick_withdrawal_entry(
                            enrollment=enrollment,
                            user=user,
                            refunded_amount=row['retained_amount'],
                            description=description,
                        )
                        created_entries.extend(entry.id for entry in entries)
                        fixed_withdrawals += len(entries)
                    elif enrollment.is_completed and 'تسجيل مسحوب بدون قيد سحب واضح' in row['issues']:
                        description = (
                            f"[QUICK_WITHDRAW #{enrollment.id}] "
                            f"إنشاء قيد سحب مفقود - {enrollment.student.full_name} - {enrollment.course.name}"
                        )
                        entries = _build_quick_withdrawal_entry(
                            enrollment=enrollment,
                            user=user,
                            refunded_amount=row['retained_amount'],
                            description=description,
                        )
                        created_entries.extend(entry.id for entry in entries)
                        fixed_withdrawals += len(entries)

                    if (
                        not enrollment.is_completed and
                        'إجمالي المقبوض أكبر من صافي التسجيل' in row['issues']
                    ):
                        fixed_receipts += _cap_quick_enrollment_receipts_to_net(enrollment, user)
                success = True
                connection.close()
                break
            except OperationalError as exc:
                last_exc = exc
                if 'database is locked' not in str(exc).lower():
                    break
                close_old_connections()
                connection.close()
                time.sleep(1.5 * (attempt + 1))
            except Exception as exc:
                last_exc = exc
                break

        if not success and last_exc:
            errors.append(f'{enrollment.student.full_name} / {enrollment.course.name}: {last_exc}')

    if cleaned_withdraw_entries:
        Account.rebuild_all_balances()

    deactivated_legacy_accounts = _deactivate_quick_legacy_withdrawal_accounts()

    return {
        'fixed_links': fixed_links,
        'fixed_withdrawals': fixed_withdrawals,
        'fixed_receipts': fixed_receipts,
        'cleaned_withdraw_entries': cleaned_withdraw_entries,
        'deactivated_legacy_accounts': deactivated_legacy_accounts,
        'created_entries': created_entries,
        'errors': errors,
    }


def _build_quick_withdrawal_fix_rows():
    rows = []
    for row in _build_quick_accounting_fix_rows():
        enrollment = row['enrollment']
        needs_fix = (
            row['legacy_entries_count'] > 0 or
            (enrollment.is_completed and row['correction_amount'] > 0) or
            (enrollment.is_completed and 'تسجيل مسحوب بدون قيد سحب واضح' in row['issues']) or
            row['existing_fix_entries_count'] > 0
        )
        if not needs_fix:
            continue

        rows.append(row)

    return rows


def _apply_quick_withdrawal_fixes(user):
    rows = _build_quick_withdrawal_fix_rows()
    fixed_withdrawals = 0
    cleaned_withdraw_entries = 0
    created_entries = []
    errors = []

    for row in rows:
        enrollment = row['enrollment']
        success = False
        last_exc = None
        for attempt in range(3):
            try:
                close_old_connections()
                _configure_sqlite_busy_timeout()
                with transaction.atomic():
                    if enrollment.is_completed and row['retained_amount'] > 0:
                        _adjust_quick_receipts_for_refund(
                            enrollment.student,
                            enrollment,
                            row['retained_amount'],
                        )

                    cleanup_result = _cleanup_quick_withdrawal_entries(enrollment, user)
                    cleaned_withdraw_entries += (
                        len(cleanup_result['reversed_ids']) +
                        len(cleanup_result['deleted_ids'])
                    )

                    if enrollment.is_completed:
                        entries = _build_quick_withdrawal_entry(
                            enrollment=enrollment,
                            user=user,
                            refunded_amount=row['retained_amount'],
                            description=(
                                f"[QUICK_WITHDRAW #{enrollment.id}] "
                                f"تصحيح سحب طالب سريع - {enrollment.student.full_name} - {enrollment.course.name}"
                            ),
                        )
                        created_entries.extend(entry.id for entry in entries)
                        fixed_withdrawals += len(entries)
                success = True
                connection.close()
                break
            except OperationalError as exc:
                last_exc = exc
                if 'database is locked' not in str(exc).lower():
                    break
                close_old_connections()
                connection.close()
                time.sleep(1.5 * (attempt + 1))
            except Exception as exc:
                last_exc = exc
                break

        if not success and last_exc:
            errors.append(f'{enrollment.student.full_name} / {enrollment.course.name}: {last_exc}')

    if cleaned_withdraw_entries:
        Account.rebuild_all_balances()

    deactivated_legacy_accounts = _deactivate_quick_legacy_withdrawal_accounts()

    return {
        'rows_count': len(rows),
        'fixed_withdrawals': fixed_withdrawals,
        'cleaned_withdraw_entries': cleaned_withdraw_entries,
        'deactivated_legacy_accounts': deactivated_legacy_accounts,
        'created_entries': created_entries,
        'errors': errors,
    }


def _append_quick_course_statement_rows(rows, course_name, student_name, student_phone, source_label, entry):
    transactions = list(entry.transactions.all())
    transactions.sort(key=lambda transaction: (transaction.is_debit, transaction.id))

    created_by = "-"
    if entry.created_by:
        created_by = entry.created_by.get_full_name() or entry.created_by.username or "-"

    posted_by = "-"
    if entry.posted_by:
        posted_by = entry.posted_by.get_full_name() or entry.posted_by.username or "-"

    for transaction in transactions:
        account_name = transaction.account.name_ar or transaction.account.name
        rows.append({
            'course_name': course_name,
            'student_name': student_name or "-",
            'student_phone': student_phone or "-",
            'entry_reference': entry.reference,
            'entry_date': entry.date.strftime('%Y-%m-%d') if entry.date else "-",
            'entry_type': entry.get_entry_type_display(),
            'entry_source': source_label,
            'entry_description': entry.description,
            'account_code': transaction.account.code,
            'account_name': account_name,
            'transaction_description': transaction.description or entry.description,
            'debit': transaction.amount if transaction.is_debit else Decimal('0'),
            'credit': transaction.amount if not transaction.is_debit else Decimal('0'),
            'entry_total': entry.total_amount or Decimal('0'),
            'posted_status': 'مرحل' if entry.is_posted else 'غير مرحل',
            'created_by': created_by,
            'posted_by': posted_by,
        })


def _extract_quick_student_id_from_entry(entry):
    for transaction in entry.transactions.select_related('account').all():
        code = (transaction.account.code or '').strip()
        if not code.startswith('1252-'):
            continue
        parts = code.split('-')
        if len(parts) < 2:
            continue
        try:
            return int(parts[1])
        except (TypeError, ValueError):
            continue
    return None


def _resolve_quick_entry_source(entry, receipt=None, enrollment=None):
    if receipt:
        return 'قيد قبض'
    if enrollment:
        return 'قيد تسجيل'

    description = (entry.description or '').lower()
    if 'quick_withdraw' in description or 'سحب' in description:
        return 'قيد سحب'
    if 'استرداد' in description:
        return 'قيد استرداد'
    if entry.entry_type == 'ADJUSTMENT':
        return 'قيد تسوية'
    return entry.get_entry_type_display()


def _build_quick_course_statement_rows(courses):
    courses = list(courses)
    rows_by_course = defaultdict(list)

    if not courses:
        return rows_by_course

    course_map = {course.id: course for course in courses}
    course_account_codes = {
        course.id: {f'2151-{course.id:03d}', f'4111-{course.id:03d}'}
        for course in courses
    }
    all_course_account_codes = {
        code for codes in course_account_codes.values() for code in codes
    }

    enrollments = list(
        QuickEnrollment.objects.filter(course__in=courses)
        .select_related('student', 'course')
        .order_by('course__name', 'student__full_name', 'id')
    )
    enrollment_ref_map = {f"QE-{enrollment.id}": enrollment for enrollment in enrollments}
    enrollment_by_course = defaultdict(list)
    for enrollment in enrollments:
        enrollment_by_course[enrollment.course_id].append(enrollment)

    receipts = list(
        QuickStudentReceipt.objects.filter(course__in=courses, journal_entry__isnull=False)
        .select_related('quick_student', 'course')
        .order_by('course__name', 'date', 'id')
    )
    receipt_by_entry_id = {
        receipt.journal_entry_id: receipt
        for receipt in receipts
        if receipt.journal_entry_id
    }
    receipt_entry_ids = set(receipt_by_entry_id.keys())

    journal_entries = list(
        JournalEntry.objects.filter(
            Q(reference__in=enrollment_ref_map.keys())
            | Q(id__in=receipt_entry_ids)
            | Q(transactions__account__code__in=all_course_account_codes)
        ).distinct().select_related(
            'created_by', 'posted_by'
        ).prefetch_related(
            Prefetch('transactions', queryset=Transaction.objects.select_related('account').order_by('id'))
        ).order_by('date', 'id')
    )

    added_entry_ids = defaultdict(set)
    quick_student_cache = {}

    for entry in journal_entries:
        entry_course_ids = set()

        receipt = receipt_by_entry_id.get(entry.id)
        if receipt and receipt.course_id in course_map:
            entry_course_ids.add(receipt.course_id)

        enrollment = enrollment_ref_map.get(entry.reference)
        if enrollment and enrollment.course_id in course_map:
            entry_course_ids.add(enrollment.course_id)

        for transaction in entry.transactions.all():
            code = (transaction.account.code or '').strip()
            for course_id, account_codes in course_account_codes.items():
                if code in account_codes:
                    entry_course_ids.add(course_id)

        if not entry_course_ids:
            continue

        student_obj = None
        if receipt and getattr(receipt, 'quick_student', None):
            student_obj = receipt.quick_student
        elif enrollment and getattr(enrollment, 'student', None):
            student_obj = enrollment.student
        else:
            quick_student_id = _extract_quick_student_id_from_entry(entry)
            if quick_student_id:
                if quick_student_id not in quick_student_cache:
                    quick_student_cache[quick_student_id] = QuickStudent.objects.filter(
                        id=quick_student_id
                    ).first()
                student_obj = quick_student_cache.get(quick_student_id)

        for course_id in sorted(entry_course_ids):
            if entry.id in added_entry_ids[course_id]:
                continue

            course = course_map.get(course_id)
            linked_enrollment = enrollment
            if not linked_enrollment and student_obj:
                linked_enrollment = next(
                    (
                        item for item in enrollment_by_course.get(course_id, [])
                        if item.student_id == getattr(student_obj, 'id', None)
                    ),
                    None,
                )

            student_name = '-'
            student_phone = '-'
            if receipt:
                student_name = receipt.student_name or getattr(receipt.quick_student, 'full_name', '-') or '-'
                student_phone = getattr(receipt.quick_student, 'phone', '-') or '-'
            elif student_obj:
                student_name = getattr(student_obj, 'full_name', '-') or '-'
                student_phone = getattr(student_obj, 'phone', '-') or '-'
            elif linked_enrollment:
                student_name = linked_enrollment.student.full_name or '-'
                student_phone = linked_enrollment.student.phone or '-'

            _append_quick_course_statement_rows(
                rows_by_course[course_id],
                course_name=course.name if course else '-',
                student_name=student_name,
                student_phone=student_phone,
                source_label=_resolve_quick_entry_source(
                    entry,
                    receipt=receipt if receipt and receipt.course_id == course_id else None,
                    enrollment=linked_enrollment if linked_enrollment and linked_enrollment.course_id == course_id else None,
                ),
                entry=entry,
            )
            added_entry_ids[course_id].add(entry.id)

    return rows_by_course


def export_quick_course_statement_excel(request):
    """Export quick course journal entries with student names per course."""
    course_type, _, report_label = _get_outstanding_course_type(request)
    academic_year_id = request.GET.get('academic_year')

    courses_qs = QuickCourse.objects.filter(is_active=True).select_related('academic_year').order_by('name')
    if course_type != 'ALL':
        courses_qs = courses_qs.filter(course_type=course_type)
    if academic_year_id:
        courses_qs = courses_qs.filter(academic_year_id=academic_year_id)

    courses = list(courses_qs)
    rows_by_course = _build_quick_course_statement_rows(courses)

    workbook = Workbook()
    workbook.remove(workbook.active)

    title_font = Font(bold=True, size=16, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    normal_font = Font(size=11)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    subheader_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    columns = [
        ("#", 6),
        ("الدورة", 24),
        ("الطالب", 24),
        ("الهاتف", 16),
        ("مصدر القيد", 14),
        ("رقم القيد", 16),
        ("تاريخ القيد", 14),
        ("نوع القيد", 18),
        ("بيان القيد", 34),
        ("رمز الحساب", 14),
        ("اسم الحساب", 24),
        ("بيان الحركة", 34),
        ("مدين", 14),
        ("دائن", 14),
        ("إجمالي القيد", 14),
        ("الحالة", 12),
        ("أنشئ بواسطة", 18),
        ("رُحّل بواسطة", 18),
    ]

    def write_sheet(ws, title, rows, include_course_name):
        ws.sheet_view.rightToLeft = True
        visible_columns = columns if include_course_name else [col for col in columns if col[0] != "الدورة"]
        total_cols = len(visible_columns)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value="كشف حساب الدورات السريعة")
        title_cell.font = title_font
        title_cell.alignment = center
        title_cell.fill = header_fill

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        meta_cell = ws.cell(row=2, column=1, value=f"الدورة/التصنيف: {title} | عدد الحركات: {len(rows)}")
        meta_cell.alignment = right
        meta_cell.fill = subheader_fill

        for col_idx, (label, width) in enumerate(visible_columns, start=1):
            cell = ws.cell(row=4, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, row in enumerate(rows, start=5):
            values = [
                row_idx - 4,
                row['course_name'],
                row['student_name'],
                row['student_phone'],
                row['entry_source'],
                row['entry_reference'],
                row['entry_date'],
                row['entry_type'],
                row['entry_description'],
                row['account_code'],
                row['account_name'],
                row['transaction_description'],
                row['debit'],
                row['credit'],
                row['entry_total'],
                row['posted_status'],
                row['created_by'],
                row['posted_by'],
            ]
            if not include_course_name:
                values.pop(1)

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = border
                cell.alignment = center if col_idx in (1, 6, 7) else right
                header_label = visible_columns[col_idx - 1][0]
                if header_label in {"مدين", "دائن", "إجمالي القيد"}:
                    cell.number_format = '#,##0.00'

        ws.freeze_panes = 'A5'

    combined_rows = []
    for course in courses:
        combined_rows.extend(rows_by_course.get(course.id, []))

    all_sheet = workbook.create_sheet("كل الدورات")
    write_sheet(all_sheet, report_label, combined_rows, include_course_name=True)

    existing_titles = {all_sheet.title}
    for course in courses:
        sheet_name = _safe_sheet_title(course.name, existing_titles)
        existing_titles.add(sheet_name)
        course_sheet = workbook.create_sheet(sheet_name)
        write_sheet(course_sheet, course.name, rows_by_course.get(course.id, []), include_course_name=False)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    timestamp = timezone.now().strftime('%Y%m%d_%H%M')
    response['Content-Disposition'] = f'attachment; filename="كشف_حساب_الدورات_السريعة_{report_label}_{timestamp}.xlsx"'
    workbook.save(response)
    return response


def export_quick_outstanding_excel(request):
    """Export quick courses outstanding report with per-course sheets."""
    course_type, _, report_label = _get_outstanding_course_type(request)
    academic_year_id = request.GET.get('academic_year')
    courses_qs = QuickCourse.objects.filter(is_active=True).select_related('academic_year').order_by('name')
    if course_type != 'ALL':
        courses_qs = courses_qs.filter(course_type=course_type)
    if academic_year_id:
        courses_qs = courses_qs.filter(academic_year_id=academic_year_id)
    start_date, end_date = _get_outstanding_date_range(request)
    course_data, _ = _build_quick_outstanding_course_summary(
        courses_qs,
        include_zero_outstanding=False,
        start_date=start_date,
        end_date=end_date,
    )
    courses = [row['course'] for row in course_data]

    enrollments = list((
        QuickEnrollment.objects
        .filter(course__in=courses, is_completed=False)
        .select_related('student', 'course', 'student__created_by', 'student__student', 'course__academic_year')
        .order_by('course__name', 'student__full_name')
    ))

    receipt_qs = QuickStudentReceipt.objects.filter(course__in=courses)

    paid_map = {}
    receipt_totals = receipt_qs.values('quick_student_id', 'course_id').annotate(
        total=Sum('paid_amount'),
        count=Count('id'),
        last=Max('date'),
    )
    for row in receipt_totals:
        paid_map[(row['quick_student_id'], row['course_id'])] = {
            'total': row['total'] or Decimal('0'),
            'count': row['count'] or 0,
            'last': row['last'],
        }

    paid_by_enrollment = {}
    receipt_by_enrollment = receipt_qs.exclude(quick_enrollment_id__isnull=True).values(
        'quick_enrollment_id'
    ).annotate(
        total=Sum('paid_amount'),
        count=Count('id'),
        last=Max('date'),
    )
    for row in receipt_by_enrollment:
        paid_by_enrollment[row['quick_enrollment_id']] = {
            'total': row['total'] or Decimal('0'),
            'count': row['count'] or 0,
            'last': row['last'],
        }

    regular_phone_set = _build_regular_phone_set()

    def student_type_label(quick_student):
        phone = _normalize_phone(quick_student.phone)
        return "طالب معهد" if phone and phone in regular_phone_set else "خارجي"

    def registered_by_label(quick_student):
        user = quick_student.created_by
        if not user:
            return "-"
        return user.get_full_name() or user.username or "-"

    workbook = Workbook()
    workbook.remove(workbook.active)

    title_font = Font(bold=True, size=16, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    normal_font = Font(size=11)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    subheader_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_sheet(ws, course_label, rows, include_course_col):
        ws.sheet_view.rightToLeft = True
        columns = [
            ("#", 6),
            ("اسم الطالب", 28),
            ("رقم الطالب", 16),
            ("رقم الهاتف", 16),
            ("نوع الطالب", 14),
            ("الحالة", 14),
            ("المسجل", 18),
            ("الفصل الدراسي", 18),
            ("نوع الدورة", 16),
            ("تاريخ التسجيل", 14),
        ]
        if include_course_col:
            columns.insert(1, ("الدورة", 26))
        columns.extend([
            ("إجمالي قبل الخصم", 16),
            ("نسبة الخصم %", 12),
            ("قيمة الخصم", 14),
            ("الصافي", 14),
            ("المدفوع", 14),
            ("المتبقي", 14),
            ("عدد الإيصالات", 12),
            ("آخر دفعة", 14),
        ])
        money_labels = {"إجمالي قبل الخصم", "قيمة الخصم", "الصافي", "المدفوع", "المتبقي"}
        percent_labels = {"نسبة الخصم %"}
        count_labels = {"عدد الإيصالات"}
        label_by_col = {idx + 1: label for idx, (label, _) in enumerate(columns)}

        total_cols = len(columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws.cell(row=1, column=1, value="تقرير المستحقات - الدورات السريعة").font = title_font
        ws.cell(row=1, column=1).alignment = center
        ws.cell(row=1, column=1).fill = header_fill

        internal_count = sum(1 for r in rows if r['student_type'] == "طالب معهد")
        external_count = sum(1 for r in rows if r['student_type'] == "خارجي")
        total_paid = sum(r['paid'] for r in rows)
        total_remaining = sum(r['remaining'] for r in rows)
        total_net = sum(r['net_amount'] for r in rows)
        total_receipts = sum(r['receipts_count'] for r in rows)
        paid_count = sum(1 for r in rows if r['payment_status'] == "مسدد")
        outstanding_count = len(rows) - paid_count

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        ws.cell(
            row=2,
            column=1,
            value=f"الدورة: {course_label} | طالب معهد: {internal_count} | خارجي: {external_count} | مسدد: {paid_count} | غير مسدد: {outstanding_count}"
        ).alignment = right
        ws.cell(row=2, column=1).fill = subheader_fill

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
        ws.cell(
            row=3,
            column=1,
            value=(
                f"إجمالي الطلاب: {len(rows)} | إجمالي الصافي: {total_net} | "
                f"إجمالي المدفوع: {total_paid} | إجمالي المتبقي: {total_remaining} | "
                f"عدد الإيصالات: {total_receipts}"
            )
        ).alignment = right
        ws.cell(row=3, column=1).fill = subheader_fill

        for col_idx, (label, width) in enumerate(columns, start=1):
            cell = ws.cell(row=4, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        row_idx = 5
        for idx, row in enumerate(rows, start=1):
            values = [
                idx,
                row['student_name'],
                row['student_number'],
                row['phone'],
                row['student_type'],
                row['payment_status'],
                row['registered_by'],
                row['academic_year'],
                row['course_type'],
                row['enrollment_date'],
            ]
            if include_course_col:
                values.insert(1, row['course_name'])
            values.extend([
                row['total_amount'],
                row['discount_percent'],
                row['discount_amount'],
                row['net_amount'],
                row['paid'],
                row['remaining'],
                row['receipts_count'],
                row['last_payment_date'],
            ])

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = border
                if col_idx in (1,):
                    cell.alignment = center
                else:
                    cell.alignment = right
                label = label_by_col.get(col_idx)
                if label in money_labels:
                    cell.number_format = '#,##0'
                elif label in percent_labels:
                    cell.number_format = '0.00'
                elif label in count_labels:
                    cell.number_format = '0'
            row_idx += 1

    def build_rows(enrollments):
        rows = []
        for enrollment in enrollments:
            student = enrollment.student
            enrollment_stats = paid_by_enrollment.get(enrollment.id)
            course_stats = paid_map.get((student.id, enrollment.course_id), None)

            paid = _format_money(
                enrollment_stats['total'] if enrollment_stats else (
                    course_stats['total'] if course_stats else Decimal('0')
                )
            )
            receipts_count = (
                enrollment_stats['count'] if enrollment_stats else (
                    course_stats['count'] if course_stats else 0
                )
            )
            last_payment = (
                enrollment_stats['last'] if enrollment_stats else (
                    course_stats['last'] if course_stats else None
                )
            )

            total_amount = _format_money(enrollment.total_amount or Decimal('0'))
            discount_percent = _format_money(enrollment.discount_percent or Decimal('0'))
            discount_amount = _format_money(enrollment.discount_amount or Decimal('0'))
            net_amount = _format_money(enrollment.net_amount or Decimal('0'))
            remaining = max(Decimal('0'), net_amount - paid)
            course = enrollment.course
            academic_year = course.academic_year.name if course and course.academic_year else "-"
            course_type_label = course.get_course_type_display() if course else "-"
            student_number = "-"
            if getattr(student, 'student', None):
                student_number = getattr(student.student, 'student_number', None) or "-"
            elif getattr(student, 'student_number', None):
                student_number = student.student_number or "-"
            rows.append({
                'course_name': enrollment.course.name,
                'student_name': student.full_name,
                'student_number': student_number,
                'phone': student.phone or "-",
                'student_type': student_type_label(student),
                'payment_status': 'غير مسدد' if remaining > 0 else 'مسدد',
                'registered_by': registered_by_label(student),
                'enrollment_date': enrollment.enrollment_date.strftime('%Y-%m-%d') if enrollment.enrollment_date else "-",
                'academic_year': academic_year,
                'course_type': course_type_label,
                'total_amount': total_amount,
                'discount_percent': discount_percent,
                'discount_amount': discount_amount,
                'net_amount': net_amount,
                'paid': paid,
                'remaining': remaining,
                'receipts_count': receipts_count,
                'last_payment_date': last_payment.strftime('%Y-%m-%d') if last_payment else "-",
            })
        return rows

    all_rows = build_rows(enrollments)
    all_sheet = workbook.create_sheet("كل الدورات")
    write_sheet(all_sheet, "كل الدورات", all_rows, include_course_col=True)

    existing_titles = {all_sheet.title}
    for course in courses:
        course_enrollments = [e for e in enrollments if e.course_id == course.id]
        course_rows = build_rows(course_enrollments)
        if not course_rows:
            continue
        sheet_name = _safe_sheet_title(course.name, existing_titles)
        existing_titles.add(sheet_name)
        ws = workbook.create_sheet(sheet_name)
        write_sheet(ws, course.name, course_rows, include_course_col=False)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    timestamp = timezone.now().strftime('%Y%m%d_%H%M')
    response['Content-Disposition'] = f'attachment; filename="تقرير_المستحقات_الدورات_السريعة_{report_label}_{timestamp}.xlsx"'
    workbook.save(response)
    return response

# ------------------------------
# Quick outstanding helpers
def _get_outstanding_course_type(request, default='INTENSIVE'):
    course_type = request.GET.get('course_type') or default
    valid_course_types = {value for value, _ in QuickCourse.COURSE_TYPE_CHOICES}
    if course_type != 'ALL' and course_type not in valid_course_types:
        course_type = default if default == 'ALL' or default in valid_course_types else 'INTENSIVE'

    label_map = {value: label for value, label in QuickCourse.COURSE_TYPE_CHOICES}
    report_label_map = {
        'INTENSIVE': 'المكثفات',
        'EXAM': 'الامتحانيات',
        'REGULAR': 'العادية',
        'WEEKEND': 'نهاية الأسبوع',
    }

    if course_type == 'ALL':
        label = 'كل الدورات'
        report_label = 'كل الدورات'
    else:
        label = label_map.get(course_type, course_type)
        report_label = report_label_map.get(course_type, label)

    return course_type, label, report_label


def _get_outstanding_date_range(request):
    start_date = parse_date(request.GET.get('start_date') or '')
    end_date = parse_date(request.GET.get('end_date') or '')

    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date

    return start_date, end_date


def _get_course_type_options():
    options = [{'value': 'ALL', 'label': 'كل الدورات'}]
    for value, label in QuickCourse.COURSE_TYPE_CHOICES:
        options.append({'value': value, 'label': label})
    return options


def _get_selected_quick_course_ids(request):
    selected_ids = []
    seen_ids = set()
    for raw_value in request.GET.getlist('course_ids'):
        try:
            course_id = int(raw_value)
        except (TypeError, ValueError):
            continue
        if course_id in seen_ids:
            continue
        selected_ids.append(course_id)
        seen_ids.add(course_id)
    return selected_ids


def _build_quick_report_course_filters(request, default_course_type='ALL'):
    course_type, course_type_label, report_label = _get_outstanding_course_type(
        request,
        default=default_course_type,
    )
    search_query = (request.GET.get('q') or '').strip()
    selected_course_ids = _get_selected_quick_course_ids(request)

    academic_year_raw = request.GET.get('academic_year')
    try:
        academic_year_id = int(academic_year_raw) if academic_year_raw else None
    except (TypeError, ValueError):
        academic_year_id = None

    base_courses = QuickCourse.objects.filter(is_active=True).select_related('academic_year')
    if course_type != 'ALL':
        base_courses = base_courses.filter(course_type=course_type)
    if academic_year_id:
        base_courses = base_courses.filter(academic_year_id=academic_year_id)

    available_courses = list(base_courses.order_by('-academic_year__start_date', 'name', 'id'))

    filtered_courses = base_courses
    if selected_course_ids:
        filtered_courses = filtered_courses.filter(id__in=selected_course_ids)
    if search_query:
        filtered_courses = filtered_courses.filter(
            Q(name__icontains=search_query) |
            Q(name_ar__icontains=search_query) |
            Q(sessions__title__icontains=search_query) |
            Q(sessions__code__icontains=search_query) |
            Q(sessions__room_name__icontains=search_query) |
            Q(sessions__room__name__icontains=search_query) |
            Q(enrollments__student__full_name__icontains=search_query) |
            Q(enrollments__student__phone__icontains=search_query)
        )

    filtered_courses = filtered_courses.order_by('-academic_year__start_date', 'name', 'id').distinct()

    return {
        'course_type': course_type,
        'course_type_label': course_type_label,
        'course_type_report_label': report_label,
        'course_type_options': _get_course_type_options(),
        'academic_year_id': academic_year_id,
        'academic_years': AcademicYear.objects.all().order_by('-start_date'),
        'search_query': search_query,
        'selected_course_ids': selected_course_ids,
        'available_courses': available_courses,
        'courses': filtered_courses,
    }


def _build_quick_session_population_report(request):
    filters = _build_quick_report_course_filters(request, default_course_type='ALL')
    courses = list(filters['courses'])
    today = timezone.localdate()

    report = {
        **filters,
        'course_rows': [],
        'total_courses': len(courses),
        'total_sessions': 0,
        'total_enrollments': 0,
        'total_assigned_students': 0,
        'total_unassigned_students': 0,
        'assignment_rate': 0,
        'courses_with_unassigned': 0,
        'largest_session': None,
        'generated_at': timezone.localtime(),
        'today': today,
    }

    if not courses:
        return report

    course_ids = [course.id for course in courses]
    enrollment_stats = {
        row['course_id']: row
        for row in (
            QuickEnrollment.objects.filter(
                course_id__in=course_ids,
                is_completed=False,
                student__is_active=True,
            )
            .values('course_id')
            .annotate(
                total_enrollments=Count('id', distinct=True),
                assigned_enrollments=Count(
                    'id',
                    filter=Q(session_assignment__session__is_active=True),
                    distinct=True,
                ),
            )
        )
    }

    sessions = list(
        QuickCourseSession.objects.filter(course_id__in=course_ids, is_active=True)
        .select_related('course', 'room', 'time_option')
        .annotate(
            assigned_count=Count(
                'session_enrollments',
                filter=Q(
                    session_enrollments__enrollment__is_completed=False,
                    session_enrollments__enrollment__student__is_active=True,
                ),
                distinct=True,
            )
        )
        .order_by('course__name', 'start_date', 'start_time', 'title', 'id')
    )

    sessions_by_course = defaultdict(list)
    largest_session = None
    total_sessions = 0

    for session in sessions:
        total_sessions += 1
        if today < session.start_date:
            lifecycle = 'UPCOMING'
            lifecycle_label = 'قادم'
        elif today > session.end_date:
            lifecycle = 'FINISHED'
            lifecycle_label = 'منتهي'
        else:
            lifecycle = 'LIVE'
            lifecycle_label = 'يعمل الآن'

        capacity = session.capacity or 0
        assigned_count = session.assigned_count or 0
        utilization = round((assigned_count / capacity) * 100) if capacity else 0
        session_row = {
            'session': session,
            'assigned_count': assigned_count,
            'capacity': capacity,
            'available_seats': max(0, capacity - assigned_count) if capacity else None,
            'seat_utilization': utilization,
            'lifecycle': lifecycle,
            'lifecycle_label': lifecycle_label,
        }
        sessions_by_course[session.course_id].append(session_row)

        candidate_key = (assigned_count, utilization, session.start_date, session.id)
        if largest_session is None or candidate_key > largest_session['sort_key']:
            largest_session = {
                'session': session,
                'assigned_count': assigned_count,
                'seat_utilization': utilization,
                'sort_key': candidate_key,
            }

    course_rows = []
    total_enrollments = 0
    total_assigned_students = 0
    total_unassigned_students = 0

    for course in courses:
        stats = enrollment_stats.get(course.id, {})
        total_course_enrollments = stats.get('total_enrollments', 0) or 0
        assigned_course_enrollments = stats.get('assigned_enrollments', 0) or 0
        unassigned_course_enrollments = max(0, total_course_enrollments - assigned_course_enrollments)
        session_rows = sessions_by_course.get(course.id, [])
        capacity_total = sum(item['capacity'] for item in session_rows if item['capacity'])
        assigned_via_sessions = sum(item['assigned_count'] for item in session_rows)

        course_rows.append({
            'course': course,
            'sessions': session_rows,
            'sessions_count': len(session_rows),
            'total_enrollments': total_course_enrollments,
            'assigned_students': assigned_course_enrollments,
            'assigned_via_sessions': assigned_via_sessions,
            'unassigned_students': unassigned_course_enrollments,
            'assignment_rate': round((assigned_course_enrollments / total_course_enrollments) * 100) if total_course_enrollments else 0,
            'capacity_total': capacity_total,
            'available_seats_total': max(0, capacity_total - assigned_via_sessions) if capacity_total else None,
            'has_unassigned_students': unassigned_course_enrollments > 0,
        })

        total_enrollments += total_course_enrollments
        total_assigned_students += assigned_course_enrollments
        total_unassigned_students += unassigned_course_enrollments

    report.update({
        'course_rows': course_rows,
        'total_courses': len(course_rows),
        'total_sessions': total_sessions,
        'total_enrollments': total_enrollments,
        'total_assigned_students': total_assigned_students,
        'total_unassigned_students': total_unassigned_students,
        'assignment_rate': round((total_assigned_students / total_enrollments) * 100) if total_enrollments else 0,
        'courses_with_unassigned': sum(1 for row in course_rows if row['has_unassigned_students']),
        'largest_session': largest_session,
    })
    return report


def _build_quick_free_students_report(request):
    filters = _build_quick_report_course_filters(request, default_course_type='ALL')
    courses = list(filters['courses'])

    report = {
        **filters,
        'course_rows': [],
        'total_courses_with_discount': 0,
        'total_discounted_students': 0,
        'total_discount_value': Decimal('0'),
        'average_discount_value': Decimal('0'),
        'largest_discount_course': None,
        'generated_at': timezone.localtime(),
    }

    if not courses:
        return report

    course_ids = [course.id for course in courses]
    enrollments = list(
        QuickEnrollment.objects.filter(
            course_id__in=course_ids,
            is_completed=False,
            student__is_active=True,
        )
        .select_related('course', 'course__academic_year', 'student', 'student__student')
        .order_by('course__name', 'student__full_name', 'id')
    )

    if not enrollments:
        return report

    receipt_totals = {
        row['quick_enrollment_id']: row['total'] or Decimal('0')
        for row in (
            QuickStudentReceipt.objects.filter(
                quick_enrollment_id__in=[enrollment.id for enrollment in enrollments]
            )
            .values('quick_enrollment_id')
            .annotate(total=Sum('paid_amount'))
        )
    }

    course_rows_map = {}
    total_discounted_students = 0
    total_discount_value = Decimal('0')

    for enrollment in enrollments:
        total_amount = enrollment.total_amount or enrollment.course.price or Decimal('0')
        net_amount = enrollment.net_amount or Decimal('0')
        discount_percent = enrollment.discount_percent or Decimal('0')
        discount_amount = enrollment.discount_amount or Decimal('0')
        discount_value = max(Decimal('0'), total_amount - net_amount)
        has_discount = (
            discount_percent > Decimal('0')
            or discount_amount > Decimal('0')
            or discount_value > Decimal('0')
        )
        if not has_discount:
            continue

        percent_discount_value = max(
            Decimal('0'),
            total_amount * (discount_percent / Decimal('100'))
        ) if total_amount > 0 and discount_percent > 0 else Decimal('0')
        fixed_discount_value = max(Decimal('0'), discount_amount)
        if discount_value <= 0:
            discount_value = percent_discount_value + fixed_discount_value

        if discount_percent > 0 and fixed_discount_value > 0:
            discount_label = 'نسبة + قيمة'
        elif discount_percent > 0:
            discount_label = 'نسبة'
        elif fixed_discount_value > 0:
            discount_label = 'قيمة'
        else:
            discount_label = 'فرق صافي'

        course_entry = course_rows_map.setdefault(
            enrollment.course_id,
            {
                'course': enrollment.course,
                'students': [],
                'discounted_students_count': 0,
                'total_discount_value': Decimal('0'),
                'students_with_payments': 0,
                'full_free_count': 0,
            },
        )

        paid_total = receipt_totals.get(enrollment.id, Decimal('0'))
        is_full_free = (
            discount_percent >= Decimal('100')
            or (total_amount > 0 and net_amount <= Decimal('0'))
            or (total_amount > 0 and discount_value >= total_amount)
        )

        course_entry['students'].append({
            'student': enrollment.student,
            'enrollment': enrollment,
            'total_amount': total_amount,
            'net_amount': net_amount,
            'discount_value': discount_value,
            'waived_amount': discount_value,
            'discount_percent': discount_percent,
            'discount_amount': discount_amount,
            'percent_discount_value': percent_discount_value,
            'fixed_discount_value': fixed_discount_value,
            'discount_label': discount_label,
            'paid_total': paid_total,
            'has_payments': paid_total > 0,
            'is_full_free': is_full_free,
        })
        course_entry['discounted_students_count'] += 1
        course_entry['total_discount_value'] += discount_value
        if paid_total > 0:
            course_entry['students_with_payments'] += 1
        if is_full_free:
            course_entry['full_free_count'] += 1

        total_discounted_students += 1
        total_discount_value += discount_value

    course_rows = []
    largest_discount_course = None
    for course in courses:
        course_entry = course_rows_map.get(course.id)
        if not course_entry:
            continue
        course_entry['students'].sort(
            key=lambda item: (
                -(item['discount_value'] or Decimal('0')),
                item['student'].full_name,
                item['student'].id,
            )
        )
        course_rows.append(course_entry)
        candidate_key = (
            course_entry['discounted_students_count'],
            course_entry['total_discount_value'],
            course.id,
        )
        if largest_discount_course is None or candidate_key > largest_discount_course['sort_key']:
            largest_discount_course = {
                'course': course_entry['course'],
                'discounted_students_count': course_entry['discounted_students_count'],
                'total_discount_value': course_entry['total_discount_value'],
                'sort_key': candidate_key,
            }

    report.update({
        'course_rows': course_rows,
        'total_courses_with_discount': len(course_rows),
        'total_discounted_students': total_discounted_students,
        'total_discount_value': total_discount_value,
        'average_discount_value': (total_discount_value / total_discounted_students) if total_discounted_students else Decimal('0'),
        'largest_discount_course': largest_discount_course,
        'total_courses_with_free': len(course_rows),
        'total_free_students': total_discounted_students,
        'total_waived_amount': total_discount_value,
        'average_waived_amount': (total_discount_value / total_discounted_students) if total_discounted_students else Decimal('0'),
        'largest_free_course': largest_discount_course,
    })
    return report


def _attach_quick_report_urls(context, request, report_url_name, print_url_name):
    query_string = request.GET.urlencode()
    report_url = reverse(report_url_name)
    print_url = reverse(print_url_name)
    if query_string:
        report_url = f'{report_url}?{query_string}'
        print_url = f'{print_url}?{query_string}'
    context.update({
        'report_url': report_url,
        'print_url': print_url,
    })
    return context


def _build_quick_outstanding_course_summary(courses, include_zero_outstanding=False, start_date=None, end_date=None):
    courses = list(courses)
    course_map = {
        course.id: {
            'course': course,
            'total_students': 0,
            'male_students': 0,
            'female_students': 0,
            'unknown_students': 0,
            'paid_students': 0,
            'outstanding_students': 0,
            'total_outstanding': Decimal('0.00'),
            'total_paid': Decimal('0.00'),
        }
        for course in courses
    }

    enrollments_qs = QuickEnrollment.objects.filter(course__in=courses, is_completed=False)
    if start_date:
        enrollments_qs = enrollments_qs.filter(enrollment_date__gte=start_date)
    if end_date:
        enrollments_qs = enrollments_qs.filter(enrollment_date__lte=end_date)

    enrollments = list(enrollments_qs.select_related('course', 'student__student'))
    if enrollments:
        receipt_totals = QuickStudentReceipt.objects.filter(
            quick_enrollment_id__in=[enrollment.id for enrollment in enrollments]
        ).values('quick_enrollment_id').annotate(total=Sum('paid_amount'))
        paid_map = {
            row['quick_enrollment_id']: (row['total'] or Decimal('0'))
            for row in receipt_totals
        }
    else:
        paid_map = {}

    for enrollment in enrollments:
        net_amount = enrollment.net_amount or Decimal('0')
        paid_total = paid_map.get(enrollment.id, Decimal('0'))
        remaining = max(Decimal('0'), net_amount - paid_total)

        stats = course_map.get(enrollment.course_id)
        if not stats:
            continue

        stats['total_students'] += 1
        student_gender = getattr(getattr(enrollment.student, 'student', None), 'gender', None)
        if student_gender == 'male':
            stats['male_students'] += 1
        elif student_gender == 'female':
            stats['female_students'] += 1
        else:
            stats['unknown_students'] += 1
        if remaining > 0:
            stats['outstanding_students'] += 1
            stats['total_outstanding'] += remaining
        else:
            stats['paid_students'] += 1
        stats['total_paid'] += paid_total

    course_data = list(course_map.values())
    if not include_zero_outstanding:
        course_data = [row for row in course_data if row['outstanding_students'] > 0]

    totals = {
        'total_courses': len(course_data),
        'total_male_students': sum(row['male_students'] for row in course_data),
        'total_female_students': sum(row['female_students'] for row in course_data),
        'total_unknown_students': sum(row['unknown_students'] for row in course_data),
        'total_outstanding_students': sum(row['outstanding_students'] for row in course_data),
        'total_outstanding_amount': sum(row['total_outstanding'] for row in course_data),
        'total_paid_students': sum(row['paid_students'] for row in course_data),
        'total_students': sum(row['total_students'] for row in course_data),
        'total_paid_amount': sum(row['total_paid'] for row in course_data),
    }

    return course_data, totals


def _build_quick_outstanding_rows(courses, start_date=None, end_date=None):
    courses = list(courses)
    if not courses:
        return [], {'grouped_courses': [], 'totals': {'total_students': 0, 'total_courses': 0, 'total_outstanding': Decimal('0')}}

    enrollments_qs = QuickEnrollment.objects.filter(course__in=courses, is_completed=False)
    if start_date:
        enrollments_qs = enrollments_qs.filter(enrollment_date__gte=start_date)
    if end_date:
        enrollments_qs = enrollments_qs.filter(enrollment_date__lte=end_date)

    enrollments = list(
        enrollments_qs.select_related('course', 'student').order_by(
            'enrollment_date', 'course__name', 'student__full_name', 'id'
        )
    )

    paid_map = {}
    if enrollments:
        paid_rows = QuickStudentReceipt.objects.filter(
            quick_enrollment_id__in=[enrollment.id for enrollment in enrollments]
        ).values('quick_enrollment_id').annotate(total=Sum('paid_amount'))
        paid_map = {
            row['quick_enrollment_id']: (row['total'] or Decimal('0'))
            for row in paid_rows
        }

    today = timezone.localdate()
    rows = []
    grouped_map = {}

    for enrollment in enrollments:
        net_amount = enrollment.net_amount or Decimal('0')
        paid_total = paid_map.get(enrollment.id, Decimal('0'))
        remaining = max(Decimal('0'), net_amount - paid_total)
        if remaining <= 0:
            continue

        enrollment_date = enrollment.enrollment_date
        days_since = (today - enrollment_date).days if enrollment_date else 0
        row = {
            'enrollment_id': enrollment.id,
            'course_id': enrollment.course_id,
            'course_name': enrollment.course.name,
            'student_id': enrollment.student_id,
            'student_name': enrollment.student.full_name,
            'phone': enrollment.student.phone,
            'enrollment_date': enrollment_date,
            'days_since_enrollment': max(days_since, 0),
            'net_amount': net_amount,
            'paid_amount': paid_total,
            'remaining': remaining,
        }
        rows.append(row)

        course_bucket = grouped_map.setdefault(enrollment.course_id, {
            'course': enrollment.course,
            'date_groups': {},
            'total_students': 0,
            'total_outstanding': Decimal('0'),
        })
        date_bucket = course_bucket['date_groups'].setdefault(enrollment_date, {
            'date': enrollment_date,
            'students': [],
            'total_students': 0,
            'total_outstanding': Decimal('0'),
            'max_days_since_enrollment': 0,
        })
        date_bucket['students'].append(row)
        date_bucket['total_students'] += 1
        date_bucket['total_outstanding'] += remaining
        date_bucket['max_days_since_enrollment'] = max(date_bucket['max_days_since_enrollment'], row['days_since_enrollment'])
        course_bucket['total_students'] += 1
        course_bucket['total_outstanding'] += remaining

    grouped_courses = []
    for course_bucket in grouped_map.values():
        grouped_courses.append({
            'course': course_bucket['course'],
            'date_groups': sorted(course_bucket['date_groups'].values(), key=lambda item: item['date']),
            'total_students': course_bucket['total_students'],
            'total_outstanding': course_bucket['total_outstanding'],
        })

    grouped_courses.sort(key=lambda item: item['course'].name)
    return rows, {
        'grouped_courses': grouped_courses,
        'totals': {
            'total_students': len(rows),
            'total_courses': len(grouped_courses),
            'total_outstanding': sum(row['remaining'] for row in rows),
        }
    }


def _withdraw_quick_enrollment(enrollment, user, withdrawal_reason='', refund_amount=None):
    student = enrollment.student
    if enrollment.is_completed:
        raise ValueError('هذه الدورة مسحوبة مسبقاً')

    with transaction.atomic():
        paid_total = QuickStudentReceipt.objects.filter(
            quick_student=student,
            quick_enrollment=enrollment,
            course=enrollment.course
        ).aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')

        refund_amount = paid_total
        refund_result = _adjust_quick_receipts_for_refund(student, enrollment, refund_amount)
        actual_refund = refund_result['refunded_amount']

        description = f"سحب طالب سريع {student.full_name} من {enrollment.course.name}"
        if withdrawal_reason:
            description = f"{description} - {withdrawal_reason}"

        created_entries = _build_quick_withdrawal_entry(
            enrollment=enrollment,
            user=user,
            refunded_amount=actual_refund,
            description=description,
        )

        enrollment.is_completed = True
        enrollment.completion_date = timezone.now().date()
        enrollment.save(update_fields=['is_completed', 'completion_date'])

    return {
        'actual_refund': actual_refund,
        'student_name': student.full_name,
        'course_name': enrollment.course.name,
        'created_entry_ids': [entry.id for entry in created_entries],
    }


def _delete_quick_enrollment(enrollment, paid_total=None):
    if enrollment.is_completed:
        raise ValueError('هذه الدورة مسحوبة مسبقاً')

    if paid_total is None:
        paid_total = _get_quick_enrollment_paid_total(enrollment, enrollment.student)

    if paid_total > Decimal('0'):
        raise ValueError('لا يمكن حذف التسجيل لأنه يحتوي على دفعات مرتبطة')

    student_name = enrollment.student.full_name
    course_name = enrollment.course.name
    enrollment.delete()

    return {
        'student_name': student_name,
        'course_name': course_name,
    }


def _snapshot_outstanding_totals(totals):
    return {
        'total_courses': int(totals.get('total_courses', 0) or 0),
        'total_students': int(totals.get('total_students', 0) or 0),
        'total_paid_students': int(totals.get('total_paid_students', 0) or 0),
        'total_outstanding_students': int(totals.get('total_outstanding_students', 0) or 0),
        'total_paid_amount': float(totals.get('total_paid_amount', 0) or 0),
        'total_outstanding_amount': float(totals.get('total_outstanding_amount', 0) or 0),
    }


def _build_outstanding_comparison(current_totals, previous_totals):
    if not previous_totals:
        return None

    def make_item(label, key, improve_when):
        current_value = current_totals.get(key, 0)
        previous_value = previous_totals.get(key, 0)
        delta = current_value - previous_value
        if delta > 0:
            trend = 'زيادة'
        elif delta < 0:
            trend = 'نقصان'
        else:
            trend = 'ثبات'

        if improve_when == 'up':
            improved = delta > 0
        elif improve_when == 'down':
            improved = delta < 0
        else:
            improved = None

        return {
            'label': label,
            'current': current_value,
            'previous': previous_value,
            'delta': delta,
            'trend': trend,
            'improved': improved,
        }

    items = [
        make_item('إجمالي الطلاب', 'total_students', None),
        make_item('الطلاب المسددين', 'total_paid_students', 'up'),
        make_item('الطلاب غير المسددين', 'total_outstanding_students', 'down'),
        make_item('إجمالي المدفوع', 'total_paid_amount', 'up'),
        make_item('إجمالي المتبقي', 'total_outstanding_amount', 'down'),
    ]

    improvement_count = sum(1 for item in items if item['improved'] is True)
    decline_count = sum(1 for item in items if item['improved'] is False)

    return {
        'items': items,
        'improvement_count': improvement_count,
        'decline_count': decline_count,
        'total_tracked': improvement_count + decline_count,
    }


class QuickOutstandingCoursesPrintView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/outstanding_course_print.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        course_type, course_type_label, report_label = _get_outstanding_course_type(self.request)
        courses = QuickCourse.objects.filter(is_active=True).select_related('academic_year').order_by('name')
        if course_type != 'ALL':
            courses = courses.filter(course_type=course_type)
        course_data, totals = _build_quick_outstanding_course_summary(courses, include_zero_outstanding=True)
        course_data = sorted(
            course_data,
            key=lambda row: (-row['total_students'], row['course'].name)
        )
        current_snapshot = _snapshot_outstanding_totals(totals)
        previous_snapshot = self.request.session.get('quick_outstanding_report_snapshot')
        previous_time = self.request.session.get('quick_outstanding_report_timestamp')
        comparison = _build_outstanding_comparison(current_snapshot, previous_snapshot)

        context.update({
            'courses': course_data,
            'totals': totals,
            'print_date': timezone.now().strftime('%Y-%m-%d %H:%M'),
            'comparison': comparison,
            'previous_report_time': previous_time,
            'course_type': course_type,
            'course_type_label': course_type_label,
            'course_type_report_label': report_label,
        })

        self.request.session['quick_outstanding_report_snapshot'] = current_snapshot
        self.request.session['quick_outstanding_report_timestamp'] = timezone.now().strftime('%Y-%m-%d %H:%M')
        return context

    
 # الفصول الدراسية
class AcademicYearListView(LoginRequiredMixin, ListView):
    model = AcademicYear
    template_name = 'quick/academic_year_list.html'
    context_object_name = 'academic_years'
    
    def get_queryset(self):
        return AcademicYear.objects.all().order_by('-start_date')

class AcademicYearCreateView(LoginRequiredMixin, CreateView):
    model = AcademicYear
    form_class = AcademicYearForm
    template_name = 'quick/academic_year_form.html'
    success_url = reverse_lazy('quick:academic_year_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'تم إضافة الفصل الدراسي بنجاح')
        return super().form_valid(form)

class CloseAcademicYearView(LoginRequiredMixin, DetailView):
    model = AcademicYear
    template_name = 'quick/academic_year_close.html'
    
    def post(self, request, *args, **kwargs):
        academic_year = self.get_object()
        password = request.POST.get('password')
        
        # التحقق من كلمة المرور
        if not request.user.check_password(password):
            messages.error(request, 'كلمة المرور غير صحيحة')
            return render(request, self.template_name, {'academic_year': academic_year})
        
        academic_year.is_closed = True
        academic_year.closed_by = request.user
        academic_year.closed_at = timezone.now()
        academic_year.save()
        
        messages.success(request, 'تم إغلاق الفصل الدراسي بنجاح')
        return redirect('quick:academic_year_list')

# الدورات السريعة
class QuickCourseListView(LoginRequiredMixin, ListView):
    model = QuickCourse
    template_name = 'quick/quick_course_list.html'
    context_object_name = 'courses'
    
    def get_queryset(self):
        return (
            QuickCourse.objects.filter(is_active=True)
            .select_related('academic_year')
            .annotate(
                enrollments_count=Count(
                    'enrollments',
                    filter=Q(enrollments__is_completed=False, enrollments__student__is_active=True),
                    distinct=True,
                ),
                sessions_count=Count('sessions', filter=Q(sessions__is_active=True), distinct=True),
            )
            .order_by('-created_at')
        )


class QuickClassroomListView(LoginRequiredMixin, ListView):
    model = Classroom
    template_name = 'quick/quick_classroom_list.html'
    context_object_name = 'classrooms'

    def get_queryset(self):
        return Classroom.objects.filter(class_type='course').order_by('name', 'id')


class QuickClassroomCreateView(LoginRequiredMixin, CreateView):
    model = Classroom
    form_class = QuickClassroomForm
    template_name = 'quick/quick_classroom_form.html'
    success_url = reverse_lazy('quick:classroom_list')

    def form_valid(self, form):
        messages.success(self.request, 'تمت إضافة كلاس/قاعة للدورات السريعة.')
        return super().form_valid(form)


class QuickClassroomUpdateView(LoginRequiredMixin, UpdateView):
    model = Classroom
    form_class = QuickClassroomForm
    template_name = 'quick/quick_classroom_form.html'
    success_url = reverse_lazy('quick:classroom_list')

    def get_queryset(self):
        return Classroom.objects.filter(class_type='course')

    def form_valid(self, form):
        messages.success(self.request, 'تم تحديث بيانات الصف/القاعة.')
        return super().form_valid(form)


class QuickStudentIntersectionView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/student_intersections.html'

    def _get_selected_course_ids(self):
        selected_ids = []
        seen_ids = set()

        for raw_value in self.request.GET.getlist('course_ids'):
            try:
                course_id = int(raw_value)
            except (TypeError, ValueError):
                continue
            if course_id not in seen_ids:
                selected_ids.append(course_id)
                seen_ids.add(course_id)

        add_course_raw = self.request.GET.get('add_course')
        if add_course_raw:
            try:
                add_course_id = int(add_course_raw)
            except (TypeError, ValueError):
                add_course_id = None
            if add_course_id and add_course_id not in seen_ids:
                selected_ids.append(add_course_id)

        return selected_ids

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        selected_course_ids = self._get_selected_course_ids()
        courses_queryset = (
            QuickCourse.objects.filter(is_active=True)
            .select_related('academic_year')
            .annotate(
                active_students_count=Count(
                    'enrollments',
                    filter=Q(enrollments__is_completed=False, enrollments__student__is_active=True),
                    distinct=True,
                )
            )
            .order_by('-academic_year__start_date', 'name')
        )

        courses_map = {course.id: course for course in courses_queryset}
        selected_courses = [courses_map[course_id] for course_id in selected_course_ids if course_id in courses_map]
        selected_course_ids = [course.id for course in selected_courses]

        intersection_ids = None
        for course_id in selected_course_ids:
            course_student_ids = set(
                QuickEnrollment.objects.filter(
                    course_id=course_id,
                    is_completed=False,
                    student__is_active=True,
                ).values_list('student_id', flat=True)
            )
            if intersection_ids is None:
                intersection_ids = course_student_ids
            else:
                intersection_ids &= course_student_ids

        matching_student_ids = sorted(intersection_ids) if intersection_ids else []
        matching_students = []
        if matching_student_ids:
            matching_students = list(
                QuickStudent.objects.filter(id__in=matching_student_ids, is_active=True)
                .select_related('student', 'academic_year')
                .annotate(
                    total_active_courses=Count(
                        'enrollments',
                        filter=Q(enrollments__is_completed=False),
                        distinct=True,
                    ),
                    matched_courses_count=Count(
                        'enrollments',
                        filter=Q(enrollments__course_id__in=selected_course_ids, enrollments__is_completed=False),
                        distinct=True,
                    ),
                )
                .order_by('full_name', 'id')
            )

        selected_course_entries = []
        for index, course in enumerate(selected_courses):
            remaining_ids = [str(course_id) for i, course_id in enumerate(selected_course_ids) if i != index]
            remove_url = reverse('quick:student_intersections')
            if remaining_ids:
                remove_url = f"{remove_url}?{urlencode([('course_ids', course_id) for course_id in remaining_ids])}"
            selected_course_entries.append({
                'course': course,
                'remove_url': remove_url,
            })

        available_courses = [course for course in courses_queryset if course.id not in selected_course_ids]

        context.update({
            'selected_courses': selected_course_entries,
            'selected_course_ids': selected_course_ids,
            'available_courses': available_courses,
            'matching_students': matching_students,
            'matching_students_count': len(matching_students),
            'selected_courses_count': len(selected_courses),
            'total_available_courses': courses_queryset.count(),
            'has_selection': bool(selected_courses),
            'intersection_mode': 'single' if len(selected_courses) == 1 else 'multi',
            'is_exact_intersection': len(selected_courses) > 1,
        })
        return context


def _resolve_quick_course_type_value(raw_value, allow_all=True):
    valid_course_types = {value for value, _ in QuickCourse.COURSE_TYPE_CHOICES}
    if allow_all and raw_value == 'ALL':
        return 'ALL'
    if raw_value in valid_course_types:
        return raw_value
    return 'INTENSIVE'


def _extract_quick_teacher_short_name(course_name):
    course_name = str(course_name or '').strip()
    if not course_name:
        return ''

    teacher_full_name = ''
    if '(' in course_name and ')' in course_name:
        inner_text = course_name.rsplit('(', 1)[-1].split(')', 1)[0].strip()
        teacher_full_name = inner_text
        for prefix in ('الأستاذة', 'الاستاذة', 'الأستاذ', 'الاستاذ'):
            if teacher_full_name.startswith(prefix):
                teacher_full_name = teacher_full_name[len(prefix):].strip()
                break
    teacher_tokens = [token for token in teacher_full_name.split() if token]
    if teacher_tokens:
        return teacher_tokens[0]
    return course_name


def _extract_quick_teacher_full_name(course_name):
    course_name = str(course_name or '').strip()
    if not course_name:
        return ''

    teacher_full_name = ''
    if '(' in course_name and ')' in course_name:
        inner_text = course_name.rsplit('(', 1)[-1].split(')', 1)[0].strip()
        teacher_full_name = inner_text
        for prefix in ('الأستاذة', 'الاستاذة', 'الأستاذ', 'الاستاذ'):
            if teacher_full_name.startswith(prefix):
                teacher_full_name = teacher_full_name[len(prefix):].strip()
                break
    return teacher_full_name or course_name


def _normalize_manual_sort_key(value):
    value = str(value or '').strip()
    return ''.join(value.split())


MANUAL_SORT_TEACHER_ORDER = [
    'علاء',
    'قصي',
    'مهند',
    'طارق',
    'عامر',
    'عبد الرحمن',
    'عبدالله',
    'عبد الوهاب',
    'عمار',
    'اللاء',
    'خالدرياضيات',
    'حالدعلوم',
    'رامي',
    'سامر',
    'ضياءجغرافيا',
    'محمد',
    'ملهم',
    'نبيل',
    'زياد',
    'ضياء تاريخ',
    'عمار جغرافيا',
    'عيسى',
    'مجد',
    'محمدفلسفة',
]
MANUAL_SORT_TEACHER_ALIASES = {
    'خالدرياضيات': ['خالد رياضيات'],
    'حالدعلوم': ['حالد علوم', 'خالدعلوم', 'خالد علوم'],
    'ضياءجغرافيا': ['ضياء جغرافيا'],
    'ضياء تاريخ': ['ضياءتاريخ'],
    'محمدفلسفة': ['محمد فلسفة'],
}
MANUAL_SORT_TEACHER_ORDER_MAP = {}
MANUAL_SORT_TEACHER_ORDER_KEYS = []
for index, item in enumerate(MANUAL_SORT_TEACHER_ORDER):
    normalized = _normalize_manual_sort_key(item)
    if normalized not in MANUAL_SORT_TEACHER_ORDER_MAP:
        MANUAL_SORT_TEACHER_ORDER_MAP[normalized] = index
        MANUAL_SORT_TEACHER_ORDER_KEYS.append((normalized, index))
    for alias in MANUAL_SORT_TEACHER_ALIASES.get(item, []):
        alias_key = _normalize_manual_sort_key(alias)
        if alias_key not in MANUAL_SORT_TEACHER_ORDER_MAP:
            MANUAL_SORT_TEACHER_ORDER_MAP[alias_key] = index
            MANUAL_SORT_TEACHER_ORDER_KEYS.append((alias_key, index))

UNASSIGNED_FOCUS_TEACHERS = {
    _normalize_manual_sort_key(name)
    for name in [
        'علاء',
        'قصي',
        'مهند',
        'طارق',
        'عد الله',
        'عبد الله',
        'عبد الوهاب',
        'عمار',
    ]
}


def _is_focus_teacher(course_name):
    teacher_full = _extract_quick_teacher_full_name(course_name)
    teacher_key = _normalize_manual_sort_key(teacher_full)
    return teacher_key in UNASSIGNED_FOCUS_TEACHERS


def _manual_course_order_index(column):
    teacher_full = _extract_quick_teacher_full_name(column['course'].name)
    subject = column.get('subject_name', '')
    keys = [
        _normalize_manual_sort_key(f'{teacher_full}{subject}'),
        _normalize_manual_sort_key(teacher_full),
    ]
    for key in keys:
        if key in MANUAL_SORT_TEACHER_ORDER_MAP:
            return MANUAL_SORT_TEACHER_ORDER_MAP[key]
        for ordered_key, index in MANUAL_SORT_TEACHER_ORDER_KEYS:
            if ordered_key and ordered_key in key:
                return index
    return len(MANUAL_SORT_TEACHER_ORDER) + 99


def _extract_quick_course_subject(course_name):
    course_name = str(course_name or '').strip()
    if not course_name:
        return ''

    base_name = course_name.split('(', 1)[0].strip()
    removable_tokens = {
        'مكثفة',
        'التاسع',
        'تاسع',
        'بكالوريا',
        'البكالوريا',
        'الصف',
        'اللغة',
    }
    subject_tokens = [token for token in base_name.split() if token and token not in removable_tokens]
    return ' '.join(subject_tokens).strip() or base_name


def _get_quick_manual_stage_options():
    return [
        {'value': 'NON_NINTH', 'label': 'البكالوريا'},
        {'value': 'NINTH', 'label': 'تاسع'},
    ]


def _quick_manual_selection_table_exists():
    try:
        return QuickManualSortingSelection._meta.db_table in connection.introspection.table_names()
    except Exception:
        return False


def _sync_quick_manual_sorting_selection(enrollment, session=None, user=None):
    if not _quick_manual_selection_table_exists():
        return
    if session is None:
        QuickManualSortingSelection.objects.filter(enrollment=enrollment).delete()
        return
    QuickManualSortingSelection.objects.update_or_create(
        enrollment=enrollment,
        defaults={
            'session': session,
            'selected_by': user,
        },
    )


def _filter_quick_courses_by_stage(courses, stage):
    if stage == 'NINTH':
        return courses.filter(name__icontains='تاسع')
    if stage == 'NON_NINTH':
        return courses.exclude(name__icontains='تاسع')
    return courses


def _build_quick_manual_sorting_payload(course_type='INTENSIVE', stage='NON_NINTH', assignment_status='ALL', course_id=None, session_id=None):
    course_type = _resolve_quick_course_type_value(course_type, allow_all=True)
    stage = (stage or 'NON_NINTH').upper()
    valid_stages = {item['value'] for item in _get_quick_manual_stage_options()}
    if stage not in valid_stages:
        stage = 'NON_NINTH'
    course_type_options = _get_course_type_options()
    course_type_labels = {item['value']: item['label'] for item in course_type_options}
    stage_options = _get_quick_manual_stage_options()
    stage_labels = {item['value']: item['label'] for item in stage_options}

    assignment_status = (assignment_status or 'ALL').upper()
    assignment_status_options = [
        {'value': 'ALL', 'label': 'الكل'},
        {'value': 'UNASSIGNED', 'label': 'غير منزلين بالكامل'},
        {'value': 'PARTIAL', 'label': 'منزلين جزئياً'},
        {'value': 'ASSIGNED', 'label': 'منزلين بالكامل'},
        {'value': 'SINGLE_ONLY', 'label': 'طلاب الكلاس الواحد فقط'},
    ]
    valid_assignment_status = {item['value'] for item in assignment_status_options}
    if assignment_status not in valid_assignment_status:
        assignment_status = 'ALL'

    courses_queryset = QuickCourse.objects.filter(is_active=True)
    if course_type != 'ALL':
        courses_queryset = courses_queryset.filter(course_type=course_type)
    courses_queryset = _filter_quick_courses_by_stage(courses_queryset, stage)
    available_courses = list(
        courses_queryset.select_related('academic_year')
        .annotate(
            active_enrollments_count=Count(
                'enrollments',
                filter=Q(enrollments__is_completed=False, enrollments__student__is_active=True),
                distinct=True,
            )
        )
        .filter(active_enrollments_count__gt=0)
        .order_by('name', 'id')
    )
    available_course_ids = {course.id for course in available_courses}
    try:
        selected_course_id = _parse_quick_posted_int(course_id) if course_id not in (None, '') else None
    except (TypeError, ValueError):
        selected_course_id = None
    if selected_course_id not in available_course_ids:
        selected_course_id = None

    courses = list(available_courses)
    selected_course = next((course for course in available_courses if course.id == selected_course_id), None)
    course_filter_options = [{'value': '', 'label': 'كل الدورات الظاهرة'}]
    course_filter_options.extend(
        {
            'value': str(course.id),
            'label': course.name,
        }
        for course in available_courses
    )

    course_ids = [course.id for course in courses]
    sessions = list(
        QuickCourseSession.objects.filter(course_id__in=course_ids, is_active=True)
        .select_related('course')
        .order_by('course__name', 'start_date', 'start_time', 'id')
    )

    sessions_by_course = defaultdict(list)
    current_loads = {}
    for session in sessions:
        sessions_by_course[session.course_id].append(session)
        current_loads[session.id] = session.enrolled_count

    course_columns = []
    for course in courses:
        option_rows = []
        active_sessions = sessions_by_course.get(course.id, [])
        for period_number, session in enumerate(active_sessions, start=1):
            session_title = (session.title or '').strip() or f"الفترة {period_number}"
            option_rows.append({
                'id': session.id,
                'period_number': period_number,
                'label': f"الفترة {period_number}",
                'session_title': session_title,
                'select_label': session_title,
                'session': session,
                'timing': (
                    f"{session.start_time.strftime('%I:%M %p')} - "
                    f"{session.end_time.strftime('%I:%M %p') if session.end_time else 'مفتوح'}"
                ),
            })
        course_columns.append({
            'course': course,
            'short_teacher_name': _extract_quick_teacher_short_name(course.name),
            'teacher_full_name': _extract_quick_teacher_full_name(course.name),
            'subject_name': _extract_quick_course_subject(course.name),
            'sessions': option_rows,
            'sessions_count': len(option_rows),
            'single_session_id': option_rows[0]['id'] if len(option_rows) == 1 else None,
            'single_period_number': option_rows[0]['period_number'] if len(option_rows) == 1 else None,
            'has_sessions': bool(option_rows),
        })

    course_columns.sort(
        key=lambda item: (
            _manual_course_order_index(item),
            -item['sessions_count'],
            item['teacher_full_name'],
            item['course'].name,
            item['course'].id,
        )
    )
    course_column_map = {item['course'].id: item for item in course_columns}
    selected_course_sessions = course_column_map.get(selected_course_id, {}).get('sessions', []) if selected_course_id else []
    selected_course_session_ids = {item['id'] for item in selected_course_sessions}
    try:
        selected_filter_session_id = _parse_quick_posted_int(session_id) if session_id not in (None, '') else None
    except (TypeError, ValueError):
        selected_filter_session_id = None
    if selected_filter_session_id not in selected_course_session_ids:
        selected_filter_session_id = None
    selected_filter_session_option = next(
        (item for item in selected_course_sessions if item['id'] == selected_filter_session_id),
        None,
    )
    session_filter_options = [{'value': '', 'label': 'كل الكلاسات'}]
    if selected_course_sessions:
        session_filter_options.extend(
            {
                'value': str(item['id']),
                'label': item['session_title'],
            }
            for item in selected_course_sessions
        )
    active_session_ids_by_course = {
        item['course'].id: {session_option['id'] for session_option in item['sessions']}
        for item in course_columns
    }
    active_period_numbers_by_session_id = {
        session_option['id']: session_option['period_number']
        for item in course_columns
        for session_option in item['sessions']
    }

    enrollments_qs = QuickEnrollment.objects.filter(
        course_id__in=course_ids,
        is_completed=False,
        student__is_active=True,
    )
    select_related_fields = [
        'student',
        'student__created_by',
        'student__student',
        'course',
        'session_assignment',
        'session_assignment__session',
    ]
    manual_selection_enabled = _quick_manual_selection_table_exists()
    if manual_selection_enabled:
        select_related_fields.extend([
            'manual_sorting_selection',
            'manual_sorting_selection__session',
        ])
    enrollments = list(
        enrollments_qs.select_related(*select_related_fields).order_by('student__full_name', 'course__name', 'id')
    )

    student_rows_by_id = {}
    unassigned_enrollments = []
    student_assignment_totals = {}
    for enrollment in enrollments:
        student = enrollment.student
        assignment = getattr(enrollment, 'session_assignment', None)
        manual_selection = getattr(enrollment, 'manual_sorting_selection', None) if manual_selection_enabled else None
        current_session = getattr(assignment, 'session', None) if assignment else None
        current_session_id = getattr(assignment, 'session_id', None)
        active_session_ids = active_session_ids_by_course.get(enrollment.course_id, set())
        is_assigned_to_active_session = bool(current_session_id and current_session_id in active_session_ids)
        has_single_active_session = len(active_session_ids) == 1
        totals = student_assignment_totals.setdefault(
            student.id,
            {
                'non_single_total': 0,
                'non_single_assigned': 0,
                'single_total': 0,
                'no_active_total': 0,
                'assigned_total': 0,
            },
        )
        has_active_sessions = len(active_session_ids) > 0
        if has_single_active_session:
            totals['single_total'] += 1
        elif has_active_sessions:
            totals['non_single_total'] += 1
            if is_assigned_to_active_session:
                totals['non_single_assigned'] += 1
        else:
            totals['no_active_total'] += 1
        if is_assigned_to_active_session:
            totals['assigned_total'] += 1
        else:
            if current_session_id and current_session and not current_session.is_active:
                issue_label = 'مربوط على فترة غير فعالة'
            elif not active_session_ids:
                issue_label = 'الدورة بلا فترات فعالة'
            else:
                issue_label = 'غير مربوط بفترة'
            if _is_focus_teacher(enrollment.course.name) and len(active_session_ids) != 1:
                unassigned_enrollments.append({
                    'student': student,
                    'enrollment': enrollment,
                    'course': enrollment.course,
                    'issue_label': issue_label,
                    'active_sessions_count': len(active_session_ids),
                    'manual_selected_session_id': getattr(manual_selection, 'session_id', None),
                    'manual_selected_period_number': (
                        active_period_numbers_by_session_id.get(manual_selection.session_id)
                        if manual_selection and getattr(manual_selection, 'session_id', None)
                        else None
                    ),
                    'current_session_title': current_session.title if current_session else '',
                    'manage_url': reverse('quick:course_sessions_manage', kwargs={'course_id': enrollment.course_id}),
                })

        student_row = student_rows_by_id.setdefault(
            student.id,
            {
                'student': student,
                'enrollments_by_course': {},
                'enrolled_courses_count': 0,
            },
        )
        student_row['enrollments_by_course'][enrollment.course_id] = enrollment
        student_row['enrolled_courses_count'] += 1

    selected_course_student_ids = None
    if selected_course_id:
        selected_course_student_ids = {
            student_id
            for student_id, student_row in student_rows_by_id.items()
            if selected_course_id in student_row['enrollments_by_course']
        }

    fully_unassigned_student_ids = set()
    fully_assigned_student_ids = set()
    partially_assigned_student_ids = set()
    single_only_student_ids = set()
    for student_id, totals in student_assignment_totals.items():
        non_single_total = totals['non_single_total']
        non_single_assigned = totals['non_single_assigned']
        single_total = totals['single_total']
        no_active_total = totals['no_active_total']
        assigned_total = totals['assigned_total']

        if non_single_total == 0:
            if no_active_total > 0:
                fully_unassigned_student_ids.add(student_id)
            elif single_total > 0 and assigned_total == 0:
                single_only_student_ids.add(student_id)
            elif assigned_total > 0:
                fully_assigned_student_ids.add(student_id)
            else:
                fully_unassigned_student_ids.add(student_id)
            continue

        if non_single_assigned == 0:
            fully_unassigned_student_ids.add(student_id)
        elif non_single_assigned < non_single_total or no_active_total > 0:
            partially_assigned_student_ids.add(student_id)
        else:
            fully_assigned_student_ids.add(student_id)
    unassigned_enrollments = [
        row for row in unassigned_enrollments
        if row['student'].id in fully_unassigned_student_ids
    ]
    if selected_course_student_ids is not None:
        unassigned_enrollments = [
            row for row in unassigned_enrollments
            if row['student'].id in selected_course_student_ids
        ]
    if selected_filter_session_id is not None:
        unassigned_enrollments = [
            row for row in unassigned_enrollments
            if row.get('manual_selected_session_id') == selected_filter_session_id
        ]
    focus_unassigned_student_ids = {row['student'].id for row in unassigned_enrollments}

    student_rows = []
    deprioritized_subjects = {'التاريخ', 'الجغرافيا', 'الفلسفة'}
    for student_row in student_rows_by_id.values():
        cells = []
        student_subjects = set()
        for column in course_columns:
            course = column['course']
            enrollment = student_row['enrollments_by_course'].get(course.id)
            if enrollment is None:
                cells.append({
                    'course_id': course.id,
                    'is_enrolled': False,
                    'status_label': 'غير مسجل',
                })
                continue

            assignment = getattr(enrollment, 'session_assignment', None)
            manual_selection = getattr(enrollment, 'manual_sorting_selection', None) if manual_selection_enabled else None
            student_subjects.add(column['subject_name'])
            selected_session_id = None
            if column['single_session_id']:
                selected_session_id = column['single_session_id']
            else:
                current_session_id = getattr(assignment, 'session_id', None)
                if current_session_id in {item['id'] for item in column['sessions']}:
                    selected_session_id = current_session_id
                elif manual_selection and getattr(manual_selection, 'session_id', None) in {item['id'] for item in column['sessions']}:
                    # Use the saved manual suggestion only when there is no current live assignment.
                    selected_session_id = manual_selection.session_id
            selected_session_option = next(
                (item for item in column['sessions'] if item['id'] == selected_session_id),
                None,
            )
            current_session_option = next(
                (item for item in column['sessions'] if item['id'] == getattr(assignment, 'session_id', None)),
                None,
            )

            cells.append({
                'course_id': course.id,
                'enrollment': enrollment,
                'is_enrolled': True,
                'status_label': 'مسجل',
                'sessions': column['sessions'],
                'has_sessions': column['has_sessions'],
                'selected_session_id': selected_session_id,
                'selected_period_number': selected_session_option['period_number'] if selected_session_option else None,
                'selected_session_title': selected_session_option['session_title'] if selected_session_option else '',
                'current_session_id': getattr(assignment, 'session_id', None),
                'current_session_title': (
                    current_session_option['session_title']
                    if current_session_option
                    else (getattr(getattr(assignment, 'session', None), 'title', None) or '')
                ),
                'manual_selected_session_id': getattr(manual_selection, 'session_id', None),
            })

        student_row['cells'] = cells
        student_row['has_deprioritized_subject'] = any(subject in deprioritized_subjects for subject in student_subjects)
        student_rows.append(student_row)

    student_rows.sort(
        key=lambda item: (
            item['has_deprioritized_subject'],
            -item['enrolled_courses_count'],
            item['student'].full_name,
            item['student'].id,
        )
    )
    if selected_course_student_ids is not None:
        student_rows = [
            row for row in student_rows
            if row['student'].id in selected_course_student_ids
        ]
    if selected_filter_session_id is not None and selected_course_id is not None:
        student_rows = [
            row for row in student_rows
            if any(
                cell['course_id'] == selected_course_id and cell.get('selected_session_id') == selected_filter_session_id
                for cell in row['cells']
            )
        ]
    student_rows_all = list(student_rows)
    if assignment_status == 'UNASSIGNED':
        student_rows = [row for row in student_rows_all if row['student'].id in fully_unassigned_student_ids]
    elif assignment_status == 'PARTIAL':
        student_rows = [row for row in student_rows_all if row['student'].id in partially_assigned_student_ids]
    elif assignment_status == 'ASSIGNED':
        student_rows = [row for row in student_rows_all if row['student'].id in fully_assigned_student_ids]
    elif assignment_status == 'SINGLE_ONLY':
        student_rows = [row for row in student_rows_all if row['student'].id in single_only_student_ids]
    else:
        student_rows = list(student_rows_all)

    unassigned_print_course_columns = [
        column for column in course_columns
        if column['sessions_count'] != 1
    ]
    unassigned_print_course_ids = [column['course'].id for column in unassigned_print_course_columns]
    student_rows_unassigned_print = []
    if unassigned_print_course_ids:
        for row in student_rows_all:
            if row['student'].id not in fully_unassigned_student_ids:
                continue
            student_rows_unassigned_print.append({
                **row,
                'cells': [
                    cell for cell in row['cells']
                    if cell['course_id'] in unassigned_print_course_ids
                ],
            })

    filtered_student_ids = {row['student'].id for row in student_rows_all}
    filtered_fully_assigned_student_ids = fully_assigned_student_ids & filtered_student_ids
    filtered_partially_assigned_student_ids = partially_assigned_student_ids & filtered_student_ids
    filtered_single_only_student_ids = single_only_student_ids & filtered_student_ids

    student_order_index = {
        row['student'].id: index
        for index, row in enumerate(student_rows_all)
    }
    course_order_index = {
        item['course'].id: index
        for index, item in enumerate(course_columns)
    }
    unassigned_enrollments.sort(
        key=lambda item: (
            student_order_index.get(item['student'].id, 10**9),
            course_order_index.get(item['course'].id, 10**9),
            item['student'].full_name,
            item['course'].name,
            item['enrollment'].id,
        )
    )

    return {
        'course_type': course_type,
        'course_type_label': course_type_labels.get(course_type, course_type),
        'course_type_options': course_type_options,
        'selected_course_id': selected_course_id,
        'selected_course_label': selected_course.name if selected_course else 'كل الدورات الظاهرة',
        'course_filter_options': course_filter_options,
        'selected_session_id': selected_filter_session_id,
        'selected_session_label': selected_filter_session_option['session_title'] if selected_filter_session_option else 'كل الكلاسات',
        'session_filter_options': session_filter_options,
        'stage': stage,
        'stage_label': stage_labels.get(stage, stage),
        'stage_options': stage_options,
        'assignment_status': assignment_status,
        'assignment_status_label': next((item['label'] for item in assignment_status_options if item['value'] == assignment_status), ''),
        'assignment_status_options': assignment_status_options,
        'courses': courses,
        'course_columns': course_columns,
        'course_column_map': course_column_map,
        'student_rows': student_rows,
        'student_rows_all': student_rows_all,
        'current_loads': current_loads,
        'unassigned_enrollments': unassigned_enrollments,
        'unassigned_enrollment_count': len(unassigned_enrollments),
        'unassigned_student_count': len(focus_unassigned_student_ids),
        'assigned_student_count': len(filtered_fully_assigned_student_ids),
        'partial_student_count': len(filtered_partially_assigned_student_ids),
        'single_only_student_count': len(filtered_single_only_student_ids),
        'course_columns_unassigned_print': unassigned_print_course_columns,
        'student_rows_unassigned_print': student_rows_unassigned_print,
        'manual_selection_enabled': manual_selection_enabled,
        'generated_at': timezone.localtime(),
    }


def _save_quick_manual_sorting_assignments(
    posted_assignments,
    enrollments_map,
    sessions_by_course,
    session_course_map,
    user,
    manual_selection_enabled=False,
):
    saved_count = 0
    created_count = 0
    updated_count = 0
    cleared_count = 0
    unchanged_count = 0
    change_summaries = []
    changed_enrollment_details = []
    validation_errors = []

    with transaction.atomic():
        for enrollment_id, raw_value in posted_assignments.items():
            enrollment = enrollments_map.get(enrollment_id)
            if not enrollment:
                continue

            available_sessions = sessions_by_course.get(enrollment.course_id, [])
            session_meta_by_id = {
                session.id: {
                    'period_number': index + 1,
                    'title': (session.title or '').strip() or f"الفترة {index + 1}",
                }
                for index, session in enumerate(available_sessions)
            }
            has_single_session = len(available_sessions) == 1
            target_session_id = None

            if raw_value:
                try:
                    target_session_id = _parse_quick_posted_int(raw_value)
                except (TypeError, ValueError):
                    validation_errors.append(
                        f"اختيار غير صالح للطالب {enrollment.student.full_name} في دورة {enrollment.course.name}."
                    )
                    continue
            elif has_single_session:
                target_session_id = available_sessions[0].id

            if target_session_id is not None and session_course_map.get(target_session_id) != enrollment.course_id:
                validation_errors.append(
                    f"الفترة المختارة لا تتبع نفس الدورة للطالب {enrollment.student.full_name}."
                )
                continue

            assignment = getattr(enrollment, 'session_assignment', None)
            current_session_id = getattr(assignment, 'session_id', None)
            current_session_label = (
                session_meta_by_id.get(current_session_id, {}).get('title')
                or (getattr(getattr(assignment, 'session', None), 'title', None) or '').strip()
                or 'فترة غير محددة'
            )
            target_session_label = (
                session_meta_by_id.get(target_session_id, {}).get('title')
                if target_session_id is not None
                else ''
            ) or 'فترة غير محددة'
            student_name = enrollment.student.full_name
            course_name = enrollment.course.name

            if target_session_id is None:
                if assignment:
                    assignment.delete()
                    cleared_count += 1
                    saved_count += 1
                    change_summaries.append(
                        f'{student_name} - {course_name}: إزالة التنزيل من {current_session_label}'
                    )
                    changed_enrollment_details.append({
                        'enrollment_id': enrollment.id,
                        'label': 'تمت إزالة التنزيل',
                        'session_title': '',
                    })
                    if manual_selection_enabled:
                        _sync_quick_manual_sorting_selection(enrollment=enrollment, session=None, user=user)
                else:
                    unchanged_count += 1
                continue

            if assignment and current_session_id == target_session_id:
                assignment.assigned_by = user
                assignment.save(update_fields=['assigned_by'])
                updated_count += 1
                saved_count += 1
                change_summaries.append(
                    f'{student_name} - {course_name}: تأكيد {target_session_label}'
                )
                changed_enrollment_details.append({
                    'enrollment_id': enrollment.id,
                    'label': 'تم تأكيد الكلاس',
                    'session_title': target_session_label,
                })
                if manual_selection_enabled:
                    _sync_quick_manual_sorting_selection(
                        enrollment=enrollment,
                        session=assignment.session,
                        user=user,
                    )
                continue

            assignment, created = QuickCourseSessionEnrollment.objects.update_or_create(
                enrollment=enrollment,
                defaults={
                    'session_id': target_session_id,
                    'assigned_by': user,
                },
            )
            if created:
                created_count += 1
                change_summaries.append(
                    f'{student_name} - {course_name}: من غير منزل إلى {target_session_label}'
                )
                changed_enrollment_details.append({
                    'enrollment_id': enrollment.id,
                    'label': 'تم تنزيل الكلاس',
                    'session_title': target_session_label,
                })
            else:
                updated_count += 1
                change_summaries.append(
                    f'{student_name} - {course_name}: من {current_session_label} إلى {target_session_label}'
                )
                changed_enrollment_details.append({
                    'enrollment_id': enrollment.id,
                    'label': 'تم تعديل الكلاس',
                    'session_title': target_session_label,
                })
            saved_count += 1

            if manual_selection_enabled:
                _sync_quick_manual_sorting_selection(
                    enrollment=enrollment,
                    session=assignment.session,
                    user=user,
                )

    return {
        'saved_count': saved_count,
        'created_count': created_count,
        'updated_count': updated_count,
        'cleared_count': cleared_count,
        'unchanged_count': unchanged_count,
        'change_summaries': change_summaries,
        'changed_enrollment_details': changed_enrollment_details,
        'validation_errors': validation_errors,
    }


@method_decorator(never_cache, name='dispatch')
class QuickManualSortingView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_manual_sorting.html'
    page_size = 30
    flash_session_key = 'quick_manual_sorting_saved_cells'

    def _get_payload(self):
        course_type = self.request.GET.get('course_type') or self.request.POST.get('course_type') or 'INTENSIVE'
        course_id = self.request.GET.get('course_id') or self.request.POST.get('course_id') or ''
        session_id = self.request.GET.get('session_id') or self.request.POST.get('session_id') or ''
        stage = self.request.GET.get('stage') or self.request.POST.get('stage') or 'NON_NINTH'
        assignment_status = (
            self.request.GET.get('assignment_status')
            or self.request.POST.get('assignment_status')
            or 'ALL'
        )
        return _build_quick_manual_sorting_payload(
            course_type=course_type,
            course_id=course_id,
            session_id=session_id,
            stage=stage,
            assignment_status=assignment_status,
        )

    def _get_page_obj(self, rows):
        paginator = Paginator(rows, self.page_size)
        page_number = self.request.GET.get('page') or self.request.POST.get('page') or 1
        return paginator.get_page(page_number)

    def _build_context(self, payload):
        page_obj = self._get_page_obj(payload['student_rows'])
        saved_cells_state = self.request.session.pop(self.flash_session_key, None)
        active_saved_cells = {}
        if (
            saved_cells_state
            and saved_cells_state.get('course_type') == payload['course_type']
            and saved_cells_state.get('course_id') == payload['selected_course_id']
            and saved_cells_state.get('session_id') == payload['selected_session_id']
            and saved_cells_state.get('stage') == payload['stage']
            and saved_cells_state.get('assignment_status') == payload['assignment_status']
            and saved_cells_state.get('page') == page_obj.number
        ):
            active_saved_cells = {
                int(item['enrollment_id']): item
                for item in saved_cells_state.get('cells', [])
                if item.get('enrollment_id') is not None
            }

        if active_saved_cells:
            for row in page_obj.object_list:
                for cell in row.get('cells', []):
                    enrollment = cell.get('enrollment')
                    if not enrollment:
                        continue
                    saved_info = active_saved_cells.get(enrollment.id)
                    if not saved_info:
                        continue
                    cell['was_just_saved'] = True
                    cell['saved_badge_label'] = saved_info.get('label') or 'تم الحفظ'
                    cell['saved_session_title'] = saved_info.get('session_title') or cell.get('selected_session_title') or ''

        base_query_items = [
            ('course_type', payload['course_type']),
            ('stage', payload['stage']),
            ('assignment_status', payload['assignment_status']),
        ]
        if payload['selected_course_id']:
            base_query_items.append(('course_id', payload['selected_course_id']))
        if payload['selected_session_id']:
            base_query_items.append(('session_id', payload['selected_session_id']))
        base_query = urlencode(base_query_items)
        print_url = f"{reverse('quick:manual_sorting_print')}?{base_query}" if payload['courses'] else ''
        unassigned_print_url = (
            f"{reverse('quick:manual_sorting_unassigned_print')}?{base_query}"
            if payload['unassigned_enrollments'] else ''
        )
        student_list_url = reverse('quick:student_list')
        all_sessions_url = reverse('quick:all_sessions_manage')
        return {
            **payload,
            'page_obj': page_obj,
            'student_rows_page': list(page_obj.object_list),
            'base_query': base_query,
            'print_url': print_url,
            'unassigned_print_url': unassigned_print_url,
            'student_list_url': student_list_url,
            'all_sessions_url': all_sessions_url,
            'total_students': len(payload['student_rows']),
            'total_courses': len(payload['course_columns']),
            'students_on_page': len(page_obj.object_list),
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(self._build_context(self._get_payload()))
        return context

    def post(self, request, *args, **kwargs):
        payload = self._get_payload()
        page_obj = self._get_page_obj(payload['student_rows'])
        manual_selection_enabled = payload.get('manual_selection_enabled', False)
        logger.info(
            'manual_sorting_post_start user=%s status=%s stage=%s course_type=%s page=%s',
            getattr(request.user, 'username', 'anonymous'),
            payload.get('assignment_status'),
            payload.get('stage'),
            payload.get('course_type'),
            getattr(page_obj, 'number', ''),
        )

        posted_assignments = {}
        for key, value in request.POST.items():
            if not key.startswith('assignment_'):
                continue
            raw_id = key.replace('assignment_', '')
            try:
                enrollment_id = _parse_quick_posted_int(raw_id)
            except (TypeError, ValueError):
                continue
            posted_assignments[enrollment_id] = (value or '').strip()

        logger.info(
            'manual_sorting_post_payload_count=%s sample=%s',
            len(posted_assignments),
            list(posted_assignments.items())[:5],
        )

        if not posted_assignments:
            logger.warning('manual_sorting_post_empty_payload')
            messages.info(request, 'لا يوجد تغييرات جديدة للحفظ في هذه الصفحة.')
            return self.render_to_response(self._build_context(payload))

        enrollments = list(
            QuickEnrollment.objects.filter(id__in=posted_assignments.keys()).select_related(
                'student',
                'course',
                'session_assignment',
                'session_assignment__session',
            )
        )
        enrollments_map = {enrollment.id: enrollment for enrollment in enrollments}

        course_ids = {enrollment.course_id for enrollment in enrollments}
        active_sessions = list(
            QuickCourseSession.objects.filter(course_id__in=course_ids, is_active=True).select_related('course')
        )
        sessions_by_course = defaultdict(list)
        session_course_map = {}
        for session in active_sessions:
            sessions_by_course[session.course_id].append(session)
            session_course_map[session.id] = session.course_id

        save_result = None
        last_lock_error = None
        for attempt in range(4):
            try:
                close_old_connections()
                _configure_sqlite_busy_timeout(timeout_ms=45000)
                save_result = _save_quick_manual_sorting_assignments(
                    posted_assignments=posted_assignments,
                    enrollments_map=enrollments_map,
                    sessions_by_course=sessions_by_course,
                    session_course_map=session_course_map,
                    user=request.user,
                    manual_selection_enabled=manual_selection_enabled,
                )
                last_lock_error = None
                break
            except OperationalError as exc:
                if 'database is locked' not in str(exc).lower():
                    raise
                last_lock_error = exc
                close_old_connections()
                connection.close()
                time.sleep(0.75 * (attempt + 1))

        if last_lock_error and save_result is None:
            logger.error('manual_sorting_post_locked user=%s', getattr(request.user, 'username', 'anonymous'))
            messages.error(request, 'تعذر حفظ التعديلات حالياً لأن قاعدة البيانات مشغولة. أعد المحاولة بعد ثوانٍ.')
            return self.render_to_response(self._build_context(payload))

        saved_count = save_result['saved_count']
        created_count = save_result['created_count']
        updated_count = save_result['updated_count']
        cleared_count = save_result['cleared_count']
        unchanged_count = save_result['unchanged_count']
        change_summaries = save_result['change_summaries']
        changed_enrollment_details = save_result['changed_enrollment_details']
        validation_errors = save_result['validation_errors']
        logger.info(
            'manual_sorting_post_result saved=%s created=%s updated=%s cleared=%s unchanged=%s errors=%s',
            saved_count,
            created_count,
            updated_count,
            cleared_count,
            unchanged_count,
            len(validation_errors),
        )

        if validation_errors:
            for error in validation_errors[:8]:
                messages.error(request, error)
            if len(validation_errors) > 8:
                messages.error(request, f"يوجد {len(validation_errors) - 8} أخطاء إضافية لم تُعرض هنا.")
            logger.warning('manual_sorting_post_validation_errors sample=%s', validation_errors[:3])
            return self.render_to_response(self._build_context(payload))

        if saved_count:
            messages.success(
                request,
                f'تم حفظ {saved_count} تعديل في صفحة الفرز الحالية. '
                f'(جديد: {created_count}، تعديل: {updated_count}، حذف: {cleared_count})'
            )
            if payload.get('assignment_status') and payload.get('assignment_status') != 'ALL':
                messages.info(request, 'قد تختفي بعض السجلات بعد الحفظ لأن حالة الطالب تغيّرت داخل الفلتر الحالي.')
            for summary in change_summaries[:8]:
                messages.info(request, summary)
            if len(change_summaries) > 8:
                messages.info(request, f'يوجد {len(change_summaries) - 8} تغييرات إضافية لم تُعرض هنا.')
        elif unchanged_count:
            messages.info(
                request,
                'لم يتم حفظ أي تعديل لأن الاختيارات المرسلة مطابقة للتوزيع الحالي. '
                'قد يظهر الطالب ضمن هذا الفلتر لأن حالة العرض محسوبة على مستوى الطالب، لا على كل مادة بشكل منفصل.'
            )
        else:
            messages.info(request, 'لا يوجد تغييرات جديدة للحفظ في هذه الصفحة.')

        if saved_count:
            request.session[self.flash_session_key] = {
                'course_type': payload['course_type'],
                'course_id': payload['selected_course_id'],
                'session_id': payload['selected_session_id'],
                'stage': payload['stage'],
                'assignment_status': payload['assignment_status'],
                'page': page_obj.number,
                'cells': changed_enrollment_details[:30],
            }

        redirect_url = reverse('quick:manual_sorting')
        redirect_query_items = [
            ('course_type', payload['course_type']),
            ('stage', payload['stage']),
            ('assignment_status', payload['assignment_status']),
            ('page', page_obj.number),
            ('_ts', int(time.time())),
        ]
        if payload['selected_course_id']:
            redirect_query_items.append(('course_id', payload['selected_course_id']))
        if payload['selected_session_id']:
            redirect_query_items.append(('session_id', payload['selected_session_id']))
        redirect_query = urlencode(redirect_query_items)
        logger.info('manual_sorting_post_redirect=%s?%s', redirect_url, redirect_query)
        return redirect(f'{redirect_url}?{redirect_query}')


@method_decorator(never_cache, name='dispatch')
class QuickManualSortingPrintView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_manual_sorting_print.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        payload = _build_quick_manual_sorting_payload(
            course_type=self.request.GET.get('course_type') or 'INTENSIVE',
            course_id=self.request.GET.get('course_id') or '',
            session_id=self.request.GET.get('session_id') or '',
            stage=self.request.GET.get('stage') or 'NON_NINTH',
            assignment_status=self.request.GET.get('assignment_status') or 'ALL',
        )
        context.update({
            **payload,
            'student_rows_all': payload['student_rows'],
            'total_students': len(payload['student_rows']),
            'total_courses': len(payload['course_columns']),
        })
        return context


@method_decorator(never_cache, name='dispatch')
class QuickManualSortingUnassignedPrintView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_manual_sorting_unassigned_print.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        payload = _build_quick_manual_sorting_payload(
            course_type=self.request.GET.get('course_type') or 'INTENSIVE',
            course_id=self.request.GET.get('course_id') or '',
            session_id=self.request.GET.get('session_id') or '',
            stage=self.request.GET.get('stage') or 'NON_NINTH',
            assignment_status=self.request.GET.get('assignment_status') or 'ALL',
        )
        context.update({
            **payload,
            'unassigned_rows': payload['unassigned_enrollments'],
            'student_rows_unassigned': payload['student_rows_unassigned_print'],
            'total_unassigned_students': payload['unassigned_student_count'],
            'total_unassigned_enrollments': payload['unassigned_enrollment_count'],
            'total_courses': len(payload['course_columns_unassigned_print']),
        })
        return context

class QuickCourseCreateView(LoginRequiredMixin, CreateView):
    model = QuickCourse
    form_class = QuickCourseForm
    template_name = 'quick/quick_course_form.html'
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'تم إضافة الدورة السريعة بنجاح')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('quick:course_detail', kwargs={'pk': self.object.pk})


def _quick_session_display_rows(course):
    today = timezone.localdate()
    sessions = list(
        course.sessions.filter(is_active=True)
        .prefetch_related('session_enrollments__enrollment__student')
        .order_by('start_date', 'start_time', 'title')
    )
    rows = []
    for session in sessions:
        rows.append({
            'session': session,
            'is_upcoming': session.start_date > today,
            'is_open': session.start_date <= today <= session.end_date,
            'is_finished': today > session.end_date,
            'attendance_taken_today': session.attendance_records.filter(attendance_date=today).count(),
        })
    return rows


def _dates_overlap(start_a, end_a, start_b, end_b):
    return start_a <= end_b and start_b <= end_a


def _times_overlap(start_a, end_a, start_b, end_b):
    if not end_a or not end_b:
        return start_a == start_b
    return start_a < end_b and start_b < end_a


def _session_conflicts_with_window(session, start_date, end_date, start_time, end_time):
    return (
        _dates_overlap(start_date, end_date, session.start_date, session.end_date)
        and _times_overlap(start_time, end_time, session.start_time, session.end_time)
    )


def _sessions_conflict(first_session, second_session):
    return _session_conflicts_with_window(
        second_session,
        first_session.start_date,
        first_session.end_date,
        first_session.start_time,
        first_session.end_time,
    )


def _build_quick_session_conflict_report(course_type='ALL', selected_course_ids=None):
    courses = QuickCourse.objects.filter(is_active=True).select_related('academic_year').order_by('name')
    if course_type != 'ALL':
        courses = courses.filter(course_type=course_type)
    if selected_course_ids:
        courses = courses.filter(id__in=selected_course_ids)

    course_list = list(courses)
    course_ids = [course.id for course in course_list]
    assignments = (
        QuickCourseSessionEnrollment.objects.filter(
            session__is_active=True,
            session__course_id__in=course_ids,
            enrollment__is_completed=False,
            enrollment__student__is_active=True,
        )
        .select_related(
            'enrollment__student',
            'session',
            'session__course',
            'session__course__academic_year',
        )
        .order_by(
            'enrollment__student__full_name',
            'session__start_date',
            'session__start_time',
            'session_id',
        )
    )

    assignments_by_student = defaultdict(list)
    for assignment in assignments:
        assignments_by_student[assignment.enrollment.student_id].append(assignment)

    conflict_rows = []
    grouped_students = []
    course_ids_with_conflicts = set()
    session_ids_with_conflicts = set()

    for student_assignments in assignments_by_student.values():
        student_conflicts = []
        for first_assignment, second_assignment in combinations(student_assignments, 2):
            first_session = first_assignment.session
            second_session = second_assignment.session
            if first_session.course_id == second_session.course_id:
                continue
            if not _sessions_conflict(first_session, second_session):
                continue

            overlap_start_date = max(first_session.start_date, second_session.start_date)
            overlap_end_date = min(first_session.end_date, second_session.end_date)
            overlap_start_time = max(first_session.start_time, second_session.start_time)
            if first_session.end_time and second_session.end_time:
                overlap_end_time = min(first_session.end_time, second_session.end_time)
            else:
                overlap_end_time = first_session.end_time or second_session.end_time

            row = {
                'student': first_assignment.enrollment.student,
                'first_assignment': first_assignment,
                'second_assignment': second_assignment,
                'first_session': first_session,
                'second_session': second_session,
                'overlap_start_date': overlap_start_date,
                'overlap_end_date': overlap_end_date,
                'overlap_start_time': overlap_start_time,
                'overlap_end_time': overlap_end_time,
            }
            conflict_rows.append(row)
            student_conflicts.append(row)
            course_ids_with_conflicts.update([first_session.course_id, second_session.course_id])
            session_ids_with_conflicts.update([first_session.id, second_session.id])

        if student_conflicts:
            grouped_students.append({
                'student': student_conflicts[0]['student'],
                'conflicts': student_conflicts,
                'conflicts_count': len(student_conflicts),
            })

    conflict_rows.sort(
        key=lambda row: (
            row['overlap_start_date'],
            row['overlap_start_time'],
            row['student'].full_name,
            row['first_session'].course.name,
        )
    )
    grouped_students.sort(
        key=lambda row: (-row['conflicts_count'], row['student'].full_name)
    )

    return {
        'courses': course_list,
        'grouped_students': grouped_students,
        'conflict_rows': conflict_rows,
        'unique_students_count': len(grouped_students),
        'unique_courses_count': len(course_ids_with_conflicts),
        'unique_sessions_count': len(session_ids_with_conflicts),
        'total_conflicts': len(conflict_rows),
    }


def _effective_capacity(option_max_capacity, room_max_capacity):
    limits = [value for value in [option_max_capacity, room_max_capacity] if value]
    return min(limits) if limits else 0


def _course_active_enrollment_count(course):
    return QuickEnrollment.objects.filter(
        course=course,
        is_completed=False,
        student__is_active=True,
    ).count()


def _find_available_room_for_option(option, needed_students=0):
    candidate_rooms = Classroom.objects.filter(class_type='course', is_active=True).order_by('name')
    room_rank = lambda room: (
        0 if _effective_capacity(option.max_capacity, getattr(room, 'max_capacity', 0)) >= needed_students else 1,
        -(_effective_capacity(option.max_capacity, getattr(room, 'max_capacity', 0)) or 0),
        room.name,
    )
    if option.preferred_room_id:
        candidate_rooms = sorted(
            list(candidate_rooms),
            key=lambda room: (
                0 if room.id == option.preferred_room_id else 1,
                *room_rank(room),
            )
        )
    else:
        candidate_rooms = sorted(list(candidate_rooms), key=room_rank)

    conflicting_sessions = QuickCourseSession.objects.filter(is_active=True).exclude(course=option.course)
    for room in candidate_rooms:
        room_conflict = conflicting_sessions.filter(room_id=room.id)
        room_has_overlap = any(
            _session_conflicts_with_window(
                session,
                option.start_date,
                option.end_date,
                option.start_time,
                option.end_time,
            )
            for session in room_conflict
        )
        if room_has_overlap:
            continue
        return room
    return option.preferred_room


def _student_has_conflict_for_session(student_id, session):
    other_sessions = (
        QuickCourseSession.objects.filter(
            is_active=True,
            session_enrollments__enrollment__student_id=student_id,
            session_enrollments__enrollment__is_completed=False,
            session_enrollments__enrollment__student__is_active=True,
        )
        .exclude(course=session.course)
        .distinct()
    )
    return any(_sessions_conflict(session, other_session) for other_session in other_sessions)


def _auto_assign_course_enrollments(course, user):
    sessions = list(
        course.sessions.filter(is_active=True)
        .order_by('start_date', 'start_time', 'id')
    )
    enrollments = list(
        QuickEnrollment.objects.filter(course=course, is_completed=False, student__is_active=True)
        .select_related('student')
        .order_by('enrollment_date', 'id')
    )

    QuickCourseSessionEnrollment.objects.filter(session__course=course).delete()
    if _quick_manual_selection_table_exists():
        QuickManualSortingSelection.objects.filter(enrollment__course=course).delete()

    assigned_count = 0
    unassigned_count = 0
    seats_used = {}

    for enrollment in enrollments:
        assigned = False
        for session in sessions:
            seats_used.setdefault(session.id, 0)
            if session.capacity and seats_used[session.id] >= session.capacity:
                continue
            if _student_has_conflict_for_session(enrollment.student_id, session):
                continue
            assignment = QuickCourseSessionEnrollment.objects.create(
                session=session,
                enrollment=enrollment,
                assigned_by=user,
            )
            _sync_quick_manual_sorting_selection(enrollment=enrollment, session=assignment.session, user=user)
            seats_used[session.id] += 1
            assigned_count += 1
            assigned = True
            break
        if not assigned:
            unassigned_count += 1

    under_minimum = sum(1 for session in sessions if session.enrolled_count and not session.meets_minimum_capacity)
    empty_sessions = sum(1 for session in sessions if session.enrolled_count == 0)
    return {
        'assigned_count': assigned_count,
        'unassigned_count': unassigned_count,
        'under_minimum': under_minimum,
        'empty_sessions': empty_sessions,
    }


def _assign_enrollment_to_available_session(enrollment, user=None):
    sessions = (
        QuickCourseSession.objects.filter(course=enrollment.course, is_active=True)
        .order_by('start_date', 'start_time', 'id')
    )
    for session in sessions:
        if session.capacity and session.enrolled_count >= session.capacity:
            continue
        if _student_has_conflict_for_session(enrollment.student_id, session):
            continue
        assignment, _created = QuickCourseSessionEnrollment.objects.update_or_create(
            enrollment=enrollment,
            defaults={'session': session, 'assigned_by': user},
        )
        _sync_quick_manual_sorting_selection(enrollment=enrollment, session=assignment.session, user=user)
        return assignment
    return None


def _assign_enrollment_to_specific_session(enrollment, session, user=None):
    if session.course_id != enrollment.course_id:
        raise ValueError('الكلاس المختار لا يتبع لهذه الدورة.')

    current_assignment = getattr(enrollment, 'session_assignment', None)
    if current_assignment and current_assignment.session_id == session.id:
        return current_assignment

    if session.capacity and session.enrolled_count >= session.capacity:
        raise ValueError('هذا الكلاس ممتلئ حالياً.')

    if _student_has_conflict_for_session(enrollment.student_id, session):
        raise ValueError('يوجد تعارض وقتي بين هذا الكلاس وأحد كلاسات الطالب الأخرى.')

    assignment, _created = QuickCourseSessionEnrollment.objects.update_or_create(
        enrollment=enrollment,
        defaults={'session': session, 'assigned_by': user},
    )
    _sync_quick_manual_sorting_selection(enrollment=enrollment, session=assignment.session, user=user)
    return assignment


def _build_quick_register_course_context(student):
    active_sessions_queryset = (
        QuickCourseSession.objects.filter(is_active=True)
        .select_related('room')
        .annotate(
            assigned_count=Count(
                'session_enrollments',
                filter=Q(
                    session_enrollments__enrollment__is_completed=False,
                    session_enrollments__enrollment__student__is_active=True,
                ),
            )
        )
        .order_by('start_date', 'start_time', 'id')
    )

    courses = list(
        QuickCourse.objects.filter(is_active=True, academic_year=student.academic_year)
        .prefetch_related(
            Prefetch('sessions', queryset=active_sessions_queryset, to_attr='active_sessions')
        )
        .order_by('name')
    )

    enrolled_course_ids = set(
        QuickEnrollment.objects.filter(student=student).values_list('course_id', flat=True)
    )

    course_session_catalog = {}
    for course in courses:
        session_rows = []
        for session in getattr(course, 'active_sessions', []):
            room_label = session.room.name if getattr(session, 'room', None) else (session.room_name or 'تحدد لاحقاً')
            session_rows.append({
                'id': session.id,
                'title': session.title,
                'code': session.display_code,
                'date_range': f'{session.start_date:%Y-%m-%d} - {session.end_date:%Y-%m-%d}',
                'time_range': (
                    f'{session.start_time:%H:%M}'
                    + (f' - {session.end_time:%H:%M}' if session.end_time else '')
                ),
                'meeting_days': session.meeting_days or 'يحدد حسب البرنامج',
                'room_name': room_label,
                'assigned_count': getattr(session, 'assigned_count', 0),
                'capacity': session.capacity or 0,
                'min_capacity': session.min_capacity or 0,
                'is_full': bool(session.capacity and getattr(session, 'assigned_count', 0) >= session.capacity),
                'is_upcoming': session.is_upcoming,
                'is_finished': session.is_finished,
            })
        course_session_catalog[str(course.id)] = session_rows

    return {
        'student': student,
        'courses': courses,
        'enrolled_course_ids': enrolled_course_ids,
        'already_enrolled_count': len(enrolled_course_ids),
        'available_courses_count': len(courses),
        'course_session_catalog': course_session_catalog,
        'print_receipts_url': None,
    }


def _format_quick_session_time_label(session):
    if session is None:
        return 'يحدد لاحقاً'

    time_range = session.start_time.strftime('%H:%M')
    if session.end_time:
        time_range += f' - {session.end_time.strftime("%H:%M")}'

    if session.meeting_days:
        return f'{session.meeting_days} | {time_range}'
    return time_range


def _generate_course_sessions_from_options(course, user):
    options = list(
        course.time_options.filter(is_active=True)
        .select_related('preferred_room')
        .order_by('priority', 'start_date', 'start_time', 'id')
    )

    QuickCourseSession.objects.filter(course=course).update(is_active=False)

    created_sessions = []
    skipped_options = []
    total_students = QuickEnrollment.objects.filter(course=course, is_completed=False, student__is_active=True).count()
    remaining_students = total_students
    remaining_options = len(options)
    for index, option in enumerate(options, start=1):
        target_students = ceil(remaining_students / remaining_options) if remaining_options else remaining_students
        room = _find_available_room_for_option(option, needed_students=target_students)
        if room is None and Classroom.objects.filter(class_type='course', is_active=True).exists():
            skipped_options.append(option.title)
            remaining_options -= 1
            continue
        room_capacity = getattr(room, 'max_capacity', 0) if room else 0
        effective_capacity = _effective_capacity(option.max_capacity, room_capacity)
        session = QuickCourseSession.objects.create(
            course=course,
            time_option=option,
            title=option.title or f"كلاس {index}",
            code=f"{course.id}-{index}",
            min_capacity=max(option.min_capacity or 1, getattr(room, 'min_capacity', 1) or 1),
            capacity=effective_capacity,
            start_date=option.start_date,
            end_date=option.end_date,
            start_time=option.start_time,
            end_time=option.end_time,
            meeting_days=option.meeting_days,
            room=room,
            room_name=(room.name if room else ''),
            notes=option.notes,
            is_active=True,
            created_by=user,
        )
        created_sessions.append(session)
        remaining_students = max(0, remaining_students - (effective_capacity or remaining_students))
        remaining_options -= 1
    return created_sessions, skipped_options


def _generate_schedule_for_courses(courses, user):
    summary = {
        'courses_processed': 0,
        'sessions_created': 0,
        'students_assigned': 0,
        'students_unassigned': 0,
        'courses_with_unassigned': [],
        'courses_without_sessions': [],
        'skipped_options': [],
    }
    ordered_courses = sorted(
        list(courses),
        key=lambda course: (-_course_active_enrollment_count(course), course.id),
    )
    for course in ordered_courses:
        created_sessions, skipped_options = _generate_course_sessions_from_options(course, user)
        assignment_result = _auto_assign_course_enrollments(course, user)
        summary['courses_processed'] += 1
        summary['sessions_created'] += len(created_sessions)
        summary['students_assigned'] += assignment_result['assigned_count']
        summary['students_unassigned'] += assignment_result['unassigned_count']
        if not created_sessions:
            summary['courses_without_sessions'].append(course.name)
        if assignment_result['unassigned_count']:
            summary['courses_with_unassigned'].append(course.name)
        summary['skipped_options'].extend(skipped_options)
    return summary


class QuickCourseDetailView(LoginRequiredMixin, DetailView):
    model = QuickCourse
    template_name = 'quick/quick_course_detail.html'
    context_object_name = 'course'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        course = self.object
        active_enrollments = (
            QuickEnrollment.objects.filter(course=course, is_completed=False, student__is_active=True)
            .select_related('student')
            .order_by('student__full_name')
        )
        assigned_ids = set(
            QuickCourseSessionEnrollment.objects.filter(session__course=course).values_list('enrollment_id', flat=True)
        )
        context.update({
            'session_form': QuickCourseSessionForm(),
            'session_rows': _quick_session_display_rows(course),
            'active_enrollments': active_enrollments,
            'unassigned_enrollments': [enrollment for enrollment in active_enrollments if enrollment.id not in assigned_ids],
            'transfer_form': QuickSessionTransferForm(course=course),
            'sessions_manage_url': reverse('quick:course_sessions_manage', kwargs={'course_id': course.id}),
            'today': timezone.localdate(),
        })
        return context


class QuickCourseSessionsManageView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_course_sessions_manage.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        course = get_object_or_404(QuickCourse.objects.select_related('academic_year'), pk=self.kwargs['course_id'], is_active=True)
        active_enrollments = (
            QuickEnrollment.objects.filter(course=course, is_completed=False, student__is_active=True)
            .select_related('student')
            .order_by('enrollment_date', 'id')
        )
        assigned_ids = set(
            QuickCourseSessionEnrollment.objects.filter(session__course=course).values_list('enrollment_id', flat=True)
        )
        context.update({
            'course': course,
            'session_form': QuickCourseSessionForm(),
            'time_option_form': QuickCourseTimeOptionForm(),
            'time_options': course.time_options.filter(is_active=True).select_related('preferred_room').order_by('priority', 'start_date', 'start_time'),
            'session_rows': _quick_session_display_rows(course),
            'active_enrollments': active_enrollments,
            'unassigned_enrollments': [enrollment for enrollment in active_enrollments if enrollment.id not in assigned_ids],
            'transfer_form': QuickSessionTransferForm(course=course),
            'today': timezone.localdate(),
        })
        return context


class QuickCourseTimeOptionsManageView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_course_time_options_manage.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        course = get_object_or_404(QuickCourse.objects.select_related('academic_year'), pk=self.kwargs['course_id'], is_active=True)
        options = course.time_options.filter(is_active=True).select_related('preferred_room').order_by('priority', 'start_date', 'start_time')
        context.update({
            'course': course,
            'time_option_form': QuickCourseTimeOptionForm(),
            'time_options': options,
            'rooms': Classroom.objects.filter(class_type='course', is_active=True).order_by('name'),
        })
        return context


@login_required
@require_POST
def quick_course_add_session(request, course_id):
    course = get_object_or_404(QuickCourse, pk=course_id, is_active=True)
    form = QuickCourseSessionForm(request.POST)
    if form.is_valid():
        session = form.save(commit=False)
        session.course = course
        session.created_by = request.user
        session.save()
        messages.success(request, 'تمت إضافة كلاس جديد للدورة.')
    else:
        messages.error(request, 'تعذر حفظ الصف. يرجى مراجعة الحقول المدخلة.')
    return redirect('quick:course_sessions_manage', course_id=course.id)


@login_required
@require_POST
def quick_course_add_time_option(request, course_id):
    course = get_object_or_404(QuickCourse, pk=course_id, is_active=True)
    form = QuickCourseTimeOptionForm(request.POST)
    if form.is_valid():
        option = form.save(commit=False)
        option.course = course
        option.created_by = request.user
        option.save()
        messages.success(request, 'تمت إضافة وقت متاح للدورة.')
    else:
        error_parts = []
        for field_name, field_errors in form.errors.items():
            label = form.fields.get(field_name).label if field_name in form.fields else 'التحقق العام'
            for error in field_errors:
                error_parts.append(f"{label}: {error}")
        messages.error(
            request,
            'تعذر حفظ الوقت المتاح. ' + ' | '.join(error_parts[:4]) if error_parts else 'تعذر حفظ الوقت المتاح.',
        )
    return redirect('quick:course_time_options_manage', course_id=course.id)


@login_required
@require_POST
def quick_course_generate_schedule(request, course_id):
    course = get_object_or_404(QuickCourse, pk=course_id, is_active=True)
    created_sessions, skipped_options = _generate_course_sessions_from_options(course, request.user)
    result = _auto_assign_course_enrollments(course, request.user)
    if skipped_options:
        messages.warning(request, f"تم تخطي بعض الخيارات بسبب تعارض القاعات أو الطلاب: {', '.join(skipped_options[:5])}")
    messages.success(
        request,
        f"تم توليد {len(created_sessions)} كلاس وتوزيع {result['assigned_count']} طالب تلقائياً."
    )
    return redirect('quick:course_sessions_manage', course_id=course.id)


@login_required
@require_POST
def quick_generate_all_schedules(request):
    active_courses = QuickCourse.objects.filter(is_active=True)
    courses_without_time_options = list(
        active_courses.filter(
            enrollments__is_completed=False,
            enrollments__student__is_active=True,
        )
        .exclude(time_options__is_active=True)
        .distinct()
        .order_by('name')
        .values_list('name', flat=True)
    )
    courses = (
        active_courses.filter(time_options__is_active=True)
        .distinct()
        .order_by('id')
    )
    summary = _generate_schedule_for_courses(courses, request.user)
    if courses_without_time_options:
        messages.warning(
            request,
            'هذه الدورات لديها طلاب لكن لا يوجد لها أي وقت متاح فعّال: '
            + ', '.join(courses_without_time_options[:8])
        )
    if summary['skipped_options']:
        messages.warning(request, 'تم تخطي بعض الأوقات بسبب التعارض: ' + ', '.join(summary['skipped_options'][:8]))
    if summary['courses_without_sessions']:
        messages.warning(
            request,
            'تعذر توليد أي كلاس لبعض الدورات رغم وجود أوقات متاحة: '
            + ', '.join(summary['courses_without_sessions'][:8])
        )
    if summary['students_unassigned']:
        messages.error(
            request,
            f"بقي {summary['students_unassigned']} طالب بدون كلاس في الدورات التالية: "
            + ', '.join(summary['courses_with_unassigned'][:8])
        )
    else:
        messages.success(
            request,
            f"تم توليد برنامج {summary['courses_processed']} دورة وإنشاء {summary['sessions_created']} كلاس "
            f"وتوزيع {summary['students_assigned']} طالب بدون أي طلاب غير موزعين."
        )
    return redirect('quick:course_list')


@login_required
@require_POST
def quick_course_session_extend(request, session_id):
    session = get_object_or_404(QuickCourseSession, pk=session_id, is_active=True)
    session.end_date = session.end_date + timedelta(days=1)
    session.save(update_fields=['end_date', 'updated_at'])
    messages.success(request, f'تمت إضافة يوم دوام إلى {session.title}.')
    return redirect('quick:course_sessions_manage', course_id=session.course_id)


@login_required
@require_POST
def quick_course_session_assign_students(request, session_id):
    session = get_object_or_404(QuickCourseSession, pk=session_id, is_active=True)
    form = QuickSessionAssignStudentsForm(request.POST, session=session)
    if not form.is_valid():
        messages.error(request, 'تعذر توزيع الطلاب على الصف.')
        return redirect('quick:course_session_students', session_id=session.id)

    selected = list(form.cleaned_data['enrollment_ids'])
    if session.capacity and selected:
        remaining = max(0, session.capacity - session.enrolled_count)
        if len(selected) > remaining:
            messages.error(request, f'السعة المتبقية في الصف هي {remaining} طالب فقط.')
            return redirect('quick:course_session_students', session_id=session.id)

    created_count = 0
    for enrollment in selected:
        assignment, created = QuickCourseSessionEnrollment.objects.get_or_create(
            enrollment=enrollment,
            defaults={'session': session, 'assigned_by': request.user},
        )
        if not created and assignment.session_id != session.id:
            assignment.session = session
            assignment.assigned_by = request.user
            assignment.save(update_fields=['session', 'assigned_by'])
            created = True
        _sync_quick_manual_sorting_selection(enrollment=enrollment, session=assignment.session, user=request.user)
        if created:
            created_count += 1

    messages.success(request, f'تم توزيع {created_count} طالب على الصف.')
    return redirect('quick:course_session_students', session_id=session.id)


@login_required
@require_POST
def quick_course_session_unassign_student(request, session_id, enrollment_id):
    session = get_object_or_404(
        QuickCourseSession.objects.select_related('course'),
        pk=session_id,
        is_active=True,
    )
    assignment = get_object_or_404(
        QuickCourseSessionEnrollment.objects.select_related('enrollment__student'),
        session_id=session.id,
        enrollment_id=enrollment_id,
    )
    redirect_url = request.POST.get('next') or reverse('quick:course_session_students', kwargs={'session_id': session.id})
    if not redirect_url.startswith('/'):
        redirect_url = reverse('quick:course_session_students', kwargs={'session_id': session.id})

    if session.attendance_records.filter(enrollment_id=assignment.enrollment_id).exists():
        messages.error(
            request,
            f'لا يمكن إزالة {assignment.enrollment.student.full_name} من {session.title} لأن لديه حضورًا مسجلًا داخل هذه الفترة.',
        )
        return redirect(redirect_url)

    with transaction.atomic():
        _sync_quick_manual_sorting_selection(
            enrollment=assignment.enrollment,
            session=None,
            user=request.user,
        )
        student_name = assignment.enrollment.student.full_name
        session_title = session.title
        assignment.delete()

    messages.success(request, f'تمت إزالة {student_name} من {session_title}.')
    return redirect(redirect_url)


@login_required
@require_POST
def quick_course_transfer_students(request, course_id):
    course = get_object_or_404(QuickCourse, pk=course_id, is_active=True)
    form = QuickSessionTransferForm(request.POST, course=course)
    if not form.is_valid():
        messages.error(request, 'تعذر نقل الطلاب بين الصفوف.')
        return redirect('quick:course_sessions_manage', course_id=course.id)

    source = form.cleaned_data['source_session']
    target = form.cleaned_data['target_session']
    enrollments = list(form.cleaned_data['enrollment_ids'])

    if target.capacity:
        remaining = max(0, target.capacity - target.enrolled_count)
        if len(enrollments) > remaining:
            messages.error(request, f'السعة المتبقية في الصف الهدف هي {remaining} طالب فقط.')
            return redirect('quick:course_sessions_manage', course_id=course.id)

    moved = 0
    for enrollment in enrollments:
        assignment = getattr(enrollment, 'session_assignment', None)
        if assignment and assignment.session_id == source.id:
            assignment.session = target
            assignment.assigned_by = request.user
            assignment.save(update_fields=['session', 'assigned_by'])
            _sync_quick_manual_sorting_selection(enrollment=enrollment, session=assignment.session, user=request.user)
            moved += 1

    messages.success(request, f'تم نقل {moved} طالب إلى {target.title}.')
    return redirect('quick:course_sessions_manage', course_id=course.id)


@login_required
@require_POST
def quick_course_auto_assign_students(request, course_id):
    course = get_object_or_404(QuickCourse, pk=course_id, is_active=True)
    result = _auto_assign_course_enrollments(course, request.user)
    messages.success(
        request,
        f"تم التوزيع التلقائي لـ {result['assigned_count']} طالب. "
        f"غير الموزعين: {result['unassigned_count']} | "
        f"كلاسات دون الحد الأدنى: {result['under_minimum']}."
    )
    return redirect('quick:course_sessions_manage', course_id=course.id)


class QuickCourseSessionStudentsView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_course_session_students.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        session = get_object_or_404(
            QuickCourseSession.objects.select_related('course', 'course__academic_year'),
            pk=self.kwargs['session_id'],
        )
        assignments = list(
            session.session_enrollments.select_related('enrollment__student')
            .order_by('enrollment__student__full_name')
        )
        attendance_enrollment_ids = set(
            session.attendance_records.values_list('enrollment_id', flat=True)
        )
        assign_form = QuickSessionAssignStudentsForm(session=session)
        context.update({
            'session': session,
            'course': session.course,
            'assignments': assignments,
            'assignment_rows': [
                {
                    'assignment': assignment,
                    'has_attendance': assignment.enrollment_id in attendance_enrollment_ids,
                }
                for assignment in assignments
            ],
            'assign_form': assign_form,
            'assigned_count': len(assignments),
            'available_enrollment_count': assign_form.fields['enrollment_ids'].queryset.count(),
            'can_manage_assignment_changes': (
                self.request.user.is_superuser
                or 'course_accounting_edit' in getattr(self.request, 'employee_permissions', set())
            ),
            'students_print_url': reverse('quick:course_session_students_print', kwargs={'session_id': session.id}),
            'today': timezone.localdate(),
        })
        return context


class QuickCourseSessionStudentsPrintView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_course_session_students_print.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        session = get_object_or_404(
            QuickCourseSession.objects.select_related('course', 'course__academic_year', 'room'),
            pk=self.kwargs['session_id'],
        )
        assignments = list(
            session.session_enrollments.select_related('enrollment__student')
            .order_by('enrollment__student__full_name')
        )
        attendance_enrollment_ids = set(
            session.attendance_records.values_list('enrollment_id', flat=True)
        )
        context.update({
            'session': session,
            'course': session.course,
            'assignments': assignments,
            'assignment_rows': [
                {
                    'assignment': assignment,
                    'has_attendance': assignment.enrollment_id in attendance_enrollment_ids,
                }
                for assignment in assignments
            ],
            'assigned_count': len(assignments),
            'generated_at': timezone.localtime(),
        })
        return context


class QuickAllSessionsManageView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_all_sessions_manage.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.localdate()
        search_query = (self.request.GET.get('q') or '').strip()
        status = (self.request.GET.get('status') or 'ALL').upper()
        course_type = _resolve_quick_course_type_value(self.request.GET.get('course_type') or 'ALL', allow_all=True)

        sessions = QuickCourseSession.objects.filter(is_active=True).select_related(
            'course',
            'course__academic_year',
            'room',
        )
        if course_type != 'ALL':
            sessions = sessions.filter(course__course_type=course_type)
        if search_query:
            sessions = sessions.filter(
                Q(title__icontains=search_query)
                | Q(course__name__icontains=search_query)
                | Q(room_name__icontains=search_query)
                | Q(room__name__icontains=search_query)
            )
        sessions = sessions.prefetch_related(
            Prefetch(
                'session_enrollments',
                queryset=QuickCourseSessionEnrollment.objects.select_related('enrollment__student').order_by('enrollment__student__full_name'),
                to_attr='prefetched_assignments',
            ),
            Prefetch(
                'attendance_records',
                queryset=QuickCourseSessionAttendance.objects.only('id', 'session_id', 'enrollment_id', 'attendance_date'),
                to_attr='prefetched_attendance_records',
            ),
        ).order_by('start_date', 'start_time', 'course__name', 'title')

        status_options = [
            {'value': 'ALL', 'label': 'كل الفترات'},
            {'value': 'LIVE', 'label': 'الجارية'},
            {'value': 'UPCOMING', 'label': 'القادمة'},
            {'value': 'FINISHED', 'label': 'المنتهية'},
        ]
        session_rows = []
        total_assigned_students = 0
        for session in sessions:
            if session.start_date > today:
                lifecycle = 'UPCOMING'
                lifecycle_label = 'قادمة'
            elif today > session.end_date:
                lifecycle = 'FINISHED'
                lifecycle_label = 'منتهية'
            else:
                lifecycle = 'LIVE'
                lifecycle_label = 'جارية'
            if status != 'ALL' and lifecycle != status:
                continue

            assignment_objects = list(getattr(session, 'prefetched_assignments', []))
            attendance_records = list(getattr(session, 'prefetched_attendance_records', []))
            attendance_enrollment_ids = {record.enrollment_id for record in attendance_records}
            assigned_count = len(assignment_objects)
            total_assigned_students += assigned_count
            session_rows.append({
                'session': session,
                'assigned_count': assigned_count,
                'attendance_count': len(attendance_enrollment_ids),
                'lifecycle': lifecycle,
                'lifecycle_label': lifecycle_label,
                'students_url': reverse('quick:course_session_students', kwargs={'session_id': session.id}),
                'students_print_url': reverse('quick:course_session_students_print', kwargs={'session_id': session.id}),
                'attendance_url': reverse('quick:course_session_attendance', kwargs={'session_id': session.id}),
                'course_manage_url': reverse('quick:course_sessions_manage', kwargs={'course_id': session.course_id}),
                'assignment_rows': [
                    {
                        'assignment': assignment,
                        'has_attendance': assignment.enrollment_id in attendance_enrollment_ids,
                    }
                    for assignment in assignment_objects
                ],
            })

        context.update({
            'today': today,
            'search_query': search_query,
            'status': status,
            'status_options': status_options,
            'course_type': course_type,
            'course_type_options': _get_course_type_options(),
            'session_rows': session_rows,
            'sessions_count': len(session_rows),
            'total_assigned_students': total_assigned_students,
            'can_manage_assignment_changes': (
                self.request.user.is_superuser
                or 'course_accounting_edit' in getattr(self.request, 'employee_permissions', set())
            ),
        })
        return context


class QuickCourseAttendanceDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_course_attendance_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.localdate()
        sessions = (
            QuickCourseSession.objects.filter(is_active=True)
            .select_related('course', 'course__academic_year')
            .prefetch_related(
                Prefetch(
                    'session_enrollments',
                    queryset=QuickCourseSessionEnrollment.objects.select_related('enrollment__student').order_by('enrollment__student__full_name'),
                    to_attr='prefetched_assignments',
                ),
                Prefetch(
                    'attendance_records',
                    queryset=QuickCourseSessionAttendance.objects.only('id', 'session_id', 'enrollment_id', 'attendance_date'),
                    to_attr='prefetched_attendance_records',
                ),
            )
            .order_by('start_date', 'start_time', 'course__name', 'title')
        )
        live_sessions = []
        upcoming_sessions = []
        archived_sessions = []
        can_manage_assignment_changes = (
            self.request.user.is_superuser
            or 'course_accounting_edit' in getattr(self.request, 'employee_permissions', set())
        )
        for session in sessions:
            assignment_objects = list(getattr(session, 'prefetched_assignments', []))
            attendance_records = list(getattr(session, 'prefetched_attendance_records', []))
            attendance_enrollment_ids = {record.enrollment_id for record in attendance_records}
            row = {
                'session': session,
                'assigned_count': len(assignment_objects),
                'attendance_taken_today': sum(1 for record in attendance_records if record.attendance_date == today),
                'current_day_number': session.get_day_number_for_date(min(today, session.end_date)) if today >= session.start_date else None,
                'assignment_rows': [
                    {
                        'assignment': assignment,
                        'has_attendance': assignment.enrollment_id in attendance_enrollment_ids,
                    }
                    for assignment in assignment_objects
                ],
            }
            if session.start_date <= today <= session.end_date:
                live_sessions.append(row)
            elif session.start_date > today:
                upcoming_sessions.append(row)
            else:
                archived_sessions.append(row)
        live_sessions.sort(key=lambda item: (item['session'].start_date, item['session'].start_time))
        upcoming_sessions.sort(key=lambda item: (item['session'].start_date, item['session'].start_time))
        archived_sessions.sort(key=lambda item: (item['session'].end_date, item['session'].start_time), reverse=True)
        context.update({
            'today': today,
            'live_sessions': live_sessions,
            'upcoming_sessions': upcoming_sessions,
            'archived_sessions': archived_sessions[:12],
            'can_manage_assignment_changes': can_manage_assignment_changes,
            'all_sessions_url': reverse('quick:all_sessions_manage'),
        })
        return context


class QuickCourseAttendanceArchiveView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_course_attendance_archive.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.localdate()
        archived_sessions = (
            QuickCourseSession.objects.filter(is_active=True, start_date__lt=today)
            .select_related('course', 'course__academic_year')
            .order_by('-start_date', '-start_time')
        )
        context['archived_sessions'] = [session for session in archived_sessions if session.is_finished]
        return context


class QuickCourseSessionAttendanceView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_course_session_attendance.html'

    def _get_session(self):
        return get_object_or_404(
            QuickCourseSession.objects.select_related('course', 'course__academic_year'),
            pk=self.kwargs['session_id'],
        )

    def _resolve_attendance_date(self, session):
        today = timezone.localdate()
        raw_value = self.request.GET.get('date') or self.request.POST.get('attendance_date')
        attendance_date = parse_date(raw_value) if raw_value else today
        if attendance_date is None:
            attendance_date = today
        if attendance_date < session.start_date:
            attendance_date = session.start_date
        if attendance_date > today:
            attendance_date = today
        if attendance_date > session.end_date:
            attendance_date = session.end_date
        return attendance_date

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        session = self._get_session()
        attendance_date = self._resolve_attendance_date(session)
        assignments = list(
            session.session_enrollments.select_related('enrollment__student')
            .order_by('enrollment__student__full_name')
        )
        records = {
            record.enrollment_id: record
            for record in session.attendance_records.filter(attendance_date=attendance_date)
        }
        form = QuickSessionAttendanceBulkForm(
            initial={'attendance_date': attendance_date},
            session=session,
            assignments=assignments,
        )
        for assignment in assignments:
            record = records.get(assignment.enrollment_id)
            if not record:
                continue
            prefix = f"student_{assignment.enrollment_id}"
            form.fields[f"{prefix}_status"].initial = record.status
            form.fields[f"{prefix}_notes"].initial = record.notes

        attendance_rows = []
        for assignment in assignments:
            prefix = f"student_{assignment.enrollment_id}"
            attendance_rows.append({
                'assignment': assignment,
                'status_field': form[f"{prefix}_status"],
                'notes_field': form[f"{prefix}_notes"],
            })

        context.update({
            'session': session,
            'course': session.course,
            'attendance_form': form,
            'assignments': assignments,
            'attendance_rows': attendance_rows,
            'attendance_date': attendance_date,
            'day_number': session.get_day_number_for_date(attendance_date),
            'today': timezone.localdate(),
            'can_take_attendance': attendance_date >= session.start_date,
            'max_attendance_date': min(timezone.localdate(), session.end_date),
        })
        return context

    def post(self, request, *args, **kwargs):
        session = self._get_session()
        attendance_date = self._resolve_attendance_date(session)
        assignments = list(
            session.session_enrollments.select_related('enrollment__student')
            .order_by('enrollment__student__full_name')
        )
        if attendance_date < session.start_date:
            messages.error(request, 'لا يمكن أخذ الحضور قبل بداية الصف.')
            return redirect('quick:course_session_attendance', session_id=session.id)

        form = QuickSessionAttendanceBulkForm(request.POST, session=session, assignments=assignments)
        if not form.is_valid():
            messages.error(request, 'تعذر حفظ الحضور.')
            return self.get(request, *args, **kwargs)

        day_number = session.get_day_number_for_date(attendance_date) or 1
        saved = 0
        for assignment in assignments:
            prefix = f"student_{assignment.enrollment_id}"
            status = form.cleaned_data.get(f"{prefix}_status") or 'present'
            notes = form.cleaned_data.get(f"{prefix}_notes", '')
            QuickCourseSessionAttendance.objects.update_or_create(
                session=session,
                enrollment=assignment.enrollment,
                attendance_date=attendance_date,
                defaults={
                    'day_number': day_number,
                    'status': status,
                    'notes': notes,
                    'created_by': request.user,
                },
            )
            saved += 1

        messages.success(request, f'تم حفظ حضور {saved} طالب لليوم رقم {day_number}.')
        return redirect(f"{reverse('quick:course_session_attendance', kwargs={'session_id': session.id})}?date={attendance_date.isoformat()}")


class QuickCourseSchedulePrintView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_course_schedule_print.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_ids = []
        for raw_value in self.request.GET.getlist('course_ids'):
            try:
                selected_ids.append(int(raw_value))
            except (TypeError, ValueError):
                continue
        course_type = self.request.GET.get('course_type') or 'INTENSIVE'
        courses = QuickCourse.objects.filter(is_active=True, course_type=course_type).select_related('academic_year')
        if selected_ids:
            courses = courses.filter(id__in=selected_ids)
        sessions = (
            QuickCourseSession.objects.filter(course__in=courses, is_active=True)
            .select_related('course', 'course__academic_year')
            .order_by('start_date', 'start_time', 'course__name', 'title')
        )
        total_students = 0
        sessions_by_course = defaultdict(list)
        for session in sessions:
            assigned_count = session.enrolled_count
            total_students += assigned_count
            sessions_by_course[session.course_id].append({
                'session': session,
                'assigned_count': assigned_count,
                'seat_utilization': (
                    round((assigned_count / session.capacity) * 100)
                    if session.capacity else 0
                ),
            })
        course_cards = []
        for course in courses.order_by('name'):
            course_sessions = sessions_by_course.get(course.id, [])
            course_cards.append({
                'course': course,
                'sessions': course_sessions,
                'students_count': sum(item['assigned_count'] for item in course_sessions),
            })
        context.update({
            'courses': courses.order_by('name'),
            'sessions': sessions,
            'course_cards': course_cards,
            'selected_course_ids': selected_ids,
            'selected_course_type': course_type,
            'course_type_choices': QuickCourse.COURSE_TYPE_CHOICES,
            'total_courses': courses.count(),
            'total_sessions': sessions.count(),
            'total_students': total_students,
            'generated_at': timezone.localtime(),
        })
        return context


class QuickCourseConflictReportView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_course_conflicts_report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_ids = []
        for raw_value in self.request.GET.getlist('course_ids'):
            try:
                selected_ids.append(int(raw_value))
            except (TypeError, ValueError):
                continue
        course_type = self.request.GET.get('course_type') or 'ALL'
        report = _build_quick_session_conflict_report(
            course_type=course_type,
            selected_course_ids=selected_ids,
        )
        context.update({
            'available_courses': report['courses'],
            'grouped_students': report['grouped_students'],
            'conflict_rows': report['conflict_rows'],
            'unique_students_count': report['unique_students_count'],
            'unique_courses_count': report['unique_courses_count'],
            'unique_sessions_count': report['unique_sessions_count'],
            'total_conflicts': report['total_conflicts'],
            'selected_course_ids': selected_ids,
            'selected_course_type': course_type,
            'course_type_choices': [('ALL', 'كل الأنواع')] + list(QuickCourse.COURSE_TYPE_CHOICES),
            'generated_at': timezone.localtime(),
        })
        return context


class QuickCourseSessionCountsReportView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_course_session_counts_report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(_build_quick_session_population_report(self.request))
        _attach_quick_report_urls(
            context,
            self.request,
            'quick:quick_course_session_counts_report',
            'quick:quick_course_session_counts_report_print',
        )
        return context


class QuickFreeStudentsReportView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_free_students_report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(_build_quick_free_students_report(self.request))
        _attach_quick_report_urls(
            context,
            self.request,
            'quick:quick_free_students_report',
            'quick:quick_free_students_report_print',
        )
        return context


class QuickCourseSessionCountsReportPrintView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_course_session_counts_report_print.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(_build_quick_session_population_report(self.request))
        _attach_quick_report_urls(
            context,
            self.request,
            'quick:quick_course_session_counts_report',
            'quick:quick_course_session_counts_report_print',
        )
        return context


class QuickFreeStudentsReportPrintView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/quick_free_students_report_print.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(_build_quick_free_students_report(self.request))
        _attach_quick_report_urls(
            context,
            self.request,
            'quick:quick_free_students_report',
            'quick:quick_free_students_report_print',
        )
        return context

# الطلاب السريعين
class QuickStudentListView(LoginRequiredMixin, ListView):
    model = QuickStudent
    template_name = 'quick/quick_student_list.html'
    context_object_name = 'students'
    
    def get_queryset(self):
        academic_year_id = self.request.GET.get('academic_year')
        gender = self.request.GET.get('gender')
        queryset = QuickStudent.objects.filter(is_active=True).select_related('student', 'academic_year')
        if academic_year_id:
            queryset = queryset.filter(academic_year_id=academic_year_id)
        if gender in ('male', 'female'):
            queryset = queryset.filter(student__gender=gender)
        elif gender == 'unknown':
            queryset = queryset.filter(Q(student__gender__isnull=True) | Q(student__gender=''))
        return queryset

    def post(self, request, *args, **kwargs):
        gender = request.POST.get('gender')
        student_ids = request.POST.getlist('student_ids')
        next_url = request.POST.get('next') or reverse('quick:student_list')

        if not student_ids:
            messages.warning(request, 'يرجى اختيار طلاب أولاً.')
            return redirect(next_url)

        if gender not in ('male', 'female', 'unknown'):
            messages.error(request, 'قيمة الجنس غير صحيحة.')
            return redirect(next_url)

        gender_value = '' if gender == 'unknown' else gender
        queryset = QuickStudent.objects.filter(id__in=student_ids).select_related('student')
        updated_count = 0
        for quick_student in queryset:
            student = getattr(quick_student, 'student', None)
            if student and student.gender != gender_value:
                student.gender = gender_value
                student.save(update_fields=['gender'])
                updated_count += 1

        if gender_value:
            messages.success(request, f'تم تحديث الجنس لـ {updated_count} طالب/طالبة.')
        else:
            messages.success(request, f'تم إزالة تحديد الجنس لـ {updated_count} طالب/طالبة.')
        return redirect(next_url)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # إحصائيات الربط التلقائي
        students = context['students']
        auto_assigned = students.filter(academic_year__isnull=False)
        unassigned = students.filter(academic_year__isnull=True)
        
        context.update({
            'academic_years': AcademicYear.objects.all().order_by('-start_date'),
            'auto_assigned_count': auto_assigned.count(),
            'unassigned_count': unassigned.count(),
        })
        return context


@require_superuser
def quick_duplicate_students_report(request):
    if request.method == 'POST':
        action = request.POST.get('action') or 'merge_one'
        group_key = _normalize_quick_student_name(request.POST.get('group_name'))
        search_query = (request.POST.get('q') or '').strip()
        phone_query = (request.POST.get('phone') or '').strip()
        scope = request.POST.get('scope') or 'active'

        if action == 'merge_all':
            duplicate_groups = _get_duplicate_groups(
                search_query=search_query,
                phone_query=phone_query,
                scope=scope,
            )
            if not duplicate_groups:
                messages.info(request, 'لا توجد مجموعات مكررة مطابقة للفلتر الحالي.')
                return redirect(
                    f"{reverse('quick:duplicate_students_report')}?{urlencode({'q': search_query, 'phone': phone_query, 'scope': scope})}"
                )

            merged_groups = 0
            total_enrollments = 0
            total_receipts = 0
            total_reversed = 0
            failed_groups = []

            for group in duplicate_groups:
                try:
                    merge_result = _merge_quick_students_by_name_with_retry(group['normalized_name'], request.user)
                except Exception as exc:
                    failed_groups.append(f'{group["display_name"]}: {exc}')
                    continue

                merged_groups += 1
                total_enrollments += merge_result['merged_enrollments']
                total_receipts += merge_result['merged_receipts']
                total_reversed += merge_result['reversed_duplicates']

            if merged_groups:
                messages.success(
                    request,
                    f'تم الدمج الجماعي لـ {merged_groups} مجموعة، '
                    f'ونُقل {total_enrollments} تسجيل و{total_receipts} إيصال، '
                    f'مع عكس {total_reversed} تسجيل مكرر.'
                )
            if failed_groups:
                for error in failed_groups[:10]:
                    messages.error(request, f'فشل دمج السجلات المكررة: {error}')
        else:
            if not group_key:
                messages.error(request, 'لم يتم تحديد الاسم المطلوب دمجه.')
                return redirect(
                    f"{reverse('quick:duplicate_students_report')}?{urlencode({'q': search_query, 'phone': phone_query, 'scope': scope})}"
                )

            try:
                merge_result = _merge_quick_students_by_name_with_retry(group_key, request.user)
            except Exception as exc:
                messages.error(request, f'فشل دمج السجلات المكررة: {exc}')
            else:
                messages.success(
                    request,
                    'تم الدمج بنجاح إلى السجل '
                    f'#{merge_result["target"].id}، '
                    f'ونُقل {merge_result["merged_enrollments"]} تسجيل و{merge_result["merged_receipts"]} إيصال، '
                    f'مع عكس {merge_result["reversed_duplicates"]} تسجيل مكرر على نفس الدورة.'
                )

        return redirect(
            f"{reverse('quick:duplicate_students_report')}?{urlencode({'q': search_query, 'phone': phone_query, 'scope': scope})}"
        )

    search_query = (request.GET.get('q') or '').strip()
    phone_query = (request.GET.get('phone') or '').strip()
    scope = request.GET.get('scope') or 'active'
    include_inactive = scope == 'all'

    enrollment_queryset = (
        QuickEnrollment.objects.select_related('course')
        .prefetch_related(
            Prefetch(
                'quickstudentreceipt_set',
                queryset=QuickStudentReceipt.objects.select_related('created_by', 'journal_entry').order_by('date', 'id')
            )
        )
        .annotate(
            paid_total=Coalesce(
                Sum('quickstudentreceipt__paid_amount'),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=10, decimal_places=2),
            )
        )
        .order_by('course__name', 'id')
    )

    students_queryset = (
        QuickStudent.objects.select_related('student', 'academic_year', 'created_by')
        .prefetch_related(
            Prefetch('enrollments', queryset=enrollment_queryset),
            Prefetch(
                'quickstudentreceipt_set',
                queryset=QuickStudentReceipt.objects.select_related(
                    'created_by', 'journal_entry', 'course', 'quick_enrollment'
                ).order_by('date', 'id')
            )
        )
        .order_by('full_name', 'id')
    )
    if not include_inactive:
        students_queryset = students_queryset.filter(is_active=True)

    grouped_students = defaultdict(list)
    for quick_student in students_queryset:
        normalized_name = _normalize_quick_student_name(quick_student.full_name)
        if normalized_name:
            grouped_students[normalized_name].append(quick_student)

    duplicate_groups = []
    duplicate_records_count = 0
    total_balance = Decimal('0')
    total_remaining = Decimal('0')
    total_enrollments = 0
    normalized_search = _normalize_quick_student_name(search_query)
    normalized_phone_search = _normalize_phone(phone_query)

    for normalized_name, students in grouped_students.items():
        if len(students) < 2:
            continue
        if normalized_search and normalized_search not in normalized_name:
            continue
        if normalized_phone_search and not any(
            normalized_phone_search in _normalize_phone(student.phone)
            for student in students
        ):
            continue

        members = []
        group_balance = Decimal('0')
        group_remaining = Decimal('0')
        group_enrollments = 0

        for quick_student in students:
            enrollments_data = []
            student_remaining = Decimal('0')
            account_created_by = '-'
            if quick_student.created_by:
                account_created_by = quick_student.created_by.get_full_name() or quick_student.created_by.username or '-'
            all_student_receipts = list(quick_student.quickstudentreceipt_set.all())
            used_receipt_ids = set()
            enrollments = list(quick_student.enrollments.all())
            enrollment_entry_refs = {f'QE-{enrollment.id}': enrollment for enrollment in enrollments}
            journal_entries = {
                entry.reference: entry
                for entry in JournalEntry.objects.filter(reference__in=enrollment_entry_refs.keys()).select_related('created_by')
            }

            for enrollment in enrollments:
                net_amount = enrollment.net_amount or Decimal('0')
                paid_amount = enrollment.paid_total or Decimal('0')
                remaining_amount = max(Decimal('0'), net_amount - paid_amount)
                student_remaining += remaining_amount
                group_enrollments += 1
                enrollment_created_by = '-'
                receipt_rows = []

                enrollment_entry = journal_entries.get(f'QE-{enrollment.id}')
                if enrollment_entry and enrollment_entry.created_by:
                    enrollment_created_by = (
                        enrollment_entry.created_by.get_full_name()
                        or enrollment_entry.created_by.username
                        or '-'
                    )

                receipts = [
                    receipt for receipt in all_student_receipts
                    if receipt.quick_enrollment_id == enrollment.id
                    or (
                        receipt.quick_enrollment_id is None
                        and receipt.course_id == enrollment.course_id
                    )
                ]
                if enrollment_created_by == '-' and receipts:
                    first_receipt_creator = receipts[0].created_by
                    if first_receipt_creator:
                        enrollment_created_by = (
                            first_receipt_creator.get_full_name()
                            or first_receipt_creator.username
                            or '-'
                        )

                for receipt in receipts:
                    receipt_created_by = '-'
                    if receipt.created_by:
                        receipt_created_by = receipt.created_by.get_full_name() or receipt.created_by.username or '-'
                    used_receipt_ids.add(receipt.id)
                    receipt_rows.append({
                        'id': receipt.id,
                        'receipt_number': receipt.receipt_number or '-',
                        'date': receipt.date,
                        'paid_amount': receipt.paid_amount or Decimal('0'),
                        'amount': receipt.amount or Decimal('0'),
                        'is_printed': receipt.is_printed,
                        'created_by': receipt_created_by,
                    })

                enrollments_data.append({
                    'course_name': enrollment.course.name if enrollment.course else '-',
                    'enrollment_date': enrollment.enrollment_date,
                    'net_amount': net_amount,
                    'paid_amount': paid_amount,
                    'remaining_amount': remaining_amount,
                    'is_completed': enrollment.is_completed,
                    'registered_by': enrollment_created_by,
                    'receipts': receipt_rows,
                    'has_receipts': bool(receipt_rows),
                    'receipts_count': len(receipt_rows),
                })

            orphan_receipts = []
            for receipt in all_student_receipts:
                if receipt.id in used_receipt_ids:
                    continue
                receipt_created_by = '-'
                if receipt.created_by:
                    receipt_created_by = receipt.created_by.get_full_name() or receipt.created_by.username or '-'
                orphan_receipts.append({
                    'id': receipt.id,
                    'receipt_number': receipt.receipt_number or '-',
                    'date': receipt.date,
                    'course_name': receipt.course_name or (receipt.course.name if receipt.course else '-'),
                    'paid_amount': receipt.paid_amount or Decimal('0'),
                    'amount': receipt.amount or Decimal('0'),
                    'is_printed': receipt.is_printed,
                    'created_by': receipt_created_by,
                })

            account_balance = quick_student.balance
            group_balance += account_balance
            group_remaining += student_remaining

            members.append({
                'student': quick_student,
                'account_created_by': account_created_by,
                'account_balance': account_balance,
                'remaining_total': student_remaining,
                'enrollments': enrollments_data,
                'enrollments_count': len(enrollments_data),
                'orphan_receipts': orphan_receipts,
                'has_orphan_receipts': bool(orphan_receipts),
            })

        members.sort(key=lambda item: item['student'].id)
        duplicate_groups.append({
            'display_name': students[0].full_name,
            'normalized_name': normalized_name,
            'students': members,
            'duplicate_count': len(members),
            'group_balance': group_balance,
            'group_remaining': group_remaining,
            'group_enrollments': group_enrollments,
        })

        duplicate_records_count += len(members)
        total_balance += group_balance
        total_remaining += group_remaining
        total_enrollments += group_enrollments

    duplicate_groups.sort(key=lambda item: (-item['duplicate_count'], item['display_name']))

    context = {
        'duplicate_groups': duplicate_groups,
        'search_query': search_query,
        'phone_query': phone_query,
        'scope': scope,
        'duplicate_names_count': len(duplicate_groups),
        'duplicate_records_count': duplicate_records_count,
        'total_balance': total_balance,
        'total_remaining': total_remaining,
        'total_enrollments': total_enrollments,
    }
    return render(request, 'quick/quick_duplicate_students_report.html', context)


@require_superuser
def quick_duplicate_students_print(request):
    group_key = _normalize_quick_student_name(request.GET.get('name'))
    phone_query = (request.GET.get('phone') or '').strip()
    scope = request.GET.get('scope') or 'active'
    duplicate_groups = _get_duplicate_groups(phone_query=phone_query, scope=scope)
    group = next((item for item in duplicate_groups if item['normalized_name'] == group_key), None)
    if not group:
        raise Http404('Duplicate group not found')

    return render(request, 'quick/quick_duplicate_students_print.html', {
        'group': group,
        'phone_query': phone_query,
        'scope': scope,
        'print_date': timezone.now(),
    })


@require_superuser
def quick_duplicate_students_full_print(request):
    search_query = (request.GET.get('q') or '').strip()
    phone_query = (request.GET.get('phone') or '').strip()
    scope = request.GET.get('scope') or 'active'
    duplicate_groups = _get_duplicate_groups(
        search_query=search_query,
        phone_query=phone_query,
        scope=scope,
    )

    total_balance = sum((group['group_balance'] for group in duplicate_groups), Decimal('0'))
    total_remaining = sum((group['group_remaining'] for group in duplicate_groups), Decimal('0'))
    total_enrollments = sum((group['group_enrollments'] for group in duplicate_groups), 0)
    total_records = sum((group['duplicate_count'] for group in duplicate_groups), 0)

    return render(request, 'quick/quick_duplicate_students_full_print.html', {
        'duplicate_groups': duplicate_groups,
        'search_query': search_query,
        'phone_query': phone_query,
        'scope': scope,
        'duplicate_names_count': len(duplicate_groups),
        'duplicate_records_count': total_records,
        'total_balance': total_balance,
        'total_remaining': total_remaining,
        'total_enrollments': total_enrollments,
        'print_date': timezone.now(),
    })


@require_superuser
def quick_accounting_fix_tool(request):
    if request.method == 'POST':
        result = _apply_quick_accounting_fixes(request.user)
        if result['fixed_links'] or result['fixed_withdrawals'] or result['fixed_receipts']:
            messages.success(
                request,
                f'تم تنفيذ التصحيح: ربط {result["fixed_links"]} قيد تسجيل '
                f'وتصحيح {result["fixed_receipts"]} قيد قبض '
                f'وإلغاء/تنظيف {result["cleaned_withdraw_entries"]} قيد سحب قديم '
                f'وإنشاء {result["fixed_withdrawals"]} قيد سحب جديد.'
            )
        else:
            messages.info(request, 'لم يتم العثور على قيود سريعة تحتاج تصحيحاً تلقائياً.')

        if result['deactivated_legacy_accounts']:
            messages.info(
                request,
                'تم تعطيل حسابات الانسحاب القديمة بدون حذف أي بيانات: '
                + ', '.join(result['deactivated_legacy_accounts'])
            )

        for error in result['errors'][:10]:
            messages.error(request, error)
        return redirect('quick:accounting_fix_tool')

    rows = _build_quick_accounting_fix_rows()
    issue_rows = [row for row in rows if row['issues']]
    context = {
        'rows': rows,
        'issue_rows': issue_rows,
        'audited_count': len(rows),
        'issues_count': len(issue_rows),
        'compliant_count': sum(1 for row in rows if row['is_compliant']),
        'missing_entry_count': sum(1 for row in rows if row['missing_entry']),
        'withdraw_fix_count': sum(
            1 for row in issue_rows
            if row['legacy_entries_count'] > 0 and row['correction_amount'] > 0
        ),
        'total_correction_amount': sum((row['correction_amount'] for row in issue_rows), Decimal('0')),
    }
    return render(request, 'quick/quick_accounting_fix_tool.html', context)


@require_superuser
def quick_withdrawal_fix_tool(request):
    if request.method == 'POST':
        result = _apply_quick_withdrawal_fixes(request.user)
        if result['fixed_withdrawals'] or result['cleaned_withdraw_entries']:
            messages.success(
                request,
                f'تم تنفيذ تصحيح قيود السحب: حذف/تنظيف {result["cleaned_withdraw_entries"]} قيد سحب قديم '
                f'وإنشاء {result["fixed_withdrawals"]} قيد عكس جديد.'
            )
        else:
            messages.info(request, 'لم يتم العثور على قيود سحب سريعة تحتاج تصحيحاً.')

        if result['deactivated_legacy_accounts']:
            messages.info(
                request,
                'تم تعطيل حسابات الانسحاب القديمة بدون حذف أي بيانات: '
                + ', '.join(result['deactivated_legacy_accounts'])
            )

        for error in result['errors'][:10]:
            messages.error(request, error)
        return redirect('quick:withdrawal_fix_tool')

    rows = _build_quick_withdrawal_fix_rows()
    context = {
        'rows': rows,
        'audited_count': len(rows),
        'legacy_count': sum(1 for row in rows if row['legacy_entries_count'] > 0),
        'retained_count': sum(1 for row in rows if row['retained_amount'] > 0),
        'missing_withdraw_count': sum(
            1 for row in rows if 'تسجيل مسحوب بدون قيد سحب واضح' in row['issues']
        ),
    }
    return render(request, 'quick/quick_withdrawal_fix_tool.html', context)


def _get_existing_quick_student_ar_account(student):
    if not student or not getattr(student, 'id', None):
        return None
    return Account.objects.filter(code=f'1252-{student.id:03d}').first()


def _get_quick_student_related_journal_entries(student):
    ar_account = _get_existing_quick_student_ar_account(student)
    if not ar_account:
        return JournalEntry.objects.none()

    entry_ids = Transaction.objects.filter(
        account=ar_account
    ).values_list('journal_entry_id', flat=True).distinct()
    return JournalEntry.objects.filter(id__in=entry_ids).distinct()


def _get_quick_student_delete_summary(student):
    enrollments_count = QuickEnrollment.objects.filter(student=student).count()
    receipts_count = QuickStudentReceipt.objects.filter(quick_student=student).count()
    print_jobs_count = QuickReceiptPrintJob.objects.filter(quick_student=student).count()
    journal_entries = _get_quick_student_related_journal_entries(student)
    journal_entries_count = journal_entries.count()
    transactions_count = Transaction.objects.filter(journal_entry__in=journal_entries).count()
    ar_account = _get_existing_quick_student_ar_account(student)

    return {
        'enrollments_count': enrollments_count,
        'receipts_count': receipts_count,
        'print_jobs_count': print_jobs_count,
        'journal_entries_count': journal_entries_count,
        'transactions_count': transactions_count,
        'ar_account_code': ar_account.code if ar_account else '-',
        'account_codes_display': ar_account.code if ar_account else '-',
    }


@login_required
def quick_delete_student(request, student_id):
    student = get_object_or_404(QuickStudent.objects.select_related('student'), pk=student_id)
    summary = _get_quick_student_delete_summary(student)

    if not request.user.is_superuser and not request.user.has_perm('quick.change_quickstudent'):
        if request.method == 'GET':
            return JsonResponse({'success': False, 'error': 'لا تملك صلاحية حذف الطلاب السريعين.'}, status=403)
        messages.error(request, 'لا تملك صلاحية حذف الطلاب السريعين.')
        return redirect('quick:student_profile', student_id=student.id)

    if request.method == 'GET':
        return JsonResponse({
            'success': True,
            'student_name': student.full_name,
            'summary': summary,
        })

    redirect_url = reverse('quick:student_list')
    linked_student = getattr(student, 'student', None)
    ar_account = _get_existing_quick_student_ar_account(student)
    journal_entries = list(_get_quick_student_related_journal_entries(student))

    with transaction.atomic():
        if journal_entries:
            JournalEntry.objects.filter(id__in=[entry.id for entry in journal_entries]).delete()

        student.delete()

        if linked_student:
            linked_student.delete()

        if ar_account and not Transaction.objects.filter(account=ar_account).exists():
            ar_account.delete()

    Account.rebuild_all_balances()

    messages.success(
        request,
        f'تم حذف الطالب السريع "{student.full_name}" مع {summary["enrollments_count"]} تسجيل، '
        f'{summary["receipts_count"]} إيصال، {summary["journal_entries_count"]} قيد، '
        f'و{summary["transactions_count"]} حركة مالية.'
    )
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'message': f'تم حذف الطالب السريع "{student.full_name}" مع كل البيانات المرتبطة به.',
        })
    return redirect(redirect_url)

class QuickStudentCreateView(LoginRequiredMixin, CreateView):
    model = QuickStudent
    form_class = QuickStudentForm
    template_name = 'quick/quick_student_form.html'

    def get_initial(self):
        initial = super().get_initial()
        active_year = AcademicYear.objects.filter(is_active=True).order_by('-start_date').first()
        if active_year:
            initial.setdefault('academic_year', active_year.id)
        return initial

    def form_valid(self, form):
        # إنشاء طالب نظامي أولاً
        from students.models import Student
        student = Student.objects.create(
            full_name=form.cleaned_data['full_name'],
            phone=form.cleaned_data['phone'],
            email='',
            gender=form.cleaned_data.get('gender', '') or '',
            is_active=True,
            added_by=self.request.user
        )
        
        form.instance.student = student
        form.instance.created_by = self.request.user
        messages.success(self.request, 'تم إضافة الطالب السريع بنجاح')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('quick:student_profile', kwargs={'student_id': self.object.pk})


@require_GET
@login_required
def quick_student_exists(request):
    field = (request.GET.get('field') or '').strip()
    value = (request.GET.get('value') or '').strip()
    exclude_id = request.GET.get('exclude_id')

    queryset = QuickStudent.objects.all().only('id', 'full_name', 'phone')
    if exclude_id and exclude_id.isdigit():
        queryset = queryset.exclude(pk=int(exclude_id))

    match = None
    if field == 'full_name' and value:
        normalized_value = _normalize_quick_name(value)
        match = next(
            (student for student in queryset if _normalize_quick_name(student.full_name) == normalized_value),
            None
        )
    elif field == 'phone' and value:
        normalized_value = _normalize_quick_phone(value)
        match = next(
            (student for student in queryset if _normalize_quick_phone(student.phone) == normalized_value),
            None
        )

    return JsonResponse({
        'exists': bool(match),
        'field': field,
        'full_name': match.full_name if match else '',
        'phone': match.phone if match else '',
        'id': match.id if match else None,
    })

class QuickStudentDetailView(LoginRequiredMixin, DetailView):
    model = QuickStudent
    template_name = 'quick/quick_student_detail.html'
    context_object_name = 'student'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['enrollments'] = QuickEnrollment.objects.filter(
            student=self.object
        ).select_related('course')
        return context

# التسجيلات السريعة
class QuickEnrollmentCreateView(LoginRequiredMixin, CreateView):
    model = QuickEnrollment
    form_class = QuickEnrollmentForm
    template_name = 'quick/quick_enrollment_form.html'
    
    def get_initial(self):
        initial = super().get_initial()
        student_id = self.request.GET.get('student')
        course_id = self.request.GET.get('course')
        
        if student_id:
            initial['student'] = student_id
        if course_id:
            course = get_object_or_404(QuickCourse, id=course_id)
            initial['course'] = course_id
            initial['net_amount'] = course.price
        
        return initial
    
    def form_valid(self, form):
        response = super().form_valid(form)
        # إنشاء القيد المحاسبي
        try:
            self.object.create_accrual_enrollment_entry(self.request.user)
            messages.success(
                self.request,
                'تم تسجيل الطالب وإنشاء القيد المحاسبي بنجاح. بقي الطالب ضمن غير المنزلين حتى يتم تنزيله يدوياً على كلاس.'
            )
        except Exception as e:
            messages.warning(self.request, f'تم التسجيل ولكن حدث خطأ في القيد المحاسبي: {str(e)}')
        return response
    
    def get_success_url(self):
        return reverse_lazy('quick:student_detail', kwargs={'pk': self.object.student.pk})

# بروفايل الطالب السريع
class QuickStudentProfileView(LoginRequiredMixin, DetailView):
    model = QuickStudent
    template_name = 'quick/quick_student_profile.html'
    context_object_name = 'student'
    
    def get_object(self):
        return get_object_or_404(QuickStudent, id=self.kwargs.get('student_id'))
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_object()
        
        try:
            # ✅ جلب التسجيلات النشطة فقط
            active_enrollments_queryset = QuickEnrollment.objects.filter(
                student=student, 
                is_completed=False
            ).select_related('course')
            
            # ✅ إنشاء قائمة بالبيانات المحسوبة للتسجيلات النشطة
            enrollment_data = []
            for enrollment in active_enrollments_queryset:
                # اربط الدفعات بهذا التسجيل نفسه لمنع خلط إيصالات تسجيل آخر
                total_paid = _get_quick_enrollment_paid_total(enrollment, student)
                
                net_amount = enrollment.net_amount or Decimal('0.00')
                balance_due = max(Decimal('0.00'), net_amount - total_paid)
                
                enrollment_data.append({
                    'enrollment': enrollment,
                    'total_paid': total_paid,
                    'balance_due': balance_due,
                    'net_amount': net_amount,
                    'is_active': not enrollment.is_completed
                })
            
            # ✅ حساب الإجماليات
            total_paid = sum(item['total_paid'] for item in enrollment_data)
            total_due = sum(item['net_amount'] for item in enrollment_data)
            total_remaining = total_due - total_paid
            
            # ✅ جلب جميع الإيصالات السريعة
            receipts = QuickStudentReceipt.objects.filter(
                quick_student=student
            ).select_related('course').order_by('-date', '-id')
            
            # ✅ التحقق من وجود تسجيلات نشطة
            has_active_enrollments = len(enrollment_data) > 0
            
            context.update({
                'enrollment_data': enrollment_data,
                'active_enrollments': enrollment_data,
                'all_enrollments': enrollment_data,
                'total_paid': total_paid,
                'total_due': total_due,
                'total_remaining': total_remaining,
                'receipts': receipts,
                'has_active_enrollments': has_active_enrollments,
                'delete_summary': _get_quick_student_delete_summary(student),
            })
            
        except Exception as e:
            messages.error(self.request, f'حدث خطأ في تحميل البيانات: {str(e)}')
            context.update({
                'enrollment_data': [],
                'active_enrollments': [],
                'all_enrollments': [],
                'total_paid': Decimal('0.00'),
                'total_due': Decimal('0.00'),
                'total_remaining': Decimal('0.00'),
                'receipts': [],
                'has_active_enrollments': False,
                'delete_summary': _get_quick_student_delete_summary(student),
            })
        
        return context
# كشف حساب الطالب السريع
class QuickStudentStatementView(LoginRequiredMixin, DetailView):
    model = QuickStudent
    template_name = 'quick/quick_student_statement.html'
    context_object_name = 'student'
    
    def get_object(self):
        return get_object_or_404(QuickStudent, id=self.kwargs.get('student_id'))
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_object()
        
        try:
            enrollments = list(
                QuickEnrollment.objects.filter(student=student)
                .select_related('course')
                .order_by('enrollment_date', 'id')
            )
            receipts = list(
                QuickStudentReceipt.objects.filter(quick_student=student)
                .select_related('course', 'quick_enrollment', 'journal_entry', 'created_by')
                .order_by('date', 'id')
            )

            enrollment_data = []
            per_course = []
            total_paid = Decimal('0.00')
            total_due = Decimal('0.00')

            for enrollment in enrollments:
                total_enrollment_paid = _get_quick_enrollment_paid_total(enrollment, student)
                net_amount = enrollment.net_amount or Decimal('0.00')
                balance_due = max(Decimal('0.00'), net_amount - total_enrollment_paid)

                enrollment_data.append({
                    'enrollment': enrollment,
                    'total_paid': total_enrollment_paid,
                    'balance_due': balance_due,
                    'net_amount': net_amount,
                    'is_active': not enrollment.is_completed,
                })
                per_course.append({
                    'course': enrollment.course,
                    'price': net_amount,
                    'paid': total_enrollment_paid,
                    'outstanding': balance_due,
                    'is_completed': enrollment.is_completed,
                    'enrollment_id': enrollment.id,
                })

                total_paid += total_enrollment_paid
                total_due += net_amount

            journal_entries = list(
                _get_quick_student_related_journal_entries(student)
                .select_related('created_by', 'posted_by')
                .prefetch_related('transactions__account')
                .order_by('date', 'id')
            )

            student_ar_account = getattr(student, 'ar_account', None)
            running_balance = Decimal('0.00')
            rows = []
            for entry in journal_entries:
                transactions = list(entry.transactions.select_related('account').all())
                transactions.sort(key=lambda tx: (not tx.is_debit, tx.id))
                for tx in transactions:
                    debit = tx.amount if tx.is_debit else Decimal('0.00')
                    credit = tx.amount if not tx.is_debit else Decimal('0.00')
                    if student_ar_account and tx.account_id == student_ar_account.id:
                        running_balance += debit - credit

                    rows.append({
                        'date': entry.date,
                        'ref': entry.reference or '-',
                        'desc': entry.description or tx.description or '-',
                        'account_code': tx.account.code,
                        'account_name': tx.account.name_ar or tx.account.name,
                        'tx_desc': tx.description or '-',
                        'debit': debit,
                        'credit': credit,
                        'balance': running_balance,
                        'created_by': (
                            entry.created_by.get_full_name()
                            or entry.created_by.username
                            if entry.created_by else '-'
                        ),
                        'entry_type': entry.get_entry_type_display(),
                    })

            total_remaining = max(Decimal('0.00'), total_due - total_paid)
            balance = student.balance

            context.update({
                'enrollment_data': enrollment_data,
                'active_enrollments': [row for row in enrollment_data if row['is_active']],
                'all_enrollments': enrollment_data,
                'total_paid': total_paid,
                'total_due': total_due,
                'total_remaining': total_remaining,
                'receipts': receipts,
                'has_active_enrollments': any(row['is_active'] for row in enrollment_data),
                'rows': rows,
                'per_course': per_course,
                'balance': balance,
                'entry_count': len(journal_entries),
            })
            
        except Exception as e:
            messages.error(self.request, f'حدث خطأ في تحميل البيانات: {str(e)}')
            context.update({
                'enrollment_data': [],
                'active_enrollments': [],
                'total_paid': Decimal('0.00'),
                'total_due': Decimal('0.00'),
                'total_remaining': Decimal('0.00'),
                'receipts': [],
                'has_active_enrollments': False,
                'rows': [],
                'per_course': [],
                'balance': Decimal('0.00'),
                'entry_count': 0,
            })
        
        return context

@require_POST
def update_quick_student_discount(request, student_id):
    """تحديث حسم الطالب السريع وتعديل القيود المرتبطة"""
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'يجب تسجيل الدخول'})
    
    student = get_object_or_404(QuickStudent, id=student_id)
    
    try:
        from decimal import Decimal
        from django.db import transaction as db_transaction
        
        discount_percent = Decimal(request.POST.get('discount_percent', '0'))
        discount_amount = Decimal(request.POST.get('discount_amount', '0'))
        discount_reason = request.POST.get('discount_reason', '')
        
        # التحقق من وجود تسجيلات نشطة
        active_enrollments = list(QuickEnrollment.objects.filter(
            student=student, 
            is_completed=False
        ))
        
        if not active_enrollments:
            return JsonResponse({
                'success': False,
                'error': 'لا توجد تسجيلات نشطة للطالب'
            })
        
        with db_transaction.atomic():
            updated_count = len(active_enrollments)
            for enrollment in active_enrollments:
                enrollment.discount_percent = discount_percent
                enrollment.discount_amount = discount_amount

            student.update_enrollment_discounts(request.user, enrollments=active_enrollments)
        
        return JsonResponse({
            'success': True,
            'message': f'تم تحديث الحسم والقيود المحاسبية لـ {updated_count} تسجيل نشط'
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"حدث خطأ في update_quick_student_discount: {str(e)}")
        
        return JsonResponse({
            'success': False,
            'error': f'حدث خطأ: {str(e)}'
        })

@require_POST
def quick_student_quick_receipt(request, student_id):
    """إنشاء إيصال فوري للطالب السريع"""
    from decimal import Decimal
    from django.db.models import Sum
    from .models import QuickStudentReceipt
    
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False, 'error': 'يجب تسجيل الدخول'}, status=401)
    
    student = get_object_or_404(QuickStudent, id=student_id)
    try:
        # Parse inputs
        course_id = request.POST.get('course_id')
        enrollment_id = request.POST.get('enrollment_id')
        amount = Decimal(request.POST.get('amount', '0'))
        paid_amount = Decimal(request.POST.get('paid_amount', '0'))
        discount_percent = Decimal(request.POST.get('discount_percent', '0'))
        discount_amount = Decimal(request.POST.get('discount_amount', '0'))
        receipt_date_str = request.POST.get('receipt_date')
        is_free = str(request.POST.get('is_free', '')).strip().lower() in {'1', 'true', 'yes', 'on'}
        
        # ✅ التصحيح: إذا كان amount صغيراً (أقل من 1000) نعتبره يحتاج أصفار
        if amount < 1000 and amount > 0:
            # نضرب في 1000 لإضافة الأصفار المفقودة
            amount = amount * 1000
        
        # معالجة تاريخ الإيصال
        if receipt_date_str:
            receipt_date = parse_date(receipt_date_str)
            if not receipt_date:
                return JsonResponse({'ok': False, 'error': 'صيغة التاريخ غير صحيحة'}, status=400)
        else:
            receipt_date = timezone.now().date()
            
    except (ValueError, TypeError, InvalidOperation) as e:
        return JsonResponse({'ok': False, 'error': f'خطأ في تنسيق الأرقام: {str(e)}'}, status=400)
    
    course = None
    remaining_amount = Decimal('0.00')
    enrollment = None
    
    try:
        if enrollment_id:
            try:
                enrollment_id = int(''.join(ch for ch in str(enrollment_id) if ch.isdigit()))
            except Exception:
                return JsonResponse({'ok': False, 'error': 'رقم التسجيل غير صالح'}, status=400)
            enrollment = QuickEnrollment.objects.get(pk=enrollment_id, student=student)
            
            if enrollment.is_completed:
                return JsonResponse({'ok': False, 'error': 'لا يمكن قطع إيصال لدورة مسحوبة'}, status=400)
                
            course = enrollment.course

            if course_id and str(course.id) != str(course_id):
                return JsonResponse({'ok': False, 'error': 'الدورة المحددة لا تطابق تسجيل الطالب'}, status=400)
            
            posted_net_amount = max(Decimal('0.00'), amount)
            zero_value_receipt = (
                is_free
                or posted_net_amount <= Decimal('0.00')
                or discount_percent >= Decimal('100')
            )
            if amount == 0 and not zero_value_receipt:
                amount = enrollment.net_amount or Decimal('0.00')
            elif zero_value_receipt:
                amount = posted_net_amount

            if (
                discount_percent != (enrollment.discount_percent or Decimal('0'))
                or discount_amount != (enrollment.discount_amount or Decimal('0'))
            ):
                enrollment.discount_percent = discount_percent
                enrollment.discount_amount = discount_amount
                enrollment.save()
                student.update_enrollment_discounts(request.user, enrollments=[enrollment])
            
            # احسب المتبقي من نفس التسجيل فقط
            total_paid = _get_quick_enrollment_paid_total(enrollment, student)
            
            net_amount = amount if amount > Decimal('0.00') or zero_value_receipt else (enrollment.net_amount or Decimal('0.00'))
            remaining_amount = max(Decimal('0.00'), net_amount - total_paid)
            
        elif course_id:
            course = QuickCourse.objects.get(pk=course_id)
            
            posted_net_amount = max(Decimal('0.00'), amount)
            zero_value_receipt = (
                is_free
                or posted_net_amount <= Decimal('0.00')
                or discount_percent >= Decimal('100')
            )
            if amount == 0 and not zero_value_receipt:
                amount = course.price or Decimal('0.00')
            elif zero_value_receipt:
                amount = posted_net_amount
                
            # البحث عن enrollment لهذه الدورة
            enrollment = QuickEnrollment.objects.filter(
                student=student, 
                course=course,
                is_completed=False
            ).first()
            
            if enrollment:
                total_paid = _get_quick_enrollment_paid_total(enrollment, student)
                if (
                    discount_percent != (enrollment.discount_percent or Decimal('0'))
                    or discount_amount != (enrollment.discount_amount or Decimal('0'))
                ):
                    enrollment.discount_percent = discount_percent
                    enrollment.discount_amount = discount_amount
                    enrollment.save()
                    student.update_enrollment_discounts(request.user, enrollments=[enrollment])

                net_amount = amount if amount > Decimal('0.00') or zero_value_receipt else (enrollment.net_amount or Decimal('0.00'))
                remaining_amount = max(Decimal('0.00'), net_amount - total_paid)
            else:
                remaining_amount = amount if amount > Decimal('0.00') or zero_value_receipt else (course.price or Decimal('0.00'))
                
    except (QuickEnrollment.DoesNotExist, QuickCourse.DoesNotExist) as e:
        return JsonResponse({'ok': False, 'error': 'الدورة أو التسجيل غير موجود'}, status=404)
    
    if paid_amount < 0:
        return JsonResponse({'ok': False, 'error': 'المبلغ المدفوع غير صالح'}, status=400)
    
    if paid_amount > remaining_amount:
        return JsonResponse({'ok': False, 'error': f'المبلغ المدفوع ({paid_amount}) يتجاوز المبلغ المتبقي ({remaining_amount})'}, status=400)
    
    # Create receipt - استخدام QuickStudentReceipt الجديد
    try:
        receipt = QuickStudentReceipt.objects.create(
            date=receipt_date,
            quick_student=student,
            student_name=student.full_name,
            course=course,
            course_name=(course.name if course else ''),
            quick_enrollment=enrollment,
            amount=amount,
            paid_amount=paid_amount,
            discount_percent=discount_percent,
            discount_amount=discount_amount,
            payment_method='CASH',
            created_by=request.user,
        )
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'فشل في إنشاء الإيصال: {str(e)}'}, status=500)
    
    journal_warning = None
    try:
        # إنشاء القيد المحاسبي
        if (receipt.paid_amount or Decimal('0')) > 0:
            receipt.create_accrual_journal_entry(request.user)
    except Exception as e:
        journal_warning = f"خطأ في القيد المحاسبي: {e}"
    
    new_remaining_amount = max(Decimal('0.00'), remaining_amount - paid_amount)
    
    from django.urls import reverse
    print_url = reverse('quick:quick_student_receipt_print', args=[receipt.id])
    return JsonResponse({
        'ok': True, 
        'receipt_id': receipt.id, 
        'print_url': print_url,
        'remaining_amount': float(new_remaining_amount),
        'warning': journal_warning
    })

@require_POST
def withdraw_quick_student(request, student_id):
    """سحب الطالب السريع من الدورة"""
    student = get_object_or_404(QuickStudent, pk=student_id)
    
    if request.method == 'POST':
        enrollment_id = request.POST.get('enrollment_id')
        withdrawal_reason = request.POST.get('withdrawal_reason', '')
        refund_amount_raw = request.POST.get('refund_amount', '0')

        if not enrollment_id:
            messages.error(request, 'لم يتم تحديد تسجيل الدورة')
            return redirect('quick:student_profile', student_id=student.id)

        try:
            enrollment = get_object_or_404(QuickEnrollment, pk=enrollment_id, student=student)

            if enrollment.is_completed:
                messages.error(request, 'هذه الدورة مسحوبة مسبقاً')
                return redirect('quick:student_profile', student_id=student.id)

            paid_total = QuickStudentReceipt.objects.filter(
                quick_student=student,
                quick_enrollment=enrollment,
                course=enrollment.course
            ).aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')

            try:
                refund_amount = Decimal(refund_amount_raw or '0')
            except InvalidOperation:
                refund_amount = Decimal('0')

            if refund_amount <= 0 and paid_total > 0:
                refund_amount = paid_total

            refund_result = _adjust_quick_receipts_for_refund(student, enrollment, refund_amount)
            actual_refund = refund_result['refunded_amount']
            refund_note = f' واسترد {actual_refund:,.0f} ل.س' if actual_refund > 0 else ''

            if getattr(enrollment, 'enrollment_journal_entry_id', None):
                try:
                    enrollment.enrollment_journal_entry.reverse_entry(
                        request.user,
                        description=f"إلغاء تسجيل سريع - {withdrawal_reason}" if withdrawal_reason else "إلغاء تسجيل سريع"
                    )
                except Exception:
                    pass

            description = f"سحب طالب سريع {student.full_name} من {enrollment.course.name}"
            if withdrawal_reason:
                description = f"{description} - {withdrawal_reason}"
            _build_quick_withdrawal_entry(
                enrollment=enrollment,
                user=request.user,
                refunded_amount=actual_refund,
                description=description,
            )

            enrollment.is_completed = True
            enrollment.completion_date = timezone.now().date()
            enrollment.save(update_fields=['is_completed', 'completion_date'])

            messages.success(request, f'تم سحب الطالب من دورة {enrollment.course.name}{refund_note} بنجاح')
            return redirect('quick:student_profile', student_id=student.id)

        except Exception as e:
            print(f"ERROR in withdraw_quick_student: {str(e)}")
            messages.error(request, f'حدث خطأ أثناء السحب: {str(e)}')
            return redirect('quick:student_profile', student_id=student.id)

@require_POST
def refund_quick_student(request, student_id):
    """استرداد مبلغ للطالب السريع"""
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False, 'error': 'يجب تسجيل الدخول'}, status=401)
    
    student = get_object_or_404(QuickStudent, pk=student_id)
    
    try:
        enrollment_id = request.POST.get('enrollment_id')
        refund_amount = Decimal(request.POST.get('refund_amount', '0'))
        refund_reason = request.POST.get('refund_reason', '')
        
        if not enrollment_id:
            return JsonResponse({'ok': False, 'error': 'لم يتم تحديد التسجيل'}, status=400)
        
        enrollment = get_object_or_404(QuickEnrollment, pk=enrollment_id, student=student)
        
        if enrollment.is_completed:
            return JsonResponse({'ok': False, 'error': 'لا يمكن استرداد مبلغ لدورة مسحوبة'}, status=400)
        
        try:
            result = _process_quick_refund(
                student,
                enrollment,
                refund_amount,
                refund_reason,
                request.user
            )
        except ValueError as exc:
            return JsonResponse({'ok': False, 'error': str(exc)}, status=400)
        except Exception as exc:
            import traceback
            print(f"خطأ في الاسترداد: {str(exc)}")
            print(traceback.format_exc())
            return JsonResponse({'ok': False, 'error': f'خطأ في الاسترداد: {str(exc)}'}, status=500)

        return JsonResponse({
            'ok': True,
            'message': f'تم استرداد {result["refund_amount"]:,.0f} ل.س بنجاح',
            'new_balance': float(result['new_balance']),
            'previous_balance': float(result['previous_balance']),
            'new_paid': float(result['new_total_paid']),
            'previous_paid': float(result['previous_paid'])
        })

    except Exception as e:
        import traceback
        print(f"خطأ في الاسترداد: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'ok': False, 'error': f'حدث خطأ في الاسترداد: {str(e)}'}, status=500)
# التقارير
class QuickOutstandingCoursesView(LoginRequiredMixin, ListView):
    template_name = 'quick/outstanding_course_list.html'
    context_object_name = 'courses'
    
    def get_queryset(self):
        course_type, course_type_label, report_label = _get_outstanding_course_type(self.request)
        self._course_type = course_type
        self._course_type_label = course_type_label
        self._course_type_report_label = report_label
        courses = QuickCourse.objects.filter(is_active=True).select_related('academic_year').order_by('name')
        if course_type != 'ALL':
            courses = courses.filter(course_type=course_type)
        course_data, totals = _build_quick_outstanding_course_summary(courses, include_zero_outstanding=True)
        self._totals = totals
        return course_data

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        totals = getattr(self, '_totals', {})
        context.update({
            'total_courses': totals.get('total_courses', 0),
            'total_outstanding_students': totals.get('total_outstanding_students', 0),
            'total_outstanding_amount': totals.get('total_outstanding_amount', Decimal('0')),
            'total_paid_amount': totals.get('total_paid_amount', Decimal('0')),
            'course_type': getattr(self, '_course_type', 'INTENSIVE'),
            'course_type_label': getattr(self, '_course_type_label', 'مكثفة'),
            'course_type_report_label': getattr(self, '_course_type_report_label', 'المكثفات'),
            'course_type_options': _get_course_type_options(),
        })
        total_courses = totals.get('total_courses', 0) or 0
        total_paid_amount = totals.get('total_paid_amount', Decimal('0'))
        context['average_paid_per_course'] = (
            (total_paid_amount / total_courses) if total_courses else Decimal('0')
        )
        return context


class QuickOutstandingCourseDetailView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/outstanding_course_detail.html'

    def get_context_data(self, course_id=None, **kwargs):
        context = super().get_context_data(**kwargs)
        course_type, course_type_label, report_label = _get_outstanding_course_type(self.request)
        course = get_object_or_404(QuickCourse, pk=course_id)

        enrollments = QuickEnrollment.objects.filter(
            course=course,
            is_completed=False
        ).select_related('student')

        rows = []
        total_outstanding = Decimal('0.00')

        for enrollment in enrollments:
            net_amount = enrollment.net_amount or Decimal('0.00')
            total_paid = QuickStudentReceipt.objects.filter(
                quick_enrollment=enrollment
            ).aggregate(total=Sum('paid_amount'))['total'] or Decimal('0.00')

            remaining = max(Decimal('0.00'), net_amount - total_paid)

            if remaining <= Decimal('0.00'):
                continue

            rows.append({
                'student_id': enrollment.student.id,
                'student_name': enrollment.student.full_name,
                'phone': enrollment.student.phone,
                'net_amount': net_amount,
                'paid_amount': total_paid,
                'remaining': remaining,
            })

            total_outstanding += remaining

        rows.sort(key=lambda r: r['remaining'], reverse=True)

        context.update({
            'course': course,
            'rows': rows,
            'total_students': len(rows),
            'total_outstanding': total_outstanding,
            'total_net_amount': sum(r['net_amount'] for r in rows),
            'total_paid_amount': sum(r['paid_amount'] for r in rows),
            'course_type': course_type,
            'course_type_label': course_type_label,
            'course_type_report_label': report_label,
        })
        return context


class QuickCourseStudentsView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/outstanding_course_students.html'

    def get_context_data(self, course_id=None, **kwargs):
        context = super().get_context_data(**kwargs)
        course_type, course_type_label, report_label = _get_outstanding_course_type(self.request)
        course = get_object_or_404(QuickCourse, pk=course_id)

        enrolled_students = self.get_students_from_enrollments(course)
        male_count = 0
        female_count = 0
        unknown_count = 0
        for item in enrolled_students:
            quick_student = item.get('student')
            gender = getattr(getattr(quick_student, 'student', None), 'gender', None)
            if gender == 'male':
                male_count += 1
            elif gender == 'female':
                female_count += 1
            else:
                unknown_count += 1

        if not enrolled_students:
            context.update({
                'course': course,
                'student_data': [],
                'total_students': 0,
                'fully_paid_count': 0,
                'outstanding_count': 0,
                'current_filter': 'all',
                'total_net_due': 0,
                'total_paid': 0,
                'total_remaining': 0,
                'all_students_count': 0,
                'students_without_receipts': 0,
                'male_count': 0,
                'female_count': 0,
                'unknown_count': 0,
                'course_type': course_type,
                'course_type_label': course_type_label,
                'course_type_report_label': report_label,
            })
            return context

        student_data, statistics = self.calculate_student_data(course, enrolled_students)
        filter_type = self.request.GET.get('filter', 'all')
        filtered_students, filtered_statistics = self.apply_filter(student_data, filter_type)

        context.update({
            'course': course,
            'student_data': filtered_students,
            'total_students': len(filtered_students),
            'fully_paid_count': statistics['fully_paid_count'],
            'outstanding_count': statistics['outstanding_count'],
            'current_filter': filter_type,
            'total_net_due': filtered_statistics['total_net_due'],
            'total_paid': filtered_statistics['total_paid'],
            'total_remaining': filtered_statistics['total_remaining'],
            'all_students_count': statistics['all_students_count'],
            'students_without_receipts': statistics['students_without_receipts'],
            'male_count': male_count,
            'female_count': female_count,
            'unknown_count': unknown_count,
            'course_type': course_type,
            'course_type_label': course_type_label,
            'course_type_report_label': report_label,
        })
        return context

    def get_students_from_enrollments(self, course):
        enrollments = QuickEnrollment.objects.filter(
            course=course,
            is_completed=False
        ).select_related('student')

        return [
            {'student': enrollment.student, 'enrollment': enrollment}
            for enrollment in enrollments
            if enrollment.student
        ]

    def calculate_student_data(self, course, students_with_enrollments):
        student_data = []
        statistics = {
            'total_net_due': Decimal('0'),
            'total_paid': Decimal('0'),
            'total_remaining': Decimal('0'),
            'students_without_receipts': 0,
            'fully_paid_count': 0,
            'outstanding_count': 0,
            'all_students_count': len(students_with_enrollments)
        }

        for item in students_with_enrollments:
            info = self.calculate_student_financial_info(item['student'], item['enrollment'])
            if not info:
                continue

            student_data.append(info)
            statistics['total_net_due'] += info['net_due']
            statistics['total_paid'] += info['paid_total']
            statistics['total_remaining'] += info['remaining']

            if not info['has_receipts']:
                statistics['students_without_receipts'] += 1

            if info['is_fully_paid']:
                statistics['fully_paid_count'] += 1
            else:
                statistics['outstanding_count'] += 1

        return student_data, statistics

    def calculate_student_financial_info(self, student, enrollment):
        try:
            course_price = enrollment.course.price or Decimal('0')
            net_due = enrollment.net_amount or Decimal('0')
            paid_total = self.calculate_paid_amount(enrollment)
            remaining = max(Decimal('0'), net_due - paid_total)
            discount_percent = enrollment.discount_percent or Decimal('0')
            is_fully_paid = (
                discount_percent >= Decimal('100')
                or net_due <= Decimal('0')
                or remaining <= Decimal('0')
            )
            has_receipts = paid_total > Decimal('0')
            is_free = net_due <= Decimal('0') or discount_percent >= Decimal('100')

            return {
                'student': student,
                'enrollment': enrollment,
                'course_price': course_price,
                'net_due': net_due,
                'paid_total': paid_total,
                'remaining': remaining,
                'is_fully_paid': is_fully_paid,
                'has_receipts': has_receipts,
                'is_free': is_free,
            }
        except Exception:
            return None

    def calculate_paid_amount(self, enrollment):
        total_paid = QuickStudentReceipt.objects.filter(
            quick_enrollment=enrollment
        ).aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
        return total_paid

    def apply_filter(self, student_data, filter_type):
        if filter_type == 'paid':
            filtered = [s for s in student_data if s['is_fully_paid']]
        elif filter_type == 'outstanding':
            filtered = [s for s in student_data if not s['is_fully_paid']]
        else:
            filtered = student_data

        filtered.sort(key=lambda x: (
            not x['is_fully_paid'],
            -x['remaining']
        ))

        filtered_statistics = {
            'total_net_due': sum(s['net_due'] for s in filtered),
            'total_paid': sum(s['paid_total'] for s in filtered),
            'total_remaining': sum(s['remaining'] for s in filtered)
        }

        return filtered, filtered_statistics


@login_required
def register_quick_course(request, student_id):
    """تسجيل طالب سريع في دورة"""
    student = get_object_or_404(QuickStudent, id=student_id)
    context = _build_quick_register_course_context(student)
    courses = context['courses']
    
    if request.method == 'POST':
        course_ids = request.POST.getlist('course_ids')
        if not course_ids:
            messages.error(request, 'يرجى اختيار دورة واحدة على الأقل')
            return redirect('quick:register_quick_course', student_id=student_id)

        seen = []
        for cid in course_ids:
            if cid and cid not in seen:
                seen.append(cid)

        available_courses = QuickCourse.objects.filter(
            id__in=seen,
            is_active=True,
            academic_year=student.academic_year
        ).prefetch_related(
            Prefetch(
                'sessions',
                queryset=QuickCourseSession.objects.filter(is_active=True).order_by('start_date', 'start_time', 'id'),
                to_attr='active_sessions',
            )
        )
        available_map = {str(course.id): course for course in available_courses}

        created_enrollments = 0
        assigned_sessions = 0
        created_receipts = []
        created_enrollment_ids = []
        warnings = []

        for cid in seen:
            course = available_map.get(cid)
            if not course:
                continue

            existing = QuickEnrollment.objects.filter(student=student, course=course).exists()
            if existing:
                warnings.append(f'التسجيل للدورة "{course.name}" موجود مسبقاً، تم تجاهلها.')
                continue

            selected_session_id = (request.POST.get(f'session_id_{course.id}') or '').strip()
            selected_session = None
            if selected_session_id:
                selected_session = next(
                    (
                        session for session in getattr(course, 'active_sessions', [])
                        if str(session.id) == selected_session_id
                    ),
                    None,
                )
                if selected_session is None:
                    warnings.append(f'الكلاس المختار لدورة "{course.name}" غير صالح وتم تجاهله.')

            enrollment = QuickEnrollment.objects.create(
                student=student,
                course=course,
                enrollment_date=timezone.now().date(),
                net_amount=course.price,
                total_amount=course.price
            )
            created_enrollments += 1
            created_enrollment_ids.append(enrollment.id)

            if selected_session is not None:
                try:
                    _assign_enrollment_to_specific_session(
                        enrollment=enrollment,
                        session=selected_session,
                        user=request.user,
                    )
                    assigned_sessions += 1
                except ValueError as exc:
                    warnings.append(
                        f'الطالب سُجل في دورة {course.name} لكن لم يتم تنزيله على الكلاس المختار: {exc}'
                    )
            elif getattr(course, 'active_sessions', []):
                warnings.append(
                    f'الطالب سُجل في دورة {course.name} لكن لم يتم اختيار كلاس لها، لذلك بقي ضمن غير المنزلين.'
                )
            else:
                warnings.append(
                    f'الطالب سُجل في دورة {course.name} ولا توجد كلاسات مفعلة لها حالياً، لذلك بقي ضمن غير المنزلين.'
                )

            try:
                enrollment.create_accrual_enrollment_entry(request.user)
            except Exception as exc:
                warnings.append(f'القيد المحاسبي لدورة {course.name} لم يُنجز: {exc}')

            pay_full = request.POST.get(f'pay_full_{course.id}')
            if pay_full:
                try:
                    receipt = QuickStudentReceipt.objects.create(
                        date=timezone.now().date(),
                        quick_student=student,
                        student_name=student.full_name,
                        course=course,
                        course_name=course.name,
                        quick_enrollment=enrollment,
                        amount=enrollment.net_amount,
                        paid_amount=enrollment.net_amount,
                        payment_method='CASH',
                        created_by=request.user
                    )
                    receipt.create_accrual_journal_entry(request.user)
                    created_receipts.append(receipt.id)
                except Exception as exc:
                    warnings.append(f'إنشاء إيصال لدورة {course.name} فشل: {exc}')

        if created_enrollments:
            summary_message = f'تم تسجيل الطالب في {created_enrollments} دورة'
            if assigned_sessions:
                summary_message += f' وتنزيله مباشرة على {assigned_sessions} كلاس'
            messages.success(request, summary_message)
        if warnings:
            for warning in warnings:
                messages.warning(request, warning)

        if created_enrollment_ids:
            query_params = {
                'enrollments': ','.join(str(enrollment_id) for enrollment_id in created_enrollment_ids),
            }
            if created_receipts:
                query_params['receipts'] = ','.join(str(rid) for rid in created_receipts)
            query = urlencode(query_params)
            return redirect(f"{reverse('quick:quick_student_times_print', args=[student_id])}?{query}")

        return redirect('quick:student_profile', student_id=student_id)

    return render(request, 'quick/register_quick_course.html', context)
@login_required
def quick_multiple_receipt_print(request, student_id):
    """طباعة مجموعة إيصالات دفعة واحدة"""
    ids_param = request.GET.get('ids', '')
    if not ids_param:
        raise Http404('Missing receipt identifiers')

    try:
        receipt_ids = [int(pk.strip()) for pk in ids_param.split(',') if pk.strip()]
    except ValueError:
        raise Http404('Invalid receipt identifiers')

    receipts = QuickStudentReceipt.objects.filter(
        id__in=receipt_ids,
        quick_student_id=student_id
    ).order_by('id')

    if not receipts.exists():
        raise Http404('No receipts found')

    receipt_items = []
    for receipt in receipts:
        course_price = (receipt.amount + receipt.discount_amount) if receipt.discount_amount else receipt.amount
        net_due = receipt.amount
        remaining = max(Decimal('0.00'), course_price - (receipt.paid_amount or Decimal('0.00')))
        receipt_items.append({
            'receipt': receipt,
            'course_price': course_price,
            'discount_percent': receipt.discount_percent,
            'net_due': net_due,
            'remaining': remaining,
            'receipt_date': receipt.date,
        })

    student = QuickStudent.objects.filter(id=student_id).first()
    return render(request, 'quick/quick_multiple_receipt_print.html', {
        'receipts': receipt_items,
        'student': student,
        'return_url': reverse('quick:student_profile', args=[student_id]),
        'local_agent_url': settings.QUICK_LOCAL_AGENT_URL,
        'server_printer_enabled': (
            settings.QUICK_RECEIPT_PRINTER_ENABLED or settings.QUICK_RECEIPT_PRINTER_DUMMY
        ),
    })


@login_required
def quick_student_times_print(request, student_id):
    ids_param = request.GET.get('enrollments', '')
    receipt_ids_param = request.GET.get('receipts', '')
    if not ids_param:
        raise Http404('Missing enrollment identifiers')

    try:
        enrollment_ids = [int(pk.strip()) for pk in ids_param.split(',') if pk.strip()]
    except ValueError:
        raise Http404('Invalid enrollment identifiers')

    student = get_object_or_404(QuickStudent, id=student_id)
    enrollments = list(
        QuickEnrollment.objects.filter(id__in=enrollment_ids, student_id=student_id)
        .select_related('course', 'session_assignment__session')
        .order_by('id')
    )
    if not enrollments:
        raise Http404('No enrollments found')

    schedule_items = []
    for enrollment in enrollments:
        assignment = getattr(enrollment, 'session_assignment', None)
        session = getattr(assignment, 'session', None)
        schedule_items.append({
            'course_name': enrollment.course.name if enrollment.course else '-',
            'time_label': _format_quick_session_time_label(session),
        })

    return render(request, 'quick/quick_student_times_print.html', {
        'student': student,
        'schedule_items': schedule_items,
        'return_url': reverse('quick:student_profile', args=[student_id]),
        'receipts_print_url': (
            f"{reverse('quick:quick_multiple_receipt_print', args=[student_id])}?ids={receipt_ids_param}"
            if receipt_ids_param else ''
        ),
    })


def _build_quick_receipt_payload(receipts, student_id):
    items = []
    for receipt in receipts:
        course_name = receipt.course.name if receipt.course else (receipt.course_name or '-')
        student_name = receipt.quick_student.full_name if receipt.quick_student else (receipt.student_name or '-')
        net_due = receipt.amount if receipt.amount is not None else (
            receipt.quick_enrollment.net_amount if receipt.quick_enrollment else Decimal('0')
        )
        paid_amount = receipt.paid_amount or Decimal('0')
        remaining = max(Decimal('0'), net_due - paid_amount)
        items.append({
            'id': receipt.id,
            'number': receipt.receipt_number or str(receipt.id),
            'date': receipt.date.strftime('%Y-%m-%d') if receipt.date else '',
            'student': student_name,
            'course': course_name,
            'net_due': str(net_due),
            'paid_amount': str(paid_amount),
            'remaining': str(remaining),
            'discount_percent': str(receipt.discount_percent or Decimal('0')),
            'payment_method': receipt.get_payment_method_display(),
            'notes': receipt.notes or '',
        })

    return {
        'student_id': student_id,
        'count': len(items),
        'title': settings.QUICK_RECEIPT_PRINTER_TITLE,
        'receipts': items,
    }


def _validate_print_agent(request):
    configured = settings.QUICK_PRINT_AGENT_TOKEN
    provided = request.headers.get('X-Printer-Token', '').strip()
    return bool(configured) and provided == configured


@login_required
@require_POST
def quick_multiple_receipt_payload(request, student_id):
    ids_param = request.POST.get('ids', '')
    if not ids_param:
        return JsonResponse({'ok': False, 'error': 'لم يتم تحديد الإيصالات'}, status=400)

    try:
        receipt_ids = [int(pk.strip()) for pk in ids_param.split(',') if pk.strip()]
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'معرّفات الإيصالات غير صحيحة'}, status=400)

    receipts = list(
        QuickStudentReceipt.objects.filter(
            id__in=receipt_ids,
            quick_student_id=student_id
        ).select_related('quick_student', 'course', 'quick_enrollment').order_by('id')
    )
    if not receipts:
        return JsonResponse({'ok': False, 'error': 'لا توجد إيصالات للطباعة'}, status=404)

    return JsonResponse({
        'ok': True,
        'payload': _build_quick_receipt_payload(receipts, student_id),
    })


@login_required
@require_POST
def quick_multiple_receipt_enqueue_print(request, student_id):
    ids_param = request.POST.get('ids', '')
    if not ids_param:
        return JsonResponse({'ok': False, 'error': 'لم يتم تحديد الإيصالات'}, status=400)

    try:
        receipt_ids = [int(pk.strip()) for pk in ids_param.split(',') if pk.strip()]
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'معرّفات الإيصالات غير صحيحة'}, status=400)

    student = get_object_or_404(QuickStudent, id=student_id)
    receipts = list(
        QuickStudentReceipt.objects.filter(
            id__in=receipt_ids,
            quick_student_id=student_id
        ).select_related('quick_student', 'course', 'quick_enrollment').order_by('id')
    )
    if not receipts:
        return JsonResponse({'ok': False, 'error': 'لا توجد إيصالات للطباعة'}, status=404)

    job = QuickReceiptPrintJob.objects.create(
        created_by=request.user,
        quick_student=student,
        payload=_build_quick_receipt_payload(receipts, student_id),
        status=QuickReceiptPrintJob.STATUS_PENDING,
    )
    return JsonResponse({
        'ok': True,
        'job_id': job.id,
        'message': f'تم إنشاء مهمة الطباعة رقم {job.id}. سيقوم لابتوب الطباعة بسحبها تلقائياً.',
    })


@csrf_exempt
@require_GET
def quick_print_agent_next_job(request):
    if not _validate_print_agent(request):
        return JsonResponse({'ok': False, 'error': 'Unauthorized printer agent'}, status=403)

    with transaction.atomic():
        job = (
            QuickReceiptPrintJob.objects
            .select_for_update(skip_locked=True)
            .filter(status=QuickReceiptPrintJob.STATUS_PENDING)
            .order_by('created_at')
            .first()
        )
        if not job:
            return JsonResponse({'ok': True, 'job': None})

        job.status = QuickReceiptPrintJob.STATUS_PROCESSING
        job.picked_at = timezone.now()
        job.error_message = ''
        job.save(update_fields=['status', 'picked_at', 'error_message', 'updated_at'])

    return JsonResponse({
        'ok': True,
        'job': {
            'id': job.id,
            'payload': job.payload,
        }
    })


@csrf_exempt
@require_POST
def quick_print_agent_job_update(request, job_id):
    if not _validate_print_agent(request):
        return JsonResponse({'ok': False, 'error': 'Unauthorized printer agent'}, status=403)

    job = get_object_or_404(QuickReceiptPrintJob, id=job_id)
    status = request.POST.get('status', '').strip().lower()
    error_message = request.POST.get('error_message', '').strip()

    if status not in {QuickReceiptPrintJob.STATUS_COMPLETED, QuickReceiptPrintJob.STATUS_FAILED}:
        return JsonResponse({'ok': False, 'error': 'Invalid status'}, status=400)

    job.status = status
    job.error_message = error_message
    job.completed_at = timezone.now()
    job.save(update_fields=['status', 'error_message', 'completed_at', 'updated_at'])
    return JsonResponse({'ok': True})


@login_required
@require_POST
def quick_multiple_receipt_server_print(request, student_id):
    ids_param = request.POST.get('ids', '')
    if not ids_param:
        return JsonResponse({'ok': False, 'error': 'لم يتم تحديد الإيصالات'}, status=400)

    try:
        receipt_ids = [int(pk.strip()) for pk in ids_param.split(',') if pk.strip()]
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'معرّفات الإيصالات غير صحيحة'}, status=400)

    receipts = list(
        QuickStudentReceipt.objects.filter(
            id__in=receipt_ids,
            quick_student_id=student_id
        ).select_related('quick_student', 'course', 'quick_enrollment').order_by('id')
    )
    if not receipts:
        return JsonResponse({'ok': False, 'error': 'لا توجد إيصالات للطباعة'}, status=404)

    try:
        dummy_output = print_many_receipts(receipts)
    except QuickReceiptPrinterError as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=400)

    response = {
        'ok': True,
        'printed_count': len(receipts),
        'message': f'تم إرسال {len(receipts)} إيصال إلى طابعة السيرفر',
    }
    if settings.QUICK_RECEIPT_PRINTER_DUMMY and dummy_output:
        response['dummy_preview'] = dummy_output.decode('utf-8', errors='ignore')[:4000]
    return JsonResponse(response)

def quick_student_receipt_print(request, receipt_id):
    """طباعة إيصال الطالب السريع"""
    receipt = get_object_or_404(
        QuickStudentReceipt.objects.select_related('quick_student', 'course', 'quick_enrollment'),
        id=receipt_id
    )

    enrollment = receipt.quick_enrollment
    if enrollment:
        net_due = receipt.amount if receipt.amount is not None else (enrollment.net_amount or Decimal('0.00'))
        total_paid = _get_quick_enrollment_paid_total(enrollment, receipt.quick_student)
        remaining = max(Decimal('0.00'), net_due - total_paid)
    else:
        net_due = receipt.amount or Decimal('0.00')
        remaining = max(Decimal('0.00'), net_due - (receipt.paid_amount or Decimal('0.00')))

    course_price = (
        (net_due + (receipt.discount_amount or Decimal('0.00')))
        if receipt.discount_amount else net_due
    )
    context = {
        'receipt': receipt,
        'remaining': remaining,
        'course_price': course_price,
        'discount_percent': receipt.discount_percent,
        'net_due': net_due,
        'receipt_date': receipt.date,
    }
    
    return render(request, 'quick/quick_student_receipt_print.html', context)


# في quick/views.py - أضف هذه الدالة في النهاية

@login_required
def auto_assign_academic_years(request):
    """ربط جميع الطلاب بفصولهم الدراسية تلقائياً"""
    from students.models import Student
    from quick.models import QuickStudent, AcademicYear
    
    # ربط الطلاب السريعين
    quick_students = QuickStudent.objects.filter(academic_year__isnull=True)
    updated_count = 0
    
    for student in quick_students:
        academic_year = AcademicYear.objects.filter(
            start_date__lte=student.created_at.date(),
            end_date__gte=student.created_at.date(),
            is_active=True
        ).first()
        
        if academic_year:
            student.academic_year = academic_year
            student.save()
            updated_count += 1
    
    messages.success(request, f'تم ربط {updated_count} طالب سريع تلقائياً بالفصول الدراسية')
    return redirect('quick:student_list')


# في ملف views.py - تحديث دالة التعديل

class QuickStudentUpdateView(LoginRequiredMixin, UpdateView):
    model = QuickStudent
    form_class = QuickStudentForm
    template_name = 'quick/quick_student_update.html'
    context_object_name = 'student'
    
    def get_success_url(self):
        # ✅ التوجيه إلى بروفايل الطالب بدلاً من التفاصيل البسيطة
        return reverse_lazy('quick:student_profile', kwargs={'student_id': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'تم تحديث بيانات الطالب بنجاح')
        return super().form_valid(form)




        # أضف هذه الكلاس في قسم "الدورات السريعة" بعد QuickCourseCreateView

class QuickCourseUpdateView(LoginRequiredMixin, UpdateView):
    model = QuickCourse
    form_class = QuickCourseForm
    template_name = 'quick/quick_course_form.html'  # نفس قالب الإنشاء
    context_object_name = 'course'
    
    def get_success_url(self):
        return reverse('quick:course_detail', kwargs={'pk': self.object.pk})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_update'] = True  # للتمييز بين التعديل والإضافة
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'تم تحديث بيانات الدورة بنجاح')
        return super().form_valid(form)

@require_POST

@require_POST
def withdraw_quick_student(request, student_id):
    """Withdraw quick student from course."""
    student = get_object_or_404(QuickStudent, pk=student_id)
    enrollment_id_raw = request.POST.get('enrollment_id')
    withdrawal_reason = request.POST.get('withdrawal_reason', '')
    refund_amount_raw = request.POST.get('refund_amount', '0')
    wants_json = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    enrollment_id = ''.join(ch for ch in str(enrollment_id_raw or '') if ch.isdigit())

    if not enrollment_id:
        error_message = 'لم يتم تحديد تسجيل الدورة'
        if wants_json:
            return JsonResponse({'success': False, 'error': error_message}, status=400)
        messages.error(request, error_message)
        return redirect('quick:student_profile', student_id=student.id)

    try:
        enrollment = get_object_or_404(QuickEnrollment, pk=enrollment_id, student=student)
        try:
            refund_amount = Decimal(refund_amount_raw or '0')
        except InvalidOperation:
            refund_amount = Decimal('0')

        result = _withdraw_quick_enrollment(
            enrollment=enrollment,
            user=request.user,
            withdrawal_reason=withdrawal_reason,
            refund_amount=refund_amount,
        )
        actual_refund = result['actual_refund']
        if wants_json:
            return JsonResponse({
                'success': True,
                'student_name': result['student_name'],
                'course_name': result['course_name'],
                'actual_refund': f'{actual_refund:,.0f}',
                'created_entry_ids': result.get('created_entry_ids', []),
            })
        refund_note = f' and refunded {actual_refund:,.0f} SYP' if actual_refund > 0 else ''
        messages.success(request, f'Student withdrawn from course {enrollment.course.name}{refund_note} successfully')
    except Exception as exc:
        print(f"ERROR in withdraw_quick_student override: {exc}")
        if wants_json:
            return JsonResponse({'success': False, 'error': str(exc)}, status=400)
        messages.error(request, f'Withdrawal error: {exc}')

    return redirect('quick:student_profile', student_id=student.id)


@login_required
@require_POST
def bulk_withdraw_quick_students(request, course_id):
    course = get_object_or_404(QuickCourse, pk=course_id, is_active=True)
    enrollment_ids = request.POST.getlist('enrollment_ids')

    if not enrollment_ids:
        messages.error(request, 'لم يتم تحديد أي طالب للحذف الجماعي.')
        return redirect(reverse('quick:late_payment_course_detail', args=[course.id]))

    enrollments = list(
        QuickEnrollment.objects.filter(
            id__in=enrollment_ids,
            course=course,
            is_completed=False
        ).select_related('student', 'course')
    )

    if not enrollments:
        messages.error(request, 'لم يتم العثور على تسجيلات صالحة للحذف.')
        return redirect(reverse('quick:late_payment_course_detail', args=[course.id]))

    paid_rows = QuickStudentReceipt.objects.filter(
        quick_enrollment_id__in=[enrollment.id for enrollment in enrollments]
    ).values('quick_enrollment_id').annotate(total=Sum('paid_amount'))
    paid_map = {
        row['quick_enrollment_id']: (row['total'] or Decimal('0'))
        for row in paid_rows
    }

    deleted_count = 0
    errors = []
    for enrollment in enrollments:
        try:
            with transaction.atomic():
                _delete_quick_enrollment(
                    enrollment=enrollment,
                    paid_total=paid_map.get(enrollment.id, Decimal('0')),
                )
            deleted_count += 1
        except Exception as exc:
            errors.append(f'{enrollment.student.full_name}: {exc}')

    if deleted_count:
        messages.success(request, f'تم حذف {deleted_count} تسجيل من دورة {course.name}.')
    for error in errors[:5]:
        messages.error(request, error)

    query_string = urlencode({
        key: value for key, value in {
            'course_type': request.POST.get('course_type') or '',
            'start_date': request.POST.get('start_date') or '',
            'end_date': request.POST.get('end_date') or '',
        }.items() if value
    })
    redirect_url = reverse('quick:late_payment_course_detail', args=[course.id])
    if query_string:
        redirect_url = f'{redirect_url}?{query_string}'
    return redirect(redirect_url)


class QuickLatePaymentCoursesView(LoginRequiredMixin, ListView):
    template_name = 'quick/late_payment_course_list.html'
    context_object_name = 'courses'

    def get_queryset(self):
        course_type, course_type_label, report_label = _get_outstanding_course_type(self.request)
        start_date, end_date = _get_outstanding_date_range(self.request)
        self._course_type = course_type
        self._course_type_label = course_type_label
        self._course_type_report_label = report_label
        self._start_date = start_date
        self._end_date = end_date

        courses = QuickCourse.objects.filter(is_active=True).select_related('academic_year').order_by('name')
        if course_type != 'ALL':
            courses = courses.filter(course_type=course_type)

        course_data, totals = _build_quick_outstanding_course_summary(
            courses,
            include_zero_outstanding=False,
            start_date=start_date,
            end_date=end_date,
        )
        self._totals = totals
        return course_data

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        totals = getattr(self, '_totals', {})
        total_courses = totals.get('total_courses', 0) or 0
        total_paid_amount = totals.get('total_paid_amount', Decimal('0'))
        context.update({
            'total_courses': total_courses,
            'total_outstanding_students': totals.get('total_outstanding_students', 0),
            'total_outstanding_amount': totals.get('total_outstanding_amount', Decimal('0')),
            'total_paid_amount': total_paid_amount,
            'course_type': getattr(self, '_course_type', 'INTENSIVE'),
            'course_type_label': getattr(self, '_course_type_label', ''),
            'course_type_report_label': getattr(self, '_course_type_report_label', ''),
            'course_type_options': _get_course_type_options(),
            'start_date': getattr(self, '_start_date', None),
            'end_date': getattr(self, '_end_date', None),
            'average_paid_per_course': (total_paid_amount / total_courses) if total_courses else Decimal('0'),
        })
        return context


class QuickLatePaymentCourseDetailView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/late_payment_course_detail.html'

    def get_context_data(self, course_id=None, **kwargs):
        context = super().get_context_data(**kwargs)
        course_type, course_type_label, report_label = _get_outstanding_course_type(self.request)
        start_date, end_date = _get_outstanding_date_range(self.request)
        course = get_object_or_404(QuickCourse, pk=course_id)

        rows, detail_data = _build_quick_outstanding_rows([course], start_date=start_date, end_date=end_date)
        grouped_course = detail_data['grouped_courses'][0] if detail_data['grouped_courses'] else {
            'date_groups': [],
            'total_students': 0,
            'total_outstanding': Decimal('0'),
        }

        context.update({
            'course': course,
            'rows': rows,
            'date_groups': grouped_course['date_groups'],
            'total_students': grouped_course['total_students'],
            'total_outstanding': grouped_course['total_outstanding'],
            'total_net_amount': sum(r['net_amount'] for r in rows),
            'total_paid_amount': sum(r['paid_amount'] for r in rows),
            'course_type': course_type,
            'course_type_label': course_type_label,
            'course_type_report_label': report_label,
            'start_date': start_date,
            'end_date': end_date,
        })
        return context


class QuickLatePaymentCoursesPrintView(LoginRequiredMixin, TemplateView):
    template_name = 'quick/late_payment_course_print.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        course_type, course_type_label, report_label = _get_outstanding_course_type(self.request)
        start_date, end_date = _get_outstanding_date_range(self.request)

        courses = QuickCourse.objects.filter(is_active=True).select_related('academic_year').order_by('name')
        if course_type != 'ALL':
            courses = courses.filter(course_type=course_type)

        course_data, totals = _build_quick_outstanding_course_summary(
            courses,
            include_zero_outstanding=False,
            start_date=start_date,
            end_date=end_date,
        )
        course_data = sorted(course_data, key=lambda row: (-row['outstanding_students'], row['course'].name))
        _, detail_data = _build_quick_outstanding_rows(
            [row['course'] for row in course_data],
            start_date=start_date,
            end_date=end_date,
        )

        current_snapshot = _snapshot_outstanding_totals(totals)
        previous_snapshot = self.request.session.get('quick_outstanding_report_snapshot')
        previous_time = self.request.session.get('quick_outstanding_report_timestamp')
        comparison = _build_outstanding_comparison(current_snapshot, previous_snapshot)

        context.update({
            'courses': course_data,
            'totals': totals,
            'detail_groups': detail_data['grouped_courses'],
            'print_date': timezone.now().strftime('%Y-%m-%d %H:%M'),
            'comparison': comparison,
            'previous_report_time': previous_time,
            'course_type': course_type,
            'course_type_label': course_type_label,
            'course_type_report_label': report_label,
            'start_date': start_date,
            'end_date': end_date,
        })

        self.request.session['quick_outstanding_report_snapshot'] = current_snapshot
        self.request.session['quick_outstanding_report_timestamp'] = timezone.now().strftime('%Y-%m-%d %H:%M')
        return context
