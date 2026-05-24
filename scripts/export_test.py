import os
import sys

# Ensure project root is in PYTHONPATH
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.append(project_root)

# Set Django settings module (adjust if your settings module has a different name)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'alyaman.settings')

import django
django.setup()

from accounts.views import ChartOfAccountsView
from django.test import RequestFactory
from openpyxl import load_workbook

# Create request to trigger Excel export
request = RequestFactory().get('/accounts/chart-of-accounts/?export=excel')
from django.contrib.auth import get_user_model
User = get_user_model()
# Use the first active user or create an anonymous placeholder
request.user = User.objects.first() if User.objects.exists() else None
view = ChartOfAccountsView()
view.request = request
response = view.export_to_excel()

# Save the generated Excel file
excel_path = os.path.join(project_root, 'chart_of_accounts_test.xlsx')
with open(excel_path, 'wb') as f:
    f.write(response.content)
print('Excel file saved to:', excel_path)

# Load and preview first few rows (header row is 5)
wb = load_workbook(excel_path, data_only=True)
ws = wb.active
print('Sheet title:', ws.title)
for row in ws.iter_rows(min_row=5, max_row=10, values_only=True):
    print(row)
